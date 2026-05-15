import re
import json
import datetime
import urllib.request
import urllib.parse
from io import BytesIO
from dataclasses import dataclass
from typing import List, Tuple

import boto3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter


def lambda_handler(event, context):
    @dataclass(frozen=True)
    class Config:
        bucket: str = "bi-automations"
        forecast_key: str = "Forecast/forecast.csv"
        actuals_key: str = "Forecast/actuals.csv"
        input_key: str = "Forecast/Forecast_Input.xlsx"
        out_key: str = "Forecast/verify_actuals_vs_forecast.xlsx"

    CFG = Config()
    S3 = boto3.client("s3")

    # ── Style constants ───────────────────────────────────────────────
    WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
    GRAY_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
    THIN = Side(style="thin", color="000000")
    THICK = Side(style="thick", color="000000")
    HEADER_BOTTOM_BORDER = Border(bottom=THIN)
    BOLD_FONT = Font(bold=True)
    BOLD_FONT_LARGE = Font(bold=True, size=13)
    GRAY_FONT = Font(color="595959")
    GRAY_FONT_BOLD = Font(bold=True, color="595959")
    LEFT_ALIGN = Alignment(horizontal="left")
    RIGHT_ALIGN = Alignment(horizontal="right")
    INDENT_ALIGN = Alignment(horizontal="left", indent=1)
    PERCENT_FMT = "0.0%"
    INTEGER_FMT = "#,##0"

    _WEEK_RE = re.compile(r"^(?P<y>\d{4})-(?P<w>\d{2})$")

    # ── Helpers ───────────────────────────────────────────────────────

    def parse_week_key(s: str) -> Tuple[int, int]:
        if not isinstance(s, str):
            return (-1, -1)
        m = _WEEK_RE.match(s.strip())
        return (int(m.group("y")), int(m.group("w"))) if m else (-1, -1)

    def shift_week_year(week_key: str, new_year: int) -> str:
        if not isinstance(week_key, str) or "-" not in week_key:
            return f"{new_year}-00"
        _, ww = week_key.split("-", 1)
        return f"{new_year}-{ww.zfill(2)}"

    def safe_sheet_base(name: str) -> str:
        bad = r'[:\\/?*\[\]]'
        s = re.sub(bad, " ", str(name)).strip()
        s = re.sub(r"\s+", " ", s)
        return s[:31] if s else "Sheet"

    def dedupe_sheet_name(base: str, used: set) -> str:
        name = base
        i = 2
        while name in used:
            suffix = f" {i}"
            name = (base[:31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
            i += 1
        used.add(name)
        return name

    def to_num(s: pd.Series) -> pd.Series:
        return pd.to_numeric(s, errors="coerce").fillna(0).astype(float)

    def get_val(agg, key):
        try:
            return float(agg.loc[key])
        except (KeyError, TypeError):
            return 0.0

    def write_week_header_row(ws, all_weeks, row=1):
        ws.cell(row=row, column=1, value="").fill = WHITE_FILL
        for wi, wk in enumerate(all_weeks):
            c = wi + 2
            cell = ws.cell(row=row, column=c, value=wk)
            cell.font = BOLD_FONT
            cell.border = HEADER_BOTTOM_BORDER
            cell.alignment = RIGHT_ALIGN
            cell.fill = WHITE_FILL

    def write_section_header(ws, current_row, label, max_c):
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = BOLD_FONT_LARGE
        cell.fill = WHITE_FILL
        cell.alignment = LEFT_ALIGN
        for c in range(2, max_c + 1):
            ws.cell(row=current_row, column=c).fill = WHITE_FILL
        return current_row + 1

    def write_block(ws, current_row, label, all_weeks, actuals_vals,
                    forecast_vals, actuals_ly_vals, fyoy_vals, yoy_vals,
                    error_vals, error_pct_vals):
        max_c = len(all_weeks) + 1

        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = BOLD_FONT
        cell.fill = WHITE_FILL
        cell.border = HEADER_BOTTOM_BORDER
        for c in range(2, max_c + 1):
            ws.cell(row=current_row, column=c).fill = WHITE_FILL
            ws.cell(row=current_row, column=c).border = HEADER_BOTTOM_BORDER
        current_row += 1

        metrics = [
            ("Actuals",    actuals_vals,    INTEGER_FMT),
            ("Forecast",   forecast_vals,   INTEGER_FMT),
            ("Actuals LY", actuals_ly_vals, INTEGER_FMT),
            ("F-YoY %",    fyoy_vals,       PERCENT_FMT),
            ("YoY %",      yoy_vals,        PERCENT_FMT),
            ("Error",      error_vals,      INTEGER_FMT),
            ("Error %",    error_pct_vals,  PERCENT_FMT),
        ]

        for metric_label, vals, fmt in metrics:
            cell_label = ws.cell(row=current_row, column=1, value=metric_label)
            cell_label.font = BOLD_FONT
            cell_label.alignment = INDENT_ALIGN
            cell_label.fill = WHITE_FILL

            for wi, v in enumerate(vals):
                c = wi + 2
                cell = ws.cell(row=current_row, column=c, value=v)
                cell.number_format = fmt
                cell.alignment = RIGHT_ALIGN
                cell.fill = WHITE_FILL

            current_row += 1

        for c in range(1, max_c + 1):
            ws.cell(row=current_row, column=c).fill = WHITE_FILL
        current_row += 1

        return current_row

    def write_block_slim(ws, current_row, label, all_weeks, actuals_vals,
                         forecast_vals, actuals_ly_vals, fyoy_vals, yoy_vals,
                         error_vals, error_pct_vals):
        max_c = len(all_weeks) + 1

        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = BOLD_FONT
        cell.fill = WHITE_FILL
        cell.border = HEADER_BOTTOM_BORDER
        for c in range(2, max_c + 1):
            ws.cell(row=current_row, column=c).fill = WHITE_FILL
            ws.cell(row=current_row, column=c).border = HEADER_BOTTOM_BORDER
        current_row += 1

        metrics = [
            ("Actuals",  actuals_vals,   INTEGER_FMT),
            ("F-YoY %",  fyoy_vals,      PERCENT_FMT),
            ("YoY %",    yoy_vals,       PERCENT_FMT),
            ("Error %",  error_pct_vals, PERCENT_FMT),
        ]

        for metric_label, vals, fmt in metrics:
            cell_label = ws.cell(row=current_row, column=1, value=metric_label)
            cell_label.font = BOLD_FONT
            cell_label.alignment = INDENT_ALIGN
            cell_label.fill = WHITE_FILL

            for wi, v in enumerate(vals):
                c = wi + 2
                cell = ws.cell(row=current_row, column=c, value=v)
                cell.number_format = fmt
                cell.alignment = RIGHT_ALIGN
                cell.fill = WHITE_FILL

            current_row += 1

        for c in range(1, max_c + 1):
            ws.cell(row=current_row, column=c).fill = WHITE_FILL
        current_row += 1

        return current_row

    def apply_cutoff_border(ws, all_weeks, cutoff_week):
        if cutoff_week not in all_weeks:
            return
        col_idx = all_weeks.index(cutoff_week) + 2
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            b = cell.border or Border()
            cell.border = Border(left=b.left, right=THICK, top=b.top, bottom=b.bottom)

    def apply_pre_cutoff_gray(ws, all_weeks, cutoff_week):
        """Gray out all cells (fill + font) in columns before the cutoff week."""
        if cutoff_week not in all_weeks:
            return
        cutoff_col = all_weeks.index(cutoff_week) + 2  # col index of cutoff week
        # Gray columns: from col 1 (labels) through cutoff_col (inclusive)
        for r in range(1, ws.max_row + 1):
            for c in range(1, cutoff_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = GRAY_FILL
                # Preserve bold on label/header cells
                if cell.font and cell.font.bold:
                    cell.font = GRAY_FONT_BOLD
                else:
                    cell.font = GRAY_FONT
                # Preserve existing borders
                # (cutoff border applied separately after this)

    def apply_week_grouping(ws, all_weeks, cutoff_week):
        if cutoff_week not in all_weeks:
            return
        cutoff_idx = all_weeks.index(cutoff_week)
        group_end_idx = cutoff_idx - 6
        if group_end_idx <= 0:
            return
        for i in range(0, group_end_idx):
            col = i + 2
            col_dim = ws.column_dimensions[get_column_letter(col)]
            col_dim.outlineLevel = 1
            col_dim.hidden = True

    def auto_width(ws, n_weeks):
        max_col = n_weeks + 1
        for c in range(1, max_col + 1):
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                if isinstance(v, str):
                    max_len = max(max_len, len(v))
                elif isinstance(v, (int, float)):
                    nf = ws.cell(row=r, column=c).number_format
                    if nf == PERCENT_FMT:
                        max_len = max(max_len, 7)
                    else:
                        try:
                            max_len = max(max_len, len(f"{v:,.0f}"))
                        except (ValueError, TypeError):
                            max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(c)].width = max(10, min(18, max_len + 2))
        ws.column_dimensions["A"].width = max(18, ws.column_dimensions["A"].width)

    def compute_block_values(all_weeks, agg_act, agg_fc, agg_ly, key_prefix):
        _iso = datetime.date.today().isocalendar()
        current_week = f"{_iso[0]}-{_iso[1]:02d}"

        actuals_vals, forecast_vals, actuals_ly_vals = [], [], []
        fyoy_vals, yoy_vals, error_vals, error_pct_vals = [], [], [], []

        for wk in all_weeks:
            key = (*key_prefix, wk)
            a = get_val(agg_act, key)
            f = get_val(agg_fc, key)
            a_ly = get_val(agg_ly, key)

            if wk >= current_week:
                err = 0
                err_pct = 0.0
            else:
                err = a - f
                err_pct = (err / f) if f != 0 else 0.0
            fyoy = (f / a_ly - 1.0) if (f != 0 and a_ly != 0) else 0.0
            yoy = (a / a_ly - 1.0) if (a != 0 and a_ly != 0) else 0.0

            actuals_vals.append(a)
            forecast_vals.append(f)
            actuals_ly_vals.append(a_ly)
            fyoy_vals.append(fyoy)
            yoy_vals.append(yoy)
            error_vals.append(err)
            error_pct_vals.append(err_pct)

        return actuals_vals, forecast_vals, actuals_ly_vals, fyoy_vals, yoy_vals, error_vals, error_pct_vals

    def write_blank_rows(ws, current_row, n, max_c):
        for _ in range(n):
            for c in range(1, max_c + 1):
                ws.cell(row=current_row, column=c).fill = WHITE_FILL
            current_row += 1
        return current_row

    # ── Data loading ──────────────────────────────────────────────────

    response_url = event.get("response_url", "")

    try:
        # Read Forecast_Input.xlsx info sheet → cutoff week (C2)
        input_bytes = S3.get_object(Bucket=CFG.bucket, Key=CFG.input_key)["Body"].read()
        info = pd.read_excel(BytesIO(input_bytes), sheet_name="info", header=None)
        cutoff_week = str(info.iat[1, 2]).strip()
        print(f"[verify] cutoff_week: {cutoff_week}")

        # Read forecast.csv
        fc = pd.read_csv(S3.get_object(Bucket=CFG.bucket, Key=CFG.forecast_key)["Body"])
        fc["forecasted_shop"] = fc["forecasted_shop"].astype(str).str.strip()
        _blank = fc["forecasted_shop"].isin(["", "nan", "None"])
        fc.loc[_blank, "forecasted_shop"] = "Other " + fc.loc[_blank, "shoptype"].astype(str).str.strip()
        fc["FQTY"] = to_num(fc["FQTY"])
        fc["iso_week"] = fc["iso_week"].astype(str).str.strip()
        fc["forecast_product"] = fc["forecast_product"].astype(str).str.strip()
        fc["destination_region"] = fc["destination_region"].astype(str).str.strip()
        fc["shoptype"] = fc["shoptype"].astype(str).str.strip() if "shoptype" in fc.columns else ""
        print(f"[verify] forecast.csv: {len(fc)} rows")

        # Named forecast shops (excluding "Other *")
        fc_named_shops = set(
            s for s in fc["forecasted_shop"].unique()
            if not s.startswith("Other ")
        )
        # "Other [shoptype]" entries in forecast
        fc_other_shops = set(
            s for s in fc["forecasted_shop"].unique()
            if s.startswith("Other ")
        )
        print(f"[verify] forecast shops: {len(fc_named_shops)} named, {len(fc_other_shops)} Other")

        # Read actuals.csv
        act_raw = pd.read_csv(S3.get_object(Bucket=CFG.bucket, Key=CFG.actuals_key)["Body"])
        act_raw["fulldate"] = pd.to_datetime(act_raw["fulldate"], errors="coerce")
        act_raw = act_raw[act_raw["fulldate"].notna()].copy()
        act_raw["actuals"] = to_num(act_raw["actuals"])
        act_raw["week"] = act_raw["week"].astype(str).str.strip()
        act_raw["forecast_product"] = act_raw["forecast_product"].astype(str).str.strip()
        act_raw["destination_region"] = act_raw["destination_region"].astype(str).str.strip()
        act_raw["forecasted_shop"] = act_raw["forecasted_shop"].astype(str).str.strip()
        act_raw["shoptype"] = act_raw["shoptype"].astype(str).str.strip()

        t = int(act_raw["fulldate"].dt.year.max())
        print(f"[verify] year t = {t}")

        fc_weeks = set(fc.loc[fc["iso_week"].str.startswith(f"{t}-"), "iso_week"].unique())
        act_weeks = set(act_raw.loc[act_raw["week"].str.startswith(f"{t}-"), "week"].unique())
        all_weeks: List[str] = sorted(fc_weeks | act_weeks, key=parse_week_key)
        print(f"[verify] display weeks: {len(all_weeks)}")

        if not all_weeks:
            return {"ok": False, "error": "no weeks"}

        # ── Build shop-level aggregations (residual approach) ─────────
        #
        # For named shops: aggregate actuals where forecasted_shop matches
        # For "Other [shoptype]": compute as shoptype total minus named shops
        #
        # This matches FORECAST-review-region's approach.

        shop_grp = ["destination_region", "forecasted_shop", "forecast_product"]
        st_grp = ["destination_region", "shoptype", "forecast_product"]

        act_t = act_raw[act_raw["week"].str.startswith(f"{t}-")]
        act_ly = act_raw[act_raw["week"].str.startswith(f"{t - 1}-")]

        # 1) Named-shop actuals (shops that have their own forecast row)
        #    Exclude rows where forecasted_shop == shoptype AND "Other [shoptype]"
        #    exists in forecast — those actuals belong to the "Other" residual.
        #    Matches FORECAST-review-region line 218: s != shoptype.
        fc_other_shoptypes = set(s[len("Other "):] for s in fc_other_shops)
        is_named = act_t["forecasted_shop"].isin(fc_named_shops)
        is_shoptype_self = (
            (act_t["forecasted_shop"] == act_t["shoptype"]) &
            act_t["shoptype"].isin(fc_other_shoptypes)
        )
        act_t_named = act_t[is_named & ~is_shoptype_self]
        agg_named_t = (
            act_t_named.groupby(shop_grp + ["week"], dropna=False)["actuals"]
            .sum().astype(float)
        )

        is_named_ly = act_ly["forecasted_shop"].isin(fc_named_shops)
        is_shoptype_self_ly = (
            (act_ly["forecasted_shop"] == act_ly["shoptype"]) &
            act_ly["shoptype"].isin(fc_other_shoptypes)
        )
        act_ly_named = act_ly[is_named_ly & ~is_shoptype_self_ly]
        agg_named_ly = (
            act_ly_named.groupby(shop_grp + ["week"], dropna=False)["actuals"]
            .sum().astype(float)
        )

        # 2) Shoptype totals (all actuals under each shoptype)
        agg_st_t = (
            act_t.groupby(st_grp + ["week"], dropna=False)["actuals"]
            .sum().astype(float)
        )
        agg_st_ly = (
            act_ly.groupby(st_grp + ["week"], dropna=False)["actuals"]
            .sum().astype(float)
        )

        # 3) Sum of named shops per shoptype (to subtract from total)
        act_t_named_st = act_t_named.copy()
        act_t_named_st["_st"] = act_t_named_st["shoptype"]
        agg_named_by_st_t = (
            act_t_named_st.groupby(["destination_region", "_st", "forecast_product", "week"],
                                   dropna=False)["actuals"]
            .sum().astype(float)
        )
        act_ly_named_st = act_ly_named.copy()
        act_ly_named_st["_st"] = act_ly_named_st["shoptype"]
        agg_named_by_st_ly = (
            act_ly_named_st.groupby(["destination_region", "_st", "forecast_product", "week"],
                                    dropna=False)["actuals"]
            .sum().astype(float)
        )

        # 4) Compute "Other [shoptype]" residual rows and append to named aggs
        other_rows_t = []
        other_rows_ly = []
        for other_shop in fc_other_shops:
            st_name = other_shop[len("Other "):]  # e.g. "ORWO" from "Other ORWO"
            try:
                st_total = agg_st_t.xs(st_name, level="shoptype")
            except KeyError:
                continue
            for key, total_val in st_total.items():
                dest, prod, wk = key if isinstance(key, tuple) else (key,)
                named_val = 0.0
                try:
                    named_val = float(agg_named_by_st_t.loc[(dest, st_name, prod, wk)])
                except (KeyError, TypeError):
                    pass
                residual = total_val - named_val
                if residual != 0:
                    other_rows_t.append({
                        "destination_region": dest,
                        "forecasted_shop": other_shop,
                        "forecast_product": prod,
                        "week": wk,
                        "actuals": residual,
                    })

        for other_shop in fc_other_shops:
            st_name = other_shop[len("Other "):]
            try:
                st_total = agg_st_ly.xs(st_name, level="shoptype")
            except KeyError:
                continue
            for key, total_val in st_total.items():
                dest, prod, wk = key if isinstance(key, tuple) else (key,)
                named_val = 0.0
                try:
                    named_val = float(agg_named_by_st_ly.loc[(dest, st_name, prod, wk)])
                except (KeyError, TypeError):
                    pass
                residual = total_val - named_val
                if residual != 0:
                    other_rows_ly.append({
                        "destination_region": dest,
                        "forecasted_shop": other_shop,
                        "forecast_product": prod,
                        "week": wk,
                        "actuals": residual,
                    })

        # Combine named + Other into final actuals aggregations
        if other_rows_t:
            other_df_t = pd.DataFrame(other_rows_t)
            other_agg_t = other_df_t.set_index(shop_grp + ["week"])["actuals"]
            agg_actuals = pd.concat([agg_named_t, other_agg_t])
            agg_actuals = agg_actuals.groupby(level=agg_actuals.index.names).sum()
        else:
            agg_actuals = agg_named_t

        if other_rows_ly:
            other_df_ly = pd.DataFrame(other_rows_ly)
            other_agg_ly = other_df_ly.set_index(shop_grp + ["week"])["actuals"]
            agg_actuals_ly_combined = pd.concat([agg_named_ly, other_agg_ly])
            agg_actuals_ly_combined = agg_actuals_ly_combined.groupby(
                level=agg_actuals_ly_combined.index.names).sum()
        else:
            agg_actuals_ly_combined = agg_named_ly

        # Shift LY week index to year t for alignment
        if not agg_actuals_ly_combined.empty:
            idx = agg_actuals_ly_combined.index
            new_weeks = [shift_week_year(w, t) for w in idx.get_level_values("week")]
            agg_actuals_ly = agg_actuals_ly_combined.copy()
            agg_actuals_ly.index = pd.MultiIndex.from_arrays(
                [idx.get_level_values(c) for c in shop_grp] + [new_weeks],
                names=shop_grp + ["week"],
            )
        else:
            agg_actuals_ly = agg_actuals_ly_combined

        print(f"[verify] residual Other rows: t={len(other_rows_t)}, ly={len(other_rows_ly)}")

        # Forecast aggregation
        fc_t = fc[fc["iso_week"].str.startswith(f"{t}-")]
        agg_forecast = (
            fc_t.groupby(shop_grp + ["iso_week"], dropna=False)["FQTY"]
            .sum().astype(float)
        )
        agg_forecast.index = agg_forecast.index.rename({"iso_week": "week"})

        # ── Build product-level (summary) aggregations ────────────────

        prod_grp = ["destination_region", "forecast_product"]

        agg_act_prod = (
            act_t.groupby(prod_grp + ["week"], dropna=False)["actuals"]
            .sum().astype(float)
        )

        agg_ly_prod_raw = (
            act_ly.groupby(prod_grp + ["week"], dropna=False)["actuals"]
            .sum().astype(float)
        )
        if not agg_ly_prod_raw.empty:
            idx = agg_ly_prod_raw.index
            new_weeks = [shift_week_year(w, t) for w in idx.get_level_values("week")]
            agg_ly_prod = agg_ly_prod_raw.copy()
            agg_ly_prod.index = pd.MultiIndex.from_arrays(
                [idx.get_level_values(c) for c in prod_grp] + [new_weeks],
                names=prod_grp + ["week"],
            )
        else:
            agg_ly_prod = agg_ly_prod_raw

        agg_fc_prod = (
            fc_t.groupby(prod_grp + ["iso_week"], dropna=False)["FQTY"]
            .sum().astype(float)
        )
        agg_fc_prod.index = agg_fc_prod.index.rename({"iso_week": "week"})

        # ── Discover products and combos (forecast-only) ──────────────

        fc_combos = set(
            fc_t[shop_grp].drop_duplicates()
            .apply(lambda r: (r["destination_region"], r["forecasted_shop"], r["forecast_product"]), axis=1)
        )

        product_combos = {}
        for dest, shop, prod in fc_combos:
            product_combos.setdefault(prod, set()).add((dest, shop))

        product_dests = {}
        for prod, combos in product_combos.items():
            product_dests[prod] = sorted(set(d for d, s in combos))

        ytd_by_product = act_t.groupby("forecast_product")["actuals"].sum()
        products = sorted(
            product_combos.keys(),
            key=lambda p: ytd_by_product.get(p, 0),
            reverse=True,
        )
        print(f"[verify] products: {len(products)} (sorted by YTD)")

        # ── Build Excel workbook ──────────────────────────────────────

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        used_names = set()
        max_c = len(all_weeks) + 1

        # ══════════════════════════════════════════════════════════════
        #  SUMMARY SHEET — product x destination, split by region
        # ══════════════════════════════════════════════════════════════

        ws_sum = wb.create_sheet(dedupe_sheet_name("Summary", used_names))
        write_week_header_row(ws_sum, all_weeks, row=1)

        current_row = 2

        eu_prods = [(p, "EU+RoW") for p in products if "EU+RoW" in product_dests.get(p, [])]
        if eu_prods:
            current_row = write_section_header(ws_sum, current_row, "EU+RoW", max_c)
            for product, dest in eu_prods:
                vals = compute_block_values(
                    all_weeks, agg_act_prod, agg_fc_prod, agg_ly_prod,
                    key_prefix=(dest, product),
                )
                current_row = write_block(ws_sum, current_row, product, all_weeks, *vals)

        if eu_prods:
            current_row = write_blank_rows(ws_sum, current_row, 10, max_c)

        us_prods = [(p, "US+CA") for p in products if "US+CA" in product_dests.get(p, [])]
        if us_prods:
            current_row = write_section_header(ws_sum, current_row, "US+CA", max_c)
            for product, dest in us_prods:
                vals = compute_block_values(
                    all_weeks, agg_act_prod, agg_fc_prod, agg_ly_prod,
                    key_prefix=(dest, product),
                )
                current_row = write_block(ws_sum, current_row, product, all_weeks, *vals)

        auto_width(ws_sum, len(all_weeks))
        apply_pre_cutoff_gray(ws_sum, all_weeks, cutoff_week)
        apply_cutoff_border(ws_sum, all_weeks, cutoff_week)
        apply_week_grouping(ws_sum, all_weeks, cutoff_week)
        ws_sum.freeze_panes = "B2"

        # ══════════════════════════════════════════════════════════════
        #  SUMMARY 2 SHEET — Actuals + Error only
        # ══════════════════════════════════════════════════════════════

        ws_sum2 = wb.create_sheet(dedupe_sheet_name("Summary 2", used_names))
        write_week_header_row(ws_sum2, all_weeks, row=1)

        current_row = 2

        if eu_prods:
            current_row = write_section_header(ws_sum2, current_row, "EU+RoW", max_c)
            for product, dest in eu_prods:
                vals = compute_block_values(
                    all_weeks, agg_act_prod, agg_fc_prod, agg_ly_prod,
                    key_prefix=(dest, product),
                )
                current_row = write_block_slim(ws_sum2, current_row, product, all_weeks, *vals)

        if eu_prods:
            current_row = write_blank_rows(ws_sum2, current_row, 10, max_c)

        if us_prods:
            current_row = write_section_header(ws_sum2, current_row, "US+CA", max_c)
            for product, dest in us_prods:
                vals = compute_block_values(
                    all_weeks, agg_act_prod, agg_fc_prod, agg_ly_prod,
                    key_prefix=(dest, product),
                )
                current_row = write_block_slim(ws_sum2, current_row, product, all_weeks, *vals)

        auto_width(ws_sum2, len(all_weeks))
        apply_pre_cutoff_gray(ws_sum2, all_weeks, cutoff_week)
        apply_cutoff_border(ws_sum2, all_weeks, cutoff_week)
        apply_week_grouping(ws_sum2, all_weeks, cutoff_week)
        ws_sum2.freeze_panes = "B2"

        # ══════════════════════════════════════════════════════════════
        #  PER-PRODUCT SHEETS — shop breakdown, split by region
        # ══════════════════════════════════════════════════════════════

        for product in products:
            combos = product_combos.get(product, set())
            if not combos:
                continue

            sname = dedupe_sheet_name(safe_sheet_base(product), used_names)
            ws = wb.create_sheet(sname)
            write_week_header_row(ws, all_weeks, row=1)

            eu_combos = sorted((d, s) for d, s in combos if d == "EU+RoW")
            us_combos = sorted((d, s) for d, s in combos if d == "US+CA")

            current_row = 2

            if eu_combos:
                current_row = write_section_header(ws, current_row, "EU+RoW", max_c)
                for dest, shop in eu_combos:
                    vals = compute_block_values(
                        all_weeks, agg_actuals, agg_forecast, agg_actuals_ly,
                        key_prefix=(dest, shop, product),
                    )
                    current_row = write_block(ws, current_row, shop, all_weeks, *vals)

            if eu_combos and us_combos:
                current_row = write_blank_rows(ws, current_row, 10, max_c)

            if us_combos:
                current_row = write_section_header(ws, current_row, "US+CA", max_c)
                for dest, shop in us_combos:
                    vals = compute_block_values(
                        all_weeks, agg_actuals, agg_forecast, agg_actuals_ly,
                        key_prefix=(dest, shop, product),
                    )
                    current_row = write_block(ws, current_row, shop, all_weeks, *vals)

            auto_width(ws, len(all_weeks))
            apply_pre_cutoff_gray(ws, all_weeks, cutoff_week)
            apply_cutoff_border(ws, all_weeks, cutoff_week)
            apply_week_grouping(ws, all_weeks, cutoff_week)
            ws.freeze_panes = "B2"

        # ── Save to S3 ────────────────────────────────────────────────

        buf = BytesIO()
        wb.save(buf)
        S3.put_object(
            Bucket=CFG.bucket,
            Key=CFG.out_key,
            Body=buf.getvalue(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        print(f"[verify] Wrote {len(products)} product sheets + Summary to s3://{CFG.bucket}/{CFG.out_key}")

        # ── Slack: upload Excel + error alerts ────────────────────────
        file_bytes = buf.getvalue()

        _iso = datetime.date.today().isocalendar()
        current_week = f"{_iso[0]}-{_iso[1]:02d}"
        completed_weeks = [w for w in all_weeks if w < current_week][-2:]

        def get_threshold(rank):
            if rank <= 10:
                return 0.10
            if rank <= 40:
                return 0.15
            return 0.20

        # Compute product-level error per (dest, product) — sum across weeks
        prod_errors = {}
        for product in products:
            for dest, shop in product_combos.get(product, set()):
                total_a, total_f = 0.0, 0.0
                for wk in completed_weeks:
                    total_f += get_val(agg_forecast, (dest, shop, product, wk))
                    total_a += get_val(agg_actuals, (dest, shop, product, wk))
                if total_f == 0:
                    continue
                err = (total_a - total_f) / total_f
                if err == -1.0:
                    continue
                key = (dest, product)
                if key not in prod_errors:
                    prod_errors[key] = {"total_a": 0.0, "total_f": 0.0, "shops": []}
                prod_errors[key]["total_a"] += total_a
                prod_errors[key]["total_f"] += total_f
                if abs(err) > 0:
                    prod_errors[key]["shops"].append((shop, err))

        def build_region_alerts(region):
            items = []
            for (dest, product), data in prod_errors.items():
                if dest != region or data["total_f"] == 0:
                    continue
                err = (data["total_a"] - data["total_f"]) / data["total_f"]
                if err == -1.0:
                    continue
                items.append({"product": product, "err": err, "abs_err": abs(err), "shops": data["shops"]})
            items.sort(key=lambda x: x["abs_err"], reverse=True)
            alerts = []
            for rank, item in enumerate(items, 1):
                threshold = get_threshold(rank)
                if item["abs_err"] > threshold:
                    contrib = [s for s, e in item["shops"] if abs(e) > 0.10]
                    direction = "overforecasted" if item["err"] < 0 else "underforecasted"
                    tier = "TOP" if rank <= 10 else ("MID" if rank <= 40 else "LOW")
                    alerts.append({"product": item["product"], "abs_err": item["abs_err"],
                                   "direction": direction, "threshold": threshold,
                                   "tier": tier, "contrib": contrib})
            return alerts

        def format_region(region, alerts):
            if not alerts:
                return None
            lines = [f"*{region} — Forecast Error Alert ({', '.join(completed_weeks)})*"]
            current_tier = None
            tier_labels = {"TOP": "10%", "MID": "15%", "LOW": "20%"}
            n = 0
            for a in alerts:
                if a["tier"] != current_tier:
                    current_tier = a["tier"]
                    lines.append(f"\n*{current_tier} Products ({tier_labels[current_tier]} error):*")
                n += 1
                shops = ", ".join(a["contrib"]) if a["contrib"] else "all channels"
                lines.append(f"{n}. *{a['product']}* — {a['abs_err']:.1%} {a['direction']} — channels: {shops}")
            return "\n".join(lines)

        eu_alerts = build_region_alerts("EU+RoW")
        us_alerts = build_region_alerts("US+CA")
        eu_text = format_region("EU+RoW", eu_alerts)
        us_text = format_region("US+CA", us_alerts)

        # Combine into one message
        parts = []
        if eu_text:
            parts.append(eu_text)
        if us_text:
            parts.append(us_text)
        alert_comment = "\n\n".join(parts) if parts else "No threshold breaches detected."

        try:
            ssm = boto3.client("ssm")
            slack_token = ssm.get_parameter(
                Name="/forecast/slack-bot-token", WithDecryption=True
            )["Parameter"]["Value"]
            channel = ssm.get_parameter(
                Name="/forecast/slack-channel-id"
            )["Parameter"]["Value"]

            def slack_api(method, fields):
                body = urllib.parse.urlencode(fields).encode()
                req = urllib.request.Request(
                    f"https://slack.com/api/{method}",
                    data=body,
                    headers={"Authorization": f"Bearer {slack_token}",
                             "Content-Type": "application/x-www-form-urlencoded"},
                    method="POST",
                )
                resp = json.loads(urllib.request.urlopen(req, timeout=15).read())
                if not resp.get("ok"):
                    raise RuntimeError(f"Slack {method}: {resp.get('error', resp)}")
                return resp

            # Upload Excel with alert text as initial_comment
            init = slack_api("files.getUploadURLExternal", {
                "filename": "verify_actuals_vs_forecast.xlsx",
                "length": len(file_bytes),
            })
            upload_req = urllib.request.Request(
                init["upload_url"], data=file_bytes,
                headers={"Content-Type": "application/octet-stream"},
                method="POST",
            )
            urllib.request.urlopen(upload_req, timeout=30)
            slack_api("files.completeUploadExternal", {
                "files": json.dumps([{"id": init["file_id"], "title": "verify_actuals_vs_forecast.xlsx"}]),
                "channel_id": channel,
                "initial_comment": alert_comment,
            })
            total = len(eu_alerts) + len(us_alerts)
            print(f"[verify] uploaded Excel + {total} alerts to Slack")
        except Exception as slack_err:
            print(f"[verify] Slack upload failed: {slack_err}")

        return {
            "ok": True,
            "s3_key": CFG.out_key,
            "products": len(products),
            "weeks": len(all_weeks),
            "cutoff_week": cutoff_week,
            "other_residual_rows_t": len(other_rows_t),
            "alerts": len(eu_alerts) + len(us_alerts),
        }

    except Exception as e:
        print(f"[verify] ERROR: {e}")
        import traceback; traceback.print_exc()
        return {"ok": False, "error": str(e)}
