import os
import json
import re
import urllib.request
import urllib.parse
import urllib.error
from dataclasses import dataclass
from io import BytesIO
from typing import List, Tuple, Optional, Dict

import boto3
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter


def lambda_handler(event, context):
    @dataclass(frozen=True)
    class Config:
        bucket: str = "bi-automations"
        in_key: str = "Forecast/actuals.csv"
        out_key_prefix: str = "Forecast/"
        forecast_combined_key: str = "Forecast/Seperate Forecasts/seperate_forecasts_combined.csv"

        date_col: str = "fulldate"
        actuals_col: str = "actuals"
        forecast_col: str = "forecast"

        shoptype_col: str = "shoptype"
        product_col: str = "forecast_product"
        label_col: str = "TOP Products"
        week_col: str = "week"
        destination_col: str = "destination_region"
        shop_col: str = "forecasted_shop"

        excluded_products: Tuple[str, ...] = ("EXCLUDED PRODUCT", "NEW PRODUCT")
        top_n: int = 15

    CFG = Config()
    S3 = boto3.client("s3")
    _WEEK_RE = re.compile(r"^(?P<y>\d{4})-(?P<w>\d{2})$")

    # ── Styles ────────────────────────────────────────────────────────
    WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
    GRAY_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
    THIN = Side(style="thin", color="000000")
    HEADER_BOTTOM_BORDER = Border(bottom=THIN)
    ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BOLD_FONT = Font(bold=True)
    LEFT_ALIGN = Alignment(horizontal="left")
    CENTER_ALIGN = Alignment(horizontal="center")
    INDENT_ALIGN = Alignment(horizontal="left", indent=1)
    PERCENT_FMT = "0.0%"
    NUMBER_FMT = "#,##0"

    RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
    RED_FONT = Font(color="9C0006")
    GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
    GREEN_FONT = Font(color="006100")

    YEAR_T_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")

    DEST_USCA = "US+CA"
    DEST_EUROW = "EU+RoW"

    # ── Request parsing ───────────────────────────────────────────────
    def parse_region_request(evt: dict) -> str:
        raw = (evt.get("text") or evt.get("region") or "").strip()
        low = raw.lower()
        if low.startswith("region "):
            raw = raw[7:].strip()
        normed = raw.replace(" ", "").upper()
        if normed in ("US+CA", "CA+US", "USCA", "CAUS"):
            return DEST_USCA
        if normed in ("EU+ROW", "EUROW", "EU"):
            return DEST_EUROW
        return raw.strip()

    # ── Numeric / date helpers ────────────────────────────────────────
    def to_num(s: pd.Series) -> pd.Series:
        return pd.to_numeric(s, errors="coerce").fillna(0).astype(float)

    def safe_growth(curr: pd.Series, prev: pd.Series) -> pd.Series:
        curr, prev = to_num(curr), to_num(prev)
        out = np.where(prev != 0, (curr / prev) - 1.0, 0.0)
        out = np.where(np.isfinite(out), out, 0.0)
        return pd.Series(out, index=curr.index)

    def normalize_ratio(s: pd.Series) -> pd.Series:
        s = to_num(s)
        mask = s.abs() > 5
        s.loc[mask] = s.loc[mask] / 100.0
        return s

    def parse_week_key(s: str) -> Tuple[int, int]:
        if not isinstance(s, str):
            return (-1, -1)
        m = _WEEK_RE.match(s.strip())
        return (int(m.group("y")), int(m.group("w"))) if m else (-1, -1)

    def shift_week_year(week_key: str, new_year: int) -> str:
        if not isinstance(week_key, str) or "-" not in week_key:
            return f"{new_year}-00"
        _, ww = week_key.split("-", 1)
        return f"{new_year}-{ww.zfill(2)}"

    def ensure_col(df: pd.DataFrame, col: str) -> pd.Series:
        return to_num(df[col]) if col in df.columns else pd.Series(0.0, index=df.index)

    def destination_group(v) -> str:
        s = str(v or "").strip().upper()
        return DEST_USCA if s in ("US", "CA", "US+CA", "CA+US") else DEST_EUROW

    def shift_month_label_ly(m: str) -> str:
        try:
            return f"{int(m[:4]) - 1}-{m[5:7]}"
        except Exception:
            return m

    # ── Week helpers ──────────────────────────────────────────────────
    def last_3_weeks_keys_any(df: pd.DataFrame, t: int, week_col: str) -> List[str]:
        d = df[df[week_col].astype(str).str.startswith(f"{t}-")]
        keys = sorted(d[week_col].dropna().astype(str).unique().tolist(), key=parse_week_key)
        if not keys:
            return []
        keys = keys[:-1]
        return keys[-3:] if len(keys) >= 3 else keys

    def get_display_weeks(df: pd.DataFrame, week_col: str, n: int = 15) -> List[str]:
        all_wks = sorted(df[week_col].dropna().astype(str).unique().tolist(), key=parse_week_key)
        if not all_wks:
            return []
        all_wks = all_wks[:-1]
        return all_wks[-n:] if len(all_wks) >= n else all_wks

    # ── Month helpers ─────────────────────────────────────────────────
    def last_3_full_months_labels(max_dt_t: pd.Timestamp) -> List[str]:
        month_start = pd.Timestamp(year=max_dt_t.year, month=max_dt_t.month, day=1)
        last_full_month_end = month_start - pd.Timedelta(days=1)
        last_full_month = last_full_month_end.to_period("M")
        months = pd.period_range(end=last_full_month, periods=3, freq="M")
        return [str(p) for p in months]

    def get_display_months(max_dt: pd.Timestamp) -> List[str]:
        current_period = max_dt.to_period("M")
        return [str(current_period + i) for i in range(-2, 6)]

    # ── Data cleaning ─────────────────────────────────────────────────
    def clean_input(df: pd.DataFrame, cfg: Config) -> pd.DataFrame:
        df = df.copy()
        df = df[~df[cfg.product_col].astype(str).isin(cfg.excluded_products)]
        df[cfg.date_col] = pd.to_datetime(df[cfg.date_col], errors="coerce")
        df = df[df[cfg.date_col].notna()].copy()
        df[cfg.actuals_col] = to_num(df[cfg.actuals_col])
        df["_month"] = df[cfg.date_col].dt.to_period("M").astype(str)

        df[cfg.shoptype_col] = df[cfg.shoptype_col].astype(str).str.strip()
        df[cfg.product_col] = df[cfg.product_col].astype(str).str.strip()
        df[cfg.week_col] = df[cfg.week_col].astype(str).str.strip()
        df[cfg.shop_col] = df[cfg.shop_col].astype(str).str.strip()

        if cfg.destination_col in df.columns:
            df["_dest_group"] = df[cfg.destination_col].apply(destination_group)
        else:
            df["_dest_group"] = DEST_EUROW
        return df

    # ── Aggregation helpers ───────────────────────────────────────────
    def agg_sum(df: pd.DataFrame, cfg: Config, mask: pd.Series, label: str) -> pd.DataFrame:
        return (
            df[mask]
            .groupby(cfg.product_col, as_index=False)[cfg.actuals_col]
            .sum()
            .rename(columns={cfg.product_col: cfg.label_col, cfg.actuals_col: label})
        )

    def get_agg_value(agg: pd.Series, key: tuple) -> float:
        try:
            return float(agg.loc[key])
        except Exception:
            return 0.0

    # ── Forecast CSV → shops by shoptype ──────────────────────────────
    def read_forecast_shops(cfg: Config, dest_group: str) -> Dict[str, Dict[str, List[str]]]:
        try:
            obj = S3.get_object(Bucket=cfg.bucket, Key=cfg.forecast_combined_key)
            fc = pd.read_csv(obj["Body"])
        except Exception as e:
            print(f"[read_forecast_shops] Could not read forecast CSV: {e}")
            return {}

        for col in ("destination", "shoptype", "forecasted_shop", "product"):
            if col not in fc.columns:
                print(f"[read_forecast_shops] Missing column: {col}")
                return {}

        fc["_dest_group"] = fc["destination"].apply(destination_group)
        fc = fc[fc["_dest_group"] == dest_group]

        result: Dict[str, Dict[str, List[str]]] = {}
        for (product, shoptype), grp in fc.groupby(["product", "shoptype"]):
            product = str(product).strip()
            shoptype = str(shoptype).strip()
            if not shoptype:
                continue
            shops_raw = [
                s for s in grp["forecasted_shop"].astype(str).str.strip().unique()
                if s and not s.startswith("Other ")
            ]
            # Only treat as shoptype-total (no breakdown) if the sole entry
            # is the shoptype name itself. If other shops exist alongside it,
            # it's a real named shop (e.g. ORWO shop under ORWO shoptype).
            shops_raw = [
                s for s in grp["forecasted_shop"].astype(str).str.strip().unique()
                if s 
                and not s.startswith("Other ")
                and s != shoptype  # ← add this: always exclude the shoptype-named entry
            ]
            shops = sorted(set(shops_raw)) if shops_raw else []

            if shops:
                result.setdefault(product, {})[shoptype] = shops

        return result

    # ── Forecast CSV → MONTH aggs (period-based, matches combined builder) ──
    def read_forecast_month_aggs(cfg: Config, dest_group: str):
        try:
            obj = S3.get_object(Bucket=cfg.bucket, Key=cfg.forecast_combined_key)
            fc = pd.read_csv(obj["Body"])
        except Exception as e:
            print(f"[read_forecast_month_aggs] Could not read forecast CSV: {e}")
            return None, None, None

        for col in ("destination", "shoptype", "forecasted_shop", "product", "period", cfg.forecast_col):
            if col not in fc.columns:
                print(f"[read_forecast_month_aggs] Missing column: {col}")
                return None, None, None

        fc["period"] = fc["period"].astype(str).str.strip()
        fc[cfg.forecast_col] = to_num(fc[cfg.forecast_col])

        fc["_dest_group"] = fc["destination"].apply(destination_group)
        fc = fc[fc["_dest_group"] == dest_group].copy()

        fc["_month"] = fc["period"]
        fc["_product"] = fc["product"].astype(str).str.strip()
        fc["_shoptype"] = fc["shoptype"].astype(str).str.strip()
        fc["_shop"] = fc["forecasted_shop"].astype(str).str.strip()

        fc_m_total = (
            fc.groupby(["_product", "_month", "_dest_group"], dropna=False)[cfg.forecast_col]
            .sum()
            .astype(float)
        )
        fc_m_st = (
            fc.groupby(["_shoptype", "_product", "_month", "_dest_group"], dropna=False)[cfg.forecast_col]
            .sum()
            .astype(float)
        )
        fc_m_shop = (
            fc.groupby(["_shop", "_product", "_month", "_dest_group"], dropna=False)[cfg.forecast_col]
            .sum()
            .astype(float)
        )
        return fc_m_total, fc_m_st, fc_m_shop

    # ── Summary computation ───────────────────────────────────────────
    def compute_region_review(df_raw: pd.DataFrame, cfg: Config, dest_group: str):
        df = clean_input(df_raw, cfg)
        mask = df["_dest_group"] == dest_group
        df_scope = df[mask]
        if df_scope.empty:
            raise ValueError(f"No data for region {dest_group}")

        t = int(df_scope[cfg.date_col].dt.year.max())
        max_dt_t = df_scope.loc[df_scope[cfg.date_col].dt.year == t, cfg.date_col].max()
        cutoff_t1 = max_dt_t - pd.DateOffset(years=1)

        base_t = mask & (df[cfg.date_col].dt.year == t) & (df[cfg.date_col] <= max_dt_t)
        base_t1_ytd = mask & (df[cfg.date_col].dt.year == t - 1) & (df[cfg.date_col] <= cutoff_t1)
        base_t1_fy = mask & (df[cfg.date_col].dt.year == t - 1)
        base_t2_fy = mask & (df[cfg.date_col].dt.year == t - 2)

        out = (
            agg_sum(df, cfg, base_t, f"YTD {t}")
            .merge(agg_sum(df, cfg, base_t1_ytd, f"YTD {t-1}"), on=cfg.label_col, how="outer")
            .merge(agg_sum(df, cfg, base_t1_fy, f"FY {t-1}"), on=cfg.label_col, how="outer")
            .merge(agg_sum(df, cfg, base_t2_fy, f"FY {t-2}"), on=cfg.label_col, how="outer")
            .fillna(0)
        )

        last3w_t = last_3_weeks_keys_any(df, t, cfg.week_col)
        last3w_t1 = [shift_week_year(k, t - 1) for k in last3w_t]
        out = (
            out.merge(
                agg_sum(df, cfg, mask & df[cfg.week_col].isin(last3w_t), f"Last 3W {t}"),
                on=cfg.label_col,
                how="outer",
            )
            .merge(
                agg_sum(df, cfg, mask & df[cfg.week_col].isin(last3w_t1), f"Last 3W {t-1}"),
                on=cfg.label_col,
                how="outer",
            )
            .fillna(0)
        )

        last3m_t = last_3_full_months_labels(max_dt_t)
        last3m_t1 = [shift_month_label_ly(m) for m in last3m_t]
        for m_t, m_t1 in zip(last3m_t, last3m_t1):
            msum_t = agg_sum(df, cfg, mask & (df["_month"] == m_t), "_mt")
            msum_t1 = agg_sum(df, cfg, mask & (df["_month"] == m_t1), "_mt1")
            tmp = msum_t.merge(msum_t1, on=cfg.label_col, how="outer").fillna(0)
            tmp[m_t] = safe_growth(tmp["_mt"], tmp["_mt1"])
            out = out.merge(tmp[[cfg.label_col, m_t]], on=cfg.label_col, how="outer").fillna(0)

        out["YTD % growth"] = safe_growth(ensure_col(out, f"YTD {t}"), ensure_col(out, f"YTD {t-1}"))
        out["FY % growth"] = safe_growth(ensure_col(out, f"FY {t-1}"), ensure_col(out, f"FY {t-2}"))
        out["Last 3w % growth"] = safe_growth(
            ensure_col(out, f"Last 3W {t}"), ensure_col(out, f"Last 3W {t-1}")
        )
        out["% of Total YTD"] = 0.0

        out = out.sort_values(f"YTD {t}", ascending=False).reset_index(drop=True)

        out["_rank"] = np.arange(1, len(out) + 1)
        out["_fy"] = ensure_col(out, f"FY {t-1}")
        keep = (out["_rank"] <= cfg.top_n) & (out["_fy"] >= 1000.0)

        top_rows = out[keep].copy()
        other_rows = out[~keep].copy()

        helper = {cfg.label_col, "_rank", "_fy"}
        numeric_cols = [c for c in out.columns if c not in helper and pd.api.types.is_numeric_dtype(out[c])]

        total_ytd = float(to_num(out[f"YTD {t}"]).sum())

        total_vals = out[numeric_cols].sum(numeric_only=True).to_dict()
        total_label = f"TOTAL {dest_group}"
        total_row = pd.DataFrame([{cfg.label_col: total_label, **total_vals}])

        for m_t, m_t1 in zip(last3m_t, last3m_t1):
            tot_mt = float(df[mask & (df["_month"] == m_t)][cfg.actuals_col].sum())
            tot_mt1 = float(df[mask & (df["_month"] == m_t1)][cfg.actuals_col].sum())
            total_row[m_t] = (tot_mt / tot_mt1 - 1.0) if tot_mt1 != 0 else 0.0

        total_row["YTD % growth"] = safe_growth(ensure_col(total_row, f"YTD {t}"), ensure_col(total_row, f"YTD {t-1}"))
        total_row["FY % growth"] = safe_growth(ensure_col(total_row, f"FY {t-1}"), ensure_col(total_row, f"FY {t-2}"))
        total_row["Last 3w % growth"] = safe_growth(
            ensure_col(total_row, f"Last 3W {t}"), ensure_col(total_row, f"Last 3W {t-1}")
        )
        total_row["% of Total YTD"] = 1.0

        final_cols = [
            cfg.label_col,
            "% of Total YTD",
            f"YTD {t}",
            f"YTD {t-1}",
            "YTD % growth",
            f"FY {t-1}",
            f"FY {t-2}",
            "FY % growth",
            "Last 3w % growth",
            *last3m_t,
        ]

        def finalize(frame: pd.DataFrame) -> pd.DataFrame:
            frame = frame.drop(columns=["_rank", "_fy"], errors="ignore")
            if total_ytd != 0:
                is_total = frame[cfg.label_col].astype(str).str.startswith("TOTAL")
                frame.loc[~is_total, "% of Total YTD"] = to_num(frame.loc[~is_total, f"YTD {t}"]) / total_ytd
            frame["YTD % growth"] = safe_growth(ensure_col(frame, f"YTD {t}"), ensure_col(frame, f"YTD {t-1}"))
            frame["FY % growth"] = safe_growth(ensure_col(frame, f"FY {t-1}"), ensure_col(frame, f"FY {t-2}"))
            frame["Last 3w % growth"] = safe_growth(
                ensure_col(frame, f"Last 3W {t}"), ensure_col(frame, f"Last 3W {t-1}")
            )
            percent_cols = ["% of Total YTD", "YTD % growth", "FY % growth", "Last 3w % growth", *last3m_t]
            for c in percent_cols:
                if c in frame.columns:
                    frame[c] = normalize_ratio(frame[c])
            for c in final_cols:
                if c not in frame.columns:
                    frame[c] = 0.0
            return frame[final_cols].fillna(0).reset_index(drop=True)

        top_df = finalize(top_rows.sort_values(f"YTD {t}", ascending=False))
        other_df = finalize(other_rows.sort_values(f"YTD {t}", ascending=False))
        total_df = finalize(total_row)

        if not other_rows.empty:
            other_agg_vals = other_rows[numeric_cols].sum(numeric_only=True).to_dict()
            other_agg_vals[cfg.label_col] = "Other products"
            other_agg_row = finalize(pd.DataFrame([other_agg_vals]))
        else:
            other_agg_row = pd.DataFrame(columns=final_cols)

        if not other_rows.empty:
            other_total_vals = other_rows[numeric_cols].sum(numeric_only=True).to_dict()
            other_total_vals[cfg.label_col] = f"TOTAL Other {dest_group}"
            other_total_frame = pd.DataFrame([other_total_vals])
            other_product_names = set(other_rows[cfg.label_col].astype(str).tolist())
            other_product_mask = df[cfg.product_col].astype(str).isin(other_product_names)
            for m_t, m_t1 in zip(last3m_t, last3m_t1):
                ot_mt = float(df[mask & (df["_month"] == m_t) & other_product_mask][cfg.actuals_col].sum())
                ot_mt1 = float(df[mask & (df["_month"] == m_t1) & other_product_mask][cfg.actuals_col].sum())
                other_total_frame[m_t] = (ot_mt / ot_mt1 - 1.0) if ot_mt1 != 0 else 0.0
            other_total_df = finalize(other_total_frame)
            other_ytd_sum = float(to_num(other_rows[f"YTD {t}"]).sum())
            if total_ytd != 0:
                other_total_df.loc[:, "% of Total YTD"] = other_ytd_sum / total_ytd
            else:
                other_total_df.loc[:, "% of Total YTD"] = 0.0
        else:
            other_total_df = pd.DataFrame(columns=final_cols)

        return t, top_df, other_df, total_df, other_agg_row, other_total_df

    # ── Sheet-name helpers ────────────────────────────────────────────
    def safe_sheet_base(name: str) -> str:
        bad = r'[:\\/?*\[\]]'
        s = re.sub(bad, " ", str(name)).strip()
        s = re.sub(r"\s+", " ", s)
        return s[:31] if s else "Sheet"

    def dedupe_sheet_name(base: str, used: set) -> str:
        name = base
        i = 2
        while name in used:
            suffix = f" {i}"
            name = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
            i += 1
        used.add(name)
        return name

    # ── Summary-sheet formatting ──────────────────────────────────────
    def is_total_value(v) -> bool:
        return isinstance(v, str) and v.startswith("TOTAL")

    def write_df_to_ws(ws, df: pd.DataFrame, start_row: int) -> int:
        for j, col in enumerate(df.columns, 1):
            ws.cell(row=start_row, column=j, value=col)
        for i in range(len(df)):
            for j, col in enumerate(df.columns, 1):
                ws.cell(row=start_row + 1 + i, column=j, value=df.iloc[i][col])
        return start_row + len(df)

    def apply_formatting_summary_range(ws, cfg: Config, start_row: int, end_row: int):
        if end_row < start_row:
            return
        max_col = ws.max_column

        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.fill = WHITE_FILL
                cell.border = Border()

        headers = {}
        for c in range(1, max_col + 1):
            hcell = ws.cell(row=start_row, column=c)
            hcell.font = BOLD_FONT
            hcell.border = HEADER_BOTTOM_BORDER
            headers[hcell.value] = c

        rb_names = {"YTD % growth", "FY % growth", "Last 3w % growth"}
        rb_cols = [headers[n] for n in rb_names if n in headers]
        for r in range(start_row, end_row + 1):
            for c in rb_cols:
                cell = ws.cell(row=r, column=c)
                b = cell.border or Border()
                cell.border = Border(left=b.left, right=THIN, top=b.top, bottom=b.bottom)

        label_idx = headers.get(cfg.label_col)
        if label_idx:
            for r in range(start_row + 1, end_row + 1):
                v = ws.cell(row=r, column=label_idx).value
                ws.cell(row=r, column=label_idx).alignment = LEFT_ALIGN if is_total_value(v) else INDENT_ALIGN

        percent_headers = {"% of Total YTD", "YTD % growth", "FY % growth", "Last 3w % growth"}
        percent_cols = []
        for h, c in headers.items():
            if h is None:
                continue
            hs = str(h).strip()
            if hs in percent_headers or hs.count("-") == 1:
                percent_cols.append(c)
                for r in range(start_row + 1, end_row + 1):
                    ws.cell(row=r, column=c).number_format = PERCENT_FMT

        for r in range(start_row + 1, end_row + 1):
            if is_total_value(ws.cell(row=r, column=1).value):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = BOLD_FONT
                    b = cell.border or Border()
                    cell.border = Border(left=b.left, right=b.right, top=THIN, bottom=b.bottom)

        for c in range(1, max_col + 1):
            is_pct = c in percent_cols
            max_len = 0
            for r in range(start_row, end_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    s = ""
                elif r == start_row:
                    s = str(v)
                elif is_pct:
                    try:
                        s = f"{float(v)*100:.1f}%"
                    except Exception:
                        s = "0.0%"
                else:
                    s = str(v)
                max_len = max(max_len, len(s))
            ws.column_dimensions[get_column_letter(c)].width = max(8, min(60, max_len + 2))

        ytd_growth_col = headers.get("YTD % growth")
        if ytd_growth_col:
            for r in range(start_row + 1, end_row + 1):
                label_v = ws.cell(row=r, column=1).value
                if is_total_value(label_v):
                    continue
                if isinstance(label_v, str) and label_v.startswith("Other products"):
                    continue
                cell = ws.cell(row=r, column=ytd_growth_col)
                v = cell.value
                if isinstance(v, (int, float, np.integer, np.floating)):
                    fv = float(v)
                    if fv < -0.10:
                        cell.fill = RED_FILL
                        cell.font = RED_FONT
                    elif fv > 0.10:
                        cell.fill = GREEN_FILL
                        cell.font = GREEN_FONT

    # ── Product-sheet builder ─────────────────────────────────────────
    def build_product_sheet(
        ws,
        cfg: Config,
        product: str,
        dest_group: str,
        t: int,
        display_weeks: List[str],
        w_agg_st: pd.Series,
        w_agg_total: pd.Series,
        w_agg_shop: pd.Series,
        display_months: List[str],
        m_agg_st: pd.Series,
        m_agg_total: pd.Series,
        m_agg_shop: pd.Series,
        product_shops_by_st: Dict[str, List[str]],
        fc_m_total: Optional[pd.Series],
        fc_m_st: Optional[pd.Series],
        fc_m_shop: Optional[pd.Series],
        current_month_label: str,
    ):
        prod_key = str(product).strip()

        # ── Resolve shoptypes ─────────────────────────────────────────
        try:
            sub = w_agg_st.xs(prod_key, level=1)
            sub_region = sub.xs(dest_group, level=-1)
            shoptypes = sorted(sub_region.index.get_level_values(0).unique().tolist())
        except Exception:
            shoptypes = []

        if not shoptypes and not display_weeks:
            ws.cell(row=3, column=2, value="No data for this product + region.")
            return

        # ── Value accessors ───────────────────────────────────────────
        def get_st_week(shoptype, wk):
            return get_agg_value(w_agg_st, (shoptype, prod_key, wk, dest_group))

        def get_total_week(wk):
            return get_agg_value(w_agg_total, (prod_key, wk, dest_group))

        def get_shop_week(shop, wk):
            return get_agg_value(w_agg_shop, (shop, prod_key, wk, dest_group))

        def _is_month_ge_current(mo: str) -> bool:
            try:
                return (mo[:4] == str(t)) and (mo >= current_month_label)
            except Exception:
                return False

        def get_total_month_actual(mo):
            return get_agg_value(m_agg_total, (prod_key, mo, dest_group))

        def get_st_month_actual(shoptype, mo):
            return get_agg_value(m_agg_st, (shoptype, prod_key, mo, dest_group))

        def get_shop_month_actual(shop, mo):
            return get_agg_value(m_agg_shop, (shop, prod_key, mo, dest_group))

        def get_total_month(mo):
            if _is_month_ge_current(mo) and fc_m_total is not None:
                return get_agg_value(fc_m_total, (prod_key, mo, dest_group))
            return get_total_month_actual(mo)

        def get_st_month(shoptype, mo):
            if _is_month_ge_current(mo) and fc_m_st is not None:
                return get_agg_value(fc_m_st, (str(shoptype).strip(), prod_key, mo, dest_group))
            return get_st_month_actual(shoptype, mo)

        def get_shop_month(shop, mo):
            if _is_month_ge_current(mo) and fc_m_shop is not None:
                return get_agg_value(fc_m_shop, (str(shop).strip(), prod_key, mo, dest_group))
            return get_shop_month_actual(shop, mo)

        def get_total_month_ly(mo_ly):
            return get_total_month_actual(mo_ly)

        def get_st_month_ly(shoptype, mo_ly):
            return get_st_month_actual(shoptype, mo_ly)

        def get_shop_month_ly(shop, mo_ly):
            return get_shop_month_actual(shop, mo_ly)

        # ── Helper: check if Other block is all zeros across weeks+months ──
        def other_block_is_zero(st, st_shops_list):
            for mo in display_months:
                lbl_ly = shift_month_label_ly(mo)
                if _is_month_ge_current(mo) and fc_m_shop is not None:
                    other_t = get_shop_month(f"Other {st}", mo)
                else:
                    other_t = get_st_month(st, mo) - sum(get_shop_month(s, mo) for s in st_shops_list)
                other_t1 = get_st_month_ly(st, lbl_ly) - sum(get_shop_month_ly(s, lbl_ly) for s in st_shops_list)
                if other_t != 0 or other_t1 != 0:
                    return False
            for wk in display_weeks:
                wk_ly = shift_week_year(wk, parse_week_key(wk)[0] - 1)
                other_t = get_st_week(st, wk) - sum(get_shop_week(s, wk) for s in st_shops_list)
                other_t1 = get_st_week(st, wk_ly) - sum(get_shop_week(s, wk_ly) for s in st_shops_list)
                if other_t != 0 or other_t1 != 0:
                    return False
            return True

        # ══════════════════════════════════════════════════════════════
        #  UNIFIED COLUMN LAYOUT
        # ══════════════════════════════════════════════════════════════
        WEEK_COL = 2
        TOTAL_START = 3  # C
        GREY_SEP_1 = 6  # F
        GREY_SEP_2 = 7  # G
        ST_START = 8  # H

        all_blank_cols = {GREY_SEP_1, GREY_SEP_2}

        st_layout = []
        col = ST_START
        for st_idx, st in enumerate(shoptypes):
            st_shops = product_shops_by_st.get(st, [])
            st_col = col
            col += 3  # shoptype block

            shop_cols = []
            other_col = None
            group_first = col
            group_last = col - 1  # default: no group

            if st_shops:
                for shop in st_shops:
                    shop_cols.append((shop, col))
                    col += 3
                st_shops_list = [s for s, _ in shop_cols]
                # Only add Other column if it has non-zero data
                if not other_block_is_zero(st, st_shops_list):
                    other_col = col
                    col += 3
                group_last = col - 1

            st_layout.append((st, st_col, shop_cols, other_col, group_first, group_last))

            if st_idx < len(shoptypes) - 1:
                all_blank_cols.add(col)
                col += 1

        LAST_DATA_COL = col - 1 if col > ST_START else TOTAL_START + 2
        GRAY_COL = LAST_DATA_COL + 1

        # ── Row layout ────────────────────────────────────────────────
        LABEL_ROW = 2
        HEADER_ROW = 3
        DATA_START = 4

        n_weeks = len(display_weeks)
        LAST_WEEK_ROW = DATA_START + n_weeks - 1
        WEEK_TOTAL_ROW = LAST_WEEK_ROW + 1

        BLANK_ROW_1 = WEEK_TOTAL_ROW + 1
        BLANK_ROW_2 = WEEK_TOTAL_ROW + 2

        MONTH_LABEL_ROW = BLANK_ROW_2 + 1
        MONTH_HEADER_ROW = MONTH_LABEL_ROW + 1
        MONTH_DATA_START = MONTH_HEADER_ROW + 1

        n_months = len(display_months)
        LAST_MONTH_ROW = MONTH_DATA_START + n_months - 1
        MONTH_TOTAL_ROW = LAST_MONTH_ROW + 1

        SHEET_BOTTOM = MONTH_TOTAL_ROW + 1 + 200

        week_block_ranges = [(TOTAL_START, TOTAL_START + 2)]
        month_block_ranges = [(TOTAL_START, TOTAL_START + 2)]
        for st, st_col, shop_cols, other_col, gf, gl in st_layout:
            week_block_ranges.append((st_col, st_col + 2))
            month_block_ranges.append((st_col, st_col + 2))
            if shop_cols:
                for _, sc in shop_cols:
                    week_block_ranges.append((sc, sc + 2))
                    month_block_ranges.append((sc, sc + 2))
                if other_col is not None:
                    week_block_ranges.append((other_col, other_col + 2))
                    month_block_ranges.append((other_col, other_col + 2))

        # ── Helpers ───────────────────────────────────────────────────
        def write_block_header(label, col_start, label_row, header_row):
            ws.cell(row=label_row, column=col_start, value=label)
            ws.cell(row=label_row, column=col_start).font = BOLD_FONT
            ws.cell(row=label_row, column=col_start).alignment = CENTER_ALIGN
            ws.merge_cells(
                start_row=label_row,
                start_column=col_start,
                end_row=label_row,
                end_column=col_start + 2,
            )
            for dc, hdr in enumerate([str(t), str(t - 1), "YoY %"]):
                c = col_start + dc
                ws.cell(row=header_row, column=c, value=hdr)
                ws.cell(row=header_row, column=c).font = BOLD_FONT
                ws.cell(row=header_row, column=c).border = HEADER_BOTTOM_BORDER
                ws.cell(row=header_row, column=c).alignment = CENTER_ALIGN

        def write_block_data(get_val_t, get_val_t1, col_start, labels, data_start_row):
            col_t_letter = get_column_letter(col_start)
            col_t1_letter = get_column_letter(col_start + 1)
            for i, lbl in enumerate(labels):
                r = data_start_row + i
                val_t = get_val_t(lbl)
                val_t1 = get_val_t1(lbl)
                ws.cell(row=r, column=col_start, value=val_t).number_format = NUMBER_FMT
                ws.cell(row=r, column=col_start + 1, value=val_t1).number_format = NUMBER_FMT
                yoy_formula = f"=IF({col_t1_letter}{r}=0,0,{col_t_letter}{r}/{col_t1_letter}{r}-1)"
                ws.cell(row=r, column=col_start + 2, value=yoy_formula).number_format = PERCENT_FMT

        def write_other_block(st, st_shops_list, col_start, labels, data_start_row):
            col_t_letter = get_column_letter(col_start)
            col_t1_letter = get_column_letter(col_start + 1)
            for i, lbl in enumerate(labels):
                r = data_start_row + i
                lbl_ly = shift_month_label_ly(lbl)
                # Always subtract named shops from shoptype total.
                # get_st_month / get_shop_month already handle actuals vs forecast internally.
                other_t = get_st_month(st, lbl) - sum(get_shop_month(s, lbl) for s in st_shops_list)
                other_t1 = get_st_month_ly(st, lbl_ly) - sum(get_shop_month_ly(s, lbl_ly) for s in st_shops_list)
                ws.cell(row=r, column=col_start, value=other_t).number_format = NUMBER_FMT
                ws.cell(row=r, column=col_start + 1, value=other_t1).number_format = NUMBER_FMT
                yoy_formula = f"=IF({col_t1_letter}{r}=0,0,{col_t_letter}{r}/{col_t1_letter}{r}-1)"
                ws.cell(row=r, column=col_start + 2, value=yoy_formula).number_format = PERCENT_FMT

        def write_other_block_weekly(st, st_shops_list, col_start, labels, data_start_row):
            col_t_letter = get_column_letter(col_start)
            col_t1_letter = get_column_letter(col_start + 1)
            for i, wk in enumerate(labels):
                r = data_start_row + i
                wk_ly = shift_week_year(wk, parse_week_key(wk)[0] - 1)
                other_t = get_st_week(st, wk) - sum(get_shop_week(s, wk) for s in st_shops_list)
                other_t1 = get_st_week(st, wk_ly) - sum(get_shop_week(s, wk_ly) for s in st_shops_list)
                ws.cell(row=r, column=col_start, value=other_t).number_format = NUMBER_FMT
                ws.cell(row=r, column=col_start + 1, value=other_t1).number_format = NUMBER_FMT
                yoy_formula = f"=IF({col_t1_letter}{r}=0,0,{col_t_letter}{r}/{col_t1_letter}{r}-1)"
                ws.cell(row=r, column=col_start + 2, value=yoy_formula).number_format = PERCENT_FMT

        # ── Shoptype TOTAL block as SUM formulas ──────────────────────
        def write_month_shoptype_total_formulas(st_col, shop_cols_list, other_col, data_start_row, count):
            st_t_letter = get_column_letter(st_col)
            st_t1_letter = get_column_letter(st_col + 1)
            for i in range(count):
                r = data_start_row + i
                refs_t = [get_column_letter(sc) + str(r) for _, sc in shop_cols_list]
                refs_t1 = [get_column_letter(sc + 1) + str(r) for _, sc in shop_cols_list]
                if other_col is not None:
                    refs_t.append(get_column_letter(other_col) + str(r))
                    refs_t1.append(get_column_letter(other_col + 1) + str(r))
                ws.cell(row=r, column=st_col, value="=" + "+".join(refs_t)).number_format = NUMBER_FMT
                ws.cell(row=r, column=st_col + 1, value="=" + "+".join(refs_t1)).number_format = NUMBER_FMT
                yoy_formula = f"=IF({st_t1_letter}{r}=0,0,{st_t_letter}{r}/{st_t1_letter}{r}-1)"
                ws.cell(row=r, column=st_col + 2, value=yoy_formula).number_format = PERCENT_FMT

        # ── Grand TOTAL block as SUM of shoptype TOTALs ───────────────
        def write_month_grand_total_formulas(total_col, shoptype_col_list, data_start_row, count):
            tot_t_letter = get_column_letter(total_col)
            tot_t1_letter = get_column_letter(total_col + 1)
            for i in range(count):
                r = data_start_row + i
                refs_t = [get_column_letter(sc) + str(r) for sc in shoptype_col_list]
                ws.cell(row=r, column=total_col, value="=" + "+".join(refs_t)).number_format = NUMBER_FMT
                refs_t1 = [get_column_letter(sc + 1) + str(r) for sc in shoptype_col_list]
                ws.cell(row=r, column=total_col + 1, value="=" + "+".join(refs_t1)).number_format = NUMBER_FMT
                yoy_formula = f"=IF({tot_t1_letter}{r}=0,0,{tot_t_letter}{r}/{tot_t1_letter}{r}-1)"
                ws.cell(row=r, column=total_col + 2, value=yoy_formula).number_format = PERCENT_FMT

        # ── Summary TOTAL row as SUM formulas ─────────────────────────
        def write_summary_row_formulas(summary_row, data_start_row, last_data_row, block_ranges):
            for col_start, _col_end in block_ranges:
                col_t_letter = get_column_letter(col_start)
                col_t1_letter = get_column_letter(col_start + 1)
                f_t = f"=SUM({col_t_letter}{data_start_row}:{col_t_letter}{last_data_row})"
                cell_t = ws.cell(row=summary_row, column=col_start, value=f_t)
                cell_t.number_format = NUMBER_FMT
                cell_t.font = BOLD_FONT
                f_t1 = f"=SUM({col_t1_letter}{data_start_row}:{col_t1_letter}{last_data_row})"
                cell_t1 = ws.cell(row=summary_row, column=col_start + 1, value=f_t1)
                cell_t1.number_format = NUMBER_FMT
                cell_t1.font = BOLD_FONT
                yoy_formula = f"=IF({col_t1_letter}{summary_row}=0,0,{col_t_letter}{summary_row}/{col_t1_letter}{summary_row}-1)"
                cell_yoy = ws.cell(row=summary_row, column=col_start + 2, value=yoy_formula)
                cell_yoy.number_format = PERCENT_FMT
                cell_yoy.font = BOLD_FONT

        # ── Original write_summary_row (static, for week table) ───────
        def write_summary_row(summary_row, data_start_row, last_data_row, block_ranges):
            for col_start, _col_end in block_ranges:
                col_t_letter = get_column_letter(col_start)
                col_t1_letter = get_column_letter(col_start + 1)
                f_t = f"=SUM({col_t_letter}{data_start_row}:{col_t_letter}{last_data_row})"
                cell_t = ws.cell(row=summary_row, column=col_start, value=f_t)
                cell_t.number_format = NUMBER_FMT
                cell_t.font = BOLD_FONT
                f_t1 = f"=SUM({col_t1_letter}{data_start_row}:{col_t1_letter}{last_data_row})"
                cell_t1 = ws.cell(row=summary_row, column=col_start + 1, value=f_t1)
                cell_t1.number_format = NUMBER_FMT
                cell_t1.font = BOLD_FONT
                yoy_formula = f"=IF({col_t1_letter}{summary_row}=0,0,{col_t_letter}{summary_row}/{col_t1_letter}{summary_row}-1)"
                cell_yoy = ws.cell(row=summary_row, column=col_start + 2, value=yoy_formula)
                cell_yoy.number_format = PERCENT_FMT
                cell_yoy.font = BOLD_FONT

        # ==============================================================
        #  WEEK TABLE
        # ==============================================================
        ws.cell(row=HEADER_ROW, column=WEEK_COL, value="Week:")
        ws.cell(row=HEADER_ROW, column=WEEK_COL).font = BOLD_FONT
        ws.cell(row=HEADER_ROW, column=WEEK_COL).border = HEADER_BOTTOM_BORDER

        for i, wk in enumerate(display_weeks):
            ws.cell(row=DATA_START + i, column=WEEK_COL, value=wk)
            ws.cell(row=DATA_START + i, column=WEEK_COL).alignment = LEFT_ALIGN

        for st, st_col, shop_cols, other_col, gf, gl in st_layout:
            if shop_cols:
                st_shops_list = [s for s, _ in shop_cols]

                for shop, sc in shop_cols:
                    write_block_header(shop, sc, LABEL_ROW, HEADER_ROW)
                    write_block_data(
                        get_val_t=lambda wk, _s=shop: get_shop_week(_s, wk),
                        get_val_t1=lambda wk, _s=shop: get_shop_week(_s, shift_week_year(wk, parse_week_key(wk)[0] - 1)),
                        col_start=sc,
                        labels=display_weeks,
                        data_start_row=DATA_START,
                    )

                if other_col is not None:
                    write_block_header(f"Other {st}", other_col, LABEL_ROW, HEADER_ROW)
                    write_other_block_weekly(st, st_shops_list, other_col, display_weeks, DATA_START)

                write_block_header(f"TOTAL {st}", st_col, LABEL_ROW, HEADER_ROW)
                write_month_shoptype_total_formulas(st_col, shop_cols, other_col, DATA_START, n_weeks)
            else:
                write_block_header(f"TOTAL {st}", st_col, LABEL_ROW, HEADER_ROW)
                write_block_data(
                    get_val_t=lambda wk, _st=st: get_st_week(_st, wk),
                    get_val_t1=lambda wk, _st=st: get_st_week(_st, shift_week_year(wk, parse_week_key(wk)[0] - 1)),
                    col_start=st_col,
                    labels=display_weeks,
                    data_start_row=DATA_START,
                )

        shoptype_col_list_w = [st_col for _, st_col, _, _, _, _ in st_layout]
        write_block_header("TOTAL", TOTAL_START, LABEL_ROW, HEADER_ROW)
        if shoptype_col_list_w:
            write_month_grand_total_formulas(TOTAL_START, shoptype_col_list_w, DATA_START, n_weeks)
        else:
            write_block_data(
                get_val_t=lambda wk: get_total_week(wk),
                get_val_t1=lambda wk: get_total_week(shift_week_year(wk, parse_week_key(wk)[0] - 1)),
                col_start=TOTAL_START,
                labels=display_weeks,
                data_start_row=DATA_START,
            )

        ws.cell(row=WEEK_TOTAL_ROW, column=WEEK_COL, value="TOTAL")
        ws.cell(row=WEEK_TOTAL_ROW, column=WEEK_COL).font = BOLD_FONT
        write_summary_row_formulas(WEEK_TOTAL_ROW, DATA_START, LAST_WEEK_ROW, week_block_ranges)

        # ==============================================================
        #  MONTH TABLE
        # ==============================================================
        ws.cell(row=MONTH_HEADER_ROW, column=WEEK_COL, value="Month:")
        ws.cell(row=MONTH_HEADER_ROW, column=WEEK_COL).font = BOLD_FONT
        ws.cell(row=MONTH_HEADER_ROW, column=WEEK_COL).border = HEADER_BOTTOM_BORDER

        for i, mo in enumerate(display_months):
            ws.cell(row=MONTH_DATA_START + i, column=WEEK_COL, value=mo)
            ws.cell(row=MONTH_DATA_START + i, column=WEEK_COL).alignment = LEFT_ALIGN

        for st, st_col, shop_cols, other_col, gf, gl in st_layout:
            if shop_cols:
                st_shops_list = [s for s, _ in shop_cols]

                for shop, sc in shop_cols:
                    write_block_header(shop, sc, MONTH_LABEL_ROW, MONTH_HEADER_ROW)
                    write_block_data(
                        get_val_t=lambda mo, _s=shop: get_shop_month(_s, mo),
                        get_val_t1=lambda mo, _s=shop: get_shop_month_ly(_s, shift_month_label_ly(mo)),
                        col_start=sc,
                        labels=display_months,
                        data_start_row=MONTH_DATA_START,
                    )

                if other_col is not None:
                    write_block_header(f"Other {st}", other_col, MONTH_LABEL_ROW, MONTH_HEADER_ROW)
                    write_other_block(st, st_shops_list, other_col, display_months, MONTH_DATA_START)

                write_block_header(f"TOTAL {st}", st_col, MONTH_LABEL_ROW, MONTH_HEADER_ROW)
                write_month_shoptype_total_formulas(st_col, shop_cols, other_col, MONTH_DATA_START, n_months)
            else:
                write_block_header(f"TOTAL {st}", st_col, MONTH_LABEL_ROW, MONTH_HEADER_ROW)
                write_block_data(
                    get_val_t=lambda mo, _st=st: get_st_month(_st, mo),
                    get_val_t1=lambda mo, _st=st: get_st_month_ly(_st, shift_month_label_ly(mo)),
                    col_start=st_col,
                    labels=display_months,
                    data_start_row=MONTH_DATA_START,
                )

        shoptype_col_list = [st_col for _, st_col, _, _, _, _ in st_layout]
        write_block_header("TOTAL", TOTAL_START, MONTH_LABEL_ROW, MONTH_HEADER_ROW)
        if shoptype_col_list:
            write_month_grand_total_formulas(TOTAL_START, shoptype_col_list, MONTH_DATA_START, n_months)
        else:
            write_block_data(
                get_val_t=lambda mo: get_total_month(mo),
                get_val_t1=lambda mo: get_total_month_ly(shift_month_label_ly(mo)),
                col_start=TOTAL_START,
                labels=display_months,
                data_start_row=MONTH_DATA_START,
            )

        ws.cell(row=MONTH_TOTAL_ROW, column=WEEK_COL, value="TOTAL")
        ws.cell(row=MONTH_TOTAL_ROW, column=WEEK_COL).font = BOLD_FONT
        write_summary_row_formulas(MONTH_TOTAL_ROW, MONTH_DATA_START, LAST_MONTH_ROW, month_block_ranges)

        # ==============================================================
        #  COLUMN GROUPING
        # ==============================================================
        for st, st_col, shop_cols, other_col, gf, gl in st_layout:
            if shop_cols and gf <= gl:
                for c in range(gf, gl + 1):
                    col_dim = ws.column_dimensions[get_column_letter(c)]
                    col_dim.outlineLevel = 1
                    col_dim.hidden = True

        # ==============================================================
        #  FORMATTING (both tables)
        # ==============================================================
        for row in ws.iter_rows(min_row=1, max_row=SHEET_BOTTOM, min_col=1, max_col=GRAY_COL):
            for cell in row:
                cell.fill = WHITE_FILL

        def grey_empty_range(r_start, r_end, c_start, c_end):
            for r in range(r_start, r_end + 1):
                for c in range(c_start, c_end + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                        cell.fill = GRAY_FILL

        grey_empty_range(1, WEEK_TOTAL_ROW, 1, GRAY_COL)
        grey_empty_range(MONTH_LABEL_ROW, MONTH_TOTAL_ROW, 1, GRAY_COL)

        for r in (BLANK_ROW_1, BLANK_ROW_2):
            for c in range(1, GRAY_COL + 1):
                ws.cell(row=r, column=c).fill = GRAY_FILL

        for bc in all_blank_cols:
            for r in range(1, SHEET_BOTTOM + 1):
                ws.cell(row=r, column=bc).fill = GRAY_FILL

        for c in range(1, GRAY_COL + 1):
            ws.cell(row=MONTH_TOTAL_ROW + 1, column=c).fill = GRAY_FILL

        for r in range(1, SHEET_BOTTOM + 1):
            ws.cell(row=r, column=GRAY_COL).fill = GRAY_FILL

        for r in range(MONTH_TOTAL_ROW + 2, SHEET_BOTTOM + 1):
            for c in range(1, GRAY_COL + 1):
                ws.cell(row=r, column=c).fill = GRAY_FILL

        for i in range(n_weeks):
            ws.cell(row=DATA_START + i, column=WEEK_COL).font = BOLD_FONT
        for i in range(n_months):
            ws.cell(row=MONTH_DATA_START + i, column=WEEK_COL).font = BOLD_FONT

        for hdr_r in (HEADER_ROW, MONTH_HEADER_ROW):
            ws.cell(row=hdr_r, column=WEEK_COL).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        def border_label_column(start_r, count):
            for i in range(count):
                r = start_r + i
                cell = ws.cell(row=r, column=WEEK_COL)
                top = THIN if i == 0 else None
                bot = THIN if i == count - 1 else None
                cell.border = Border(left=THIN, right=THIN, top=top, bottom=bot)

        border_label_column(DATA_START, n_weeks)
        border_label_column(MONTH_DATA_START, n_months)

        def apply_block_border(block_start_col, block_end_col, top_r, bot_r):
            for r in range(top_r, bot_r + 1):
                for c in range(block_start_col, block_end_col + 1):
                    cell = ws.cell(row=r, column=c)
                    existing = cell.border or Border()
                    left = THIN if c == block_start_col else existing.left
                    right = THIN if c == block_end_col else existing.right
                    top = THIN if r == top_r else existing.top
                    bottom = THIN if r == bot_r else existing.bottom
                    cell.border = Border(left=left, right=right, top=top, bottom=bottom)

        for cs, ce in week_block_ranges:
            apply_block_border(cs, ce, LABEL_ROW, WEEK_TOTAL_ROW)
        for cs, ce in month_block_ranges:
            apply_block_border(cs, ce, MONTH_LABEL_ROW, MONTH_TOTAL_ROW)

        for total_r, blk_ranges in [(WEEK_TOTAL_ROW, week_block_ranges), (MONTH_TOTAL_ROW, month_block_ranges)]:
            ws.cell(row=total_r, column=WEEK_COL).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            for cs, ce in blk_ranges:
                for c in range(cs, ce + 1):
                    cell = ws.cell(row=total_r, column=c)
                    b = cell.border or Border()
                    cell.border = Border(left=b.left, right=b.right, top=THIN, bottom=b.bottom)

        for i, mo in enumerate(display_months):
            if mo < current_month_label:
                continue
            r = MONTH_DATA_START + i
            for cs, _ce in month_block_ranges:
                cell = ws.cell(row=r, column=cs)
                cell.fill = YEAR_T_FILL
                cell.border = ALL_BORDER

        from openpyxl.formatting.rule import CellIsRule

        week_yoy_cols = [cs + 2 for cs, _ in week_block_ranges]
        month_yoy_cols = [cs + 2 for cs, _ in month_block_ranges]

        def apply_conditional_yoy_excel(yoy_cols, data_start_r, last_data_r):
            for c in yoy_cols:
                col_letter = get_column_letter(c)
                cell_range = f"{col_letter}{data_start_r}:{col_letter}{last_data_r}"
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="lessThan",
                        formula=["-0.1"],
                        fill=RED_FILL,
                        font=RED_FONT,
                    ),
                )
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="greaterThan",
                        formula=["0.1"],
                        fill=GREEN_FILL,
                        font=GREEN_FONT,
                    ),
                )

        apply_conditional_yoy_excel(week_yoy_cols, DATA_START, LAST_WEEK_ROW)
        apply_conditional_yoy_excel(month_yoy_cols, MONTH_DATA_START, LAST_MONTH_ROW)

        def display_len(val, fmt=None):
            if val is None:
                return 0
            if isinstance(val, str):
                if val.startswith("="):
                    if fmt == PERCENT_FMT:
                        return 7
                    return 10
                return len(val)
            if fmt == PERCENT_FMT and isinstance(val, (int, float, np.integer, np.floating)):
                try:
                    return len(f"{float(val)*100:.1f}%")
                except Exception:
                    return len(str(val))
            if isinstance(val, (int, float, np.integer, np.floating)):
                try:
                    return len(f"{float(val):,.0f}")
                except Exception:
                    return len(str(val))
            return len(str(val))

        WIDTH_ROW = 32
        ws.column_dimensions[get_column_letter(1)].width = 3
        for bc in all_blank_cols:
            ws.column_dimensions[get_column_letter(bc)].width = 3
        ws.column_dimensions[get_column_letter(GRAY_COL)].width = 3

        for c in range(WEEK_COL, LAST_DATA_COL + 1):
            if c in all_blank_cols:
                continue
            cell = ws.cell(row=WIDTH_ROW, column=c)
            max_len = display_len(cell.value, cell.number_format)
            ws.column_dimensions[get_column_letter(c)].width = max(10, min(18, max_len + 2))

        ws.sheet_view.zoomScale = 70
        ws.freeze_panes = "G1"

    # ── Excel builder ─────────────────────────────────────────────────
    def build_excel_bytes(df_raw: pd.DataFrame, cfg: Config, dest_group: str) -> bytes:
        df_clean = clean_input(df_raw, cfg)

        w_agg_st = (
            df_clean.groupby([cfg.shoptype_col, cfg.product_col, cfg.week_col, "_dest_group"], dropna=False)[
                cfg.actuals_col
            ]
            .sum()
            .astype(float)
        )
        w_agg_total = (
            df_clean.groupby([cfg.product_col, cfg.week_col, "_dest_group"], dropna=False)[cfg.actuals_col]
            .sum()
            .astype(float)
        )
        m_agg_st = (
            df_clean.groupby([cfg.shoptype_col, cfg.product_col, "_month", "_dest_group"], dropna=False)[
                cfg.actuals_col
            ]
            .sum()
            .astype(float)
        )
        m_agg_total = (
            df_clean.groupby([cfg.product_col, "_month", "_dest_group"], dropna=False)[cfg.actuals_col]
            .sum()
            .astype(float)
        )
        m_agg_shop = (
            df_clean.groupby([cfg.shop_col, cfg.product_col, "_month", "_dest_group"], dropna=False)[cfg.actuals_col]
            .sum()
            .astype(float)
        )
        w_agg_shop = (
            df_clean.groupby([cfg.shop_col, cfg.product_col, cfg.week_col, "_dest_group"], dropna=False)[
                cfg.actuals_col
            ]
            .sum()
            .astype(float)
        )

        df_region = df_clean[df_clean["_dest_group"] == dest_group]
        display_weeks = get_display_weeks(df_region, cfg.week_col, n=15)

        max_dt_region = df_region[cfg.date_col].max()
        display_months = get_display_months(max_dt_region)
        current_month_label = str(max_dt_region.to_period("M"))

        fc_m_total, fc_m_st, fc_m_shop = read_forecast_month_aggs(cfg, dest_group)

        all_shops_by_product = read_forecast_shops(cfg, dest_group)

        t, top_df, other_df, total_df, other_agg_row, other_total_df = compute_region_review(df_raw, cfg, dest_group)
        top_products = top_df[cfg.label_col].astype(str).tolist()

        buf = BytesIO()
        used = set()

        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            wb = writer.book

            summary_name = dedupe_sheet_name(safe_sheet_base(dest_group), used)
            ws_sum = wb.create_sheet(summary_name)
            writer.sheets[summary_name] = ws_sum
            if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
                try:
                    wb.remove(wb["Sheet"])
                except Exception:
                    pass

            r = 1
            end_top = write_df_to_ws(ws_sum, top_df, start_row=r)

            if not other_agg_row.empty:
                end_top += 1
                for j, col in enumerate(other_agg_row.columns, 1):
                    ws_sum.cell(row=end_top, column=j, value=other_agg_row.iloc[0][col])

            end_top += 1
            for j, col in enumerate(total_df.columns, 1):
                ws_sum.cell(row=end_top, column=j, value=total_df.iloc[0][col])

            apply_formatting_summary_range(ws_sum, cfg, start_row=r, end_row=end_top)

            blank_start = end_top + 1
            for br in range(blank_start, blank_start + 15):
                for bc in range(1, ws_sum.max_column + 1):
                    ws_sum.cell(row=br, column=bc).fill = WHITE_FILL

            other_start = blank_start + 15
            if not other_df.empty:
                end_other = write_df_to_ws(ws_sum, other_df, start_row=other_start)
                end_other += 1
                for j, col in enumerate(other_total_df.columns, 1):
                    ws_sum.cell(row=end_other, column=j, value=other_total_df.iloc[0][col])
                apply_formatting_summary_range(ws_sum, cfg, start_row=other_start, end_row=end_other)

            for prod in top_products:
                sname = dedupe_sheet_name(safe_sheet_base(prod), used)
                pws = wb.create_sheet(sname)
                writer.sheets[sname] = pws

                build_product_sheet(
                    ws=pws,
                    cfg=cfg,
                    product=prod,
                    dest_group=dest_group,
                    t=t,
                    display_weeks=display_weeks,
                    w_agg_st=w_agg_st,
                    w_agg_total=w_agg_total,
                    w_agg_shop=w_agg_shop,
                    display_months=display_months,
                    m_agg_st=m_agg_st,
                    m_agg_total=m_agg_total,
                    m_agg_shop=m_agg_shop,
                    product_shops_by_st=all_shops_by_product.get(prod, {}),
                    fc_m_total=fc_m_total,
                    fc_m_st=fc_m_st,
                    fc_m_shop=fc_m_shop,
                    current_month_label=current_month_label,
                )

        return buf.getvalue()

    # ── S3 I/O ────────────────────────────────────────────────────────
    def read_csv_from_s3(cfg: Config) -> pd.DataFrame:
        obj = S3.get_object(Bucket=cfg.bucket, Key=cfg.in_key)
        return pd.read_csv(obj["Body"])

    # ── Slack helpers ─────────────────────────────────────────────────
    def slack_api_post_form(method: str, token: str, fields: dict, timeout: int = 20) -> dict:
        body = urllib.parse.urlencode(fields).encode("utf-8")
        req = urllib.request.Request(
            f"https://slack.com/api/{method}",
            data=body,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/x-www-form-urlencoded; charset=utf-8",
            },
            method="POST",
        )
        try:
            resp = urllib.request.urlopen(req, timeout=timeout).read()
        except urllib.error.HTTPError as e:
            err_body = e.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Slack API HTTPError {e.code} for {method}: {err_body}") from e
        data = json.loads(resp.decode("utf-8"))
        if not data.get("ok"):
            raise RuntimeError(f"Slack API {method} failed: {data}")
        return data

    def slack_upload_xlsx(file_bytes: bytes, filename: str, message: str):
        token = os.environ["SLACK_BOT_TOKEN"]
        channels_env = os.environ["SLACK_CHANNEL_ID"]
        channel_ids = [c.strip() for c in channels_env.split(",") if c.strip()]
        if not channel_ids:
            raise ValueError("SLACK_CHANNEL_ID is empty")

        init = slack_api_post_form(
            "files.getUploadURLExternal",
            token,
            {"filename": filename, "length": str(len(file_bytes))},
        )
        upload_url, file_id = init["upload_url"], init["file_id"]

        upload_req = urllib.request.Request(
            upload_url,
            data=file_bytes,
            headers={"Content-Type": "application/octet-stream"},
            method="POST",
        )
        try:
            resp = urllib.request.urlopen(upload_req, timeout=30)
            status = getattr(resp, "status", None) or resp.getcode()
            if int(status) != 200:
                raise RuntimeError(f"Upload failed: HTTP {status}")
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"Upload HTTPError {e.code}: {e.read().decode()}") from e

        files_payload = json.dumps([{"id": file_id, "title": filename}])
        complete = {"files": files_payload, "initial_comment": message}
        if len(channel_ids) == 1:
            complete["channel_id"] = channel_ids[0]
        else:
            complete["channels"] = ",".join(channel_ids)
        slack_api_post_form("files.completeUploadExternal", token, complete)

    # ── Entrypoint ────────────────────────────────────────────────────
    region = parse_region_request(event)
    if region not in (DEST_USCA, DEST_EUROW):
        raise ValueError(f"Invalid region '{region}'. Use 'EU+RoW' or 'US+CA'.")

    df = read_csv_from_s3(CFG)
    excel_bytes = build_excel_bytes(df, CFG, region)

    slack_upload_xlsx(
        excel_bytes,
        filename=f"review_region_{region}.xlsx",
        message=f"Finished reviewing region {region}",
    )

    return {"ok": True, "mode": "region", "region": region}