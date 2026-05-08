import re
import csv
import json
import urllib.request
from io import BytesIO, StringIO
from datetime import datetime, timezone

import boto3
from openpyxl import load_workbook

S3 = boto3.client("s3")

BUCKET = "bi-automations"
PREFIX = "Forecast/Seperate Forecasts/"
INPUT_XLSX_KEY = "Forecast/Forecast_Input.xlsx"
COMBINED_KEY = f"{PREFIX}seperate_forecasts_combined.csv"
BLUE_RGB = "DCE6F1"
PERIOD_RE = re.compile(r"^\d{4}-\d{2}$")

SHOP_FILE_RE = re.compile(r"^submit_shop_(.+?)(?:\s*\(\d+\))*\.xlsx$", re.I)
SHOPTYPE_FILE_RE = re.compile(r"^submit_shoptype_(.+?)(?:\s*\(\d+\))*\.xlsx$", re.I)
REGION_FILE_RE = re.compile(r"^submit_region_(.+?)(?:\s*\(\d+\))*\.xlsx$", re.I)
ANY_SUBMIT_RE = re.compile(r"^submit_(?:shop|shoptype|region)_.+?(?:\s*\(\d+\))*\.xlsx$", re.I)

SUBMIT_LAMBDA_ARN = "arn:aws:lambda:eu-central-1:497892281264:function:FORECAST-submit"


# ═══════════════════════════════════════════════════════════════════
#  SLACK HELPERS
# ═══════════════════════════════════════════════════════════════════


def get_slack_token():
    ssm = boto3.client("ssm")
    return ssm.get_parameter(
        Name="/forecast/slack-bot-token", WithDecryption=True
    )["Parameter"]["Value"]


def post_to_channel(token: str, channel_id: str, text: str):
    req = urllib.request.Request(
        "https://slack.com/api/chat.postMessage",
        data=json.dumps({"channel": channel_id, "text": text}).encode(),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {token}",
        },
        method="POST",
    )
    urllib.request.urlopen(req, timeout=10).read()


def slack_reply(response_url: str, text: str):
    if not response_url:
        return
    req = urllib.request.Request(
        response_url,
        data=json.dumps({"response_type": "in_channel", "text": text}).encode(),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    urllib.request.urlopen(req, timeout=10).read()


def send_to_slack(token: str, channel_id: str, response_url: str, text: str):
    """Send message, chunking at 3800 chars. Prefer response_url if available."""
    def dispatch(msg):
        if response_url:
            slack_reply(response_url, msg)
        elif channel_id:
            post_to_channel(token, channel_id, msg)

    lines = text.split("\n")
    chunk_lines = []
    chunk_len = 0
    in_code = False

    for line in lines:
        line_len = len(line) + 1
        if chunk_len + line_len > 3800 and chunk_lines:
            out = "\n".join(chunk_lines)
            if in_code:
                out += "\n```"
            dispatch(out)
            chunk_lines = []
            chunk_len = 0
            if in_code:
                chunk_lines.append("```")
                chunk_len = 4

        chunk_lines.append(line)
        chunk_len += line_len
        if line.strip() == "```":
            in_code = not in_code

    if chunk_lines:
        dispatch("\n".join(chunk_lines))


# ═══════════════════════════════════════════════════════════════════
#  SLACK FILE DOWNLOAD  (moved from submit)
# ═══════════════════════════════════════════════════════════════════


def download_slack_file_to_s3(user_id, channel_id, response_url):
    """Find the user's submit file in Slack, download, upload to S3.

    Returns (s3_key, entity_name, file_bytes, file_type) or (None, None, None, None).
    """
    token = get_slack_token()

    url = f"https://slack.com/api/conversations.history?channel={channel_id}&limit=10"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    resp = json.loads(urllib.request.urlopen(req).read())

    if not resp.get("ok"):
        slack_reply(response_url, f"Slack API error: {resp.get('error', 'unknown')}")
        return None, None, None, None

    file_info = None
    for msg in resp.get("messages", []):
        if msg.get("user") != user_id:
            continue
        for f in msg.get("files", []):
            if ANY_SUBMIT_RE.match(f.get("name", "")):
                file_info = f
                break
        if file_info:
            break

    if not file_info:
        slack_reply(response_url,
            "No matching file found. Upload a `submit_shop_*.xlsx`, "
            "`submit_shoptype_*.xlsx`, or `submit_region_*.xlsx` first.")
        return None, None, None, None

    file_name = file_info["name"]
    m = SHOP_FILE_RE.match(file_name)
    if m:
        entity_name = m.group(1).strip()
        file_type = "shop"
    else:
        m = SHOPTYPE_FILE_RE.match(file_name)
        if m:
            entity_name = m.group(1).strip()
            file_type = "shoptype"
        else:
            m = REGION_FILE_RE.match(file_name)
            entity_name = m.group(1).strip() if m else file_name
            file_type = "region"

    req = urllib.request.Request(
        file_info["url_private"],
        headers={"Authorization": f"Bearer {token}"},
    )
    file_bytes = urllib.request.urlopen(req).read()

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    base, ext = file_name.rsplit(".", 1)
    # Strip Slack duplicate counter like " (14)" before adding timestamp
    base = re.sub(r"\s*\(\d+\)$", "", base)
    timestamped_name = f"{base}_{ts}.{ext}"
    s3_key = f"{PREFIX}{timestamped_name}"
    S3.put_object(Bucket=BUCKET, Key=s3_key, Body=file_bytes)
    print(f"[submit-verify] Uploaded {file_name} to S3: {s3_key}")

    return s3_key, entity_name, file_bytes, file_type


# ═══════════════════════════════════════════════════════════════════
#  EXCEL HELPERS
# ═══════════════════════════════════════════════════════════════════


def s3_bytes(bucket, key):
    return S3.get_object(Bucket=bucket, Key=key)["Body"].read()


def cell_to_period(cv):
    if isinstance(cv, str):
        s = cv.strip()
        if PERIOD_RE.match(s):
            return s
    elif hasattr(cv, "strftime"):
        return cv.strftime("%Y-%m")
    return None


def is_na(v):
    return v is None or (isinstance(v, str) and v.strip().lower() == "na")


def is_blue(cell):
    fill = getattr(cell, "fill", None)
    if not fill or not getattr(fill, "patternType", None):
        return False
    rgb = getattr(getattr(fill, "fgColor", None), "rgb", None)
    return bool(rgb) and rgb.upper().endswith(BLUE_RGB)


def find_row(ws, label, start, end):
    for r in range(start, end + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip() == label:
            return r
    return None


def iter_avail_rows(ws, avail_row):
    r = avail_row + 1
    while r <= ws.max_row:
        v1 = ws.cell(r, 1).value
        v2 = ws.cell(r, 2).value
        if (v1 is None or (isinstance(v1, str) and not v1.strip())) and \
           (v2 is None or (isinstance(v2, str) and not str(v2).strip())):
            return
        yield (r, "" if v1 is None else str(v1).strip(), "" if v2 is None else str(v2).strip())
        r += 1


def find_summary_sheet(wb):
    for name in wb.sheetnames:
        ws = wb[name]
        v = ws.cell(1, 1).value
        if isinstance(v, str) and v.strip() == "TOP Products":
            return ws
    return wb[wb.sheetnames[0]]


_HEADER_SKIP_LABELS = {
    "Actuals", "Actuals LY", "YoY %", "F-YoY %", "Forecast",
    "Calculations", "Metric", "Available formats:",
}


def _find_header_row(ws):
    for r in range(1, min(80, ws.max_row) + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip() == "Product name":
            return r

    for r in range(1, min(80, ws.max_row) + 1):
        if not (cell_to_period(ws.cell(r, 3).value) or cell_to_period(ws.cell(r, 4).value)):
            continue
        v = ws.cell(r, 1).value
        label = str(v).strip() if v else ""
        if label in _HEADER_SKIP_LABELS or label.startswith("Destination:"):
            continue
        return r

    return -1


def read_forecast_config(bucket, xlsx_key):
    """Read current_month from Forecast_Input.xlsx config."""
    import pandas as pd
    raw = s3_bytes(bucket, xlsx_key)
    info = pd.read_excel(BytesIO(raw), sheet_name="info", header=None)
    return str(info.iat[3, 2]).strip()


# ═══════════════════════════════════════════════════════════════════
#  VALIDATION LOGIC
# ═══════════════════════════════════════════════════════════════════


def extract_sheet_data(ws, header_row):
    """Extract actuals, forecast (blue cells), and format totals from a product sheet."""
    month_cols = {}
    for c in range(3, ws.max_column + 1):
        p = cell_to_period(ws.cell(header_row, c).value)
        if p:
            month_cols[c] = p
    if not month_cols:
        return None

    scan_start = header_row
    scan_end = min(ws.max_row, header_row + 250)

    actuals_row = find_row(ws, "Actuals", scan_start, scan_end)
    actuals_ly_row = find_row(ws, "Actuals LY", scan_start, scan_end)
    forecast_row = find_row(ws, "Forecast", scan_start, scan_end)
    avail_row = find_row(ws, "Available formats:", scan_start, scan_end)

    actuals = {}
    actuals_ly = {}
    forecast = {}
    forecast_na = set()
    format_totals = {}

    for col, period in month_cols.items():
        if actuals_row:
            val = ws.cell(actuals_row, col).value
            if val is not None and not is_na(val):
                try:
                    actuals[period] = float(val)
                except (ValueError, TypeError):
                    pass

        if actuals_ly_row:
            val = ws.cell(actuals_ly_row, col).value
            if val is not None and not is_na(val):
                try:
                    actuals_ly[period] = float(val)
                except (ValueError, TypeError):
                    pass

        if forecast_row:
            cell = ws.cell(forecast_row, col)
            if is_blue(cell):
                if is_na(cell.value):
                    forecast_na.add(period)
                else:
                    try:
                        forecast[period] = float(cell.value)
                    except (ValueError, TypeError):
                        pass

        if avail_row:
            fmt_sum = 0.0
            has_any = False
            for rr, fmt, frame in iter_avail_rows(ws, avail_row):
                cell = ws.cell(rr, col)
                if is_blue(cell) and not is_na(cell.value):
                    try:
                        fmt_sum += float(cell.value)
                        has_any = True
                    except (ValueError, TypeError):
                        pass
            if has_any:
                format_totals[period] = fmt_sum

    return {
        "month_cols": month_cols,
        "actuals": actuals,
        "actuals_ly": actuals_ly,
        "forecast": forecast,
        "forecast_na": forecast_na,
        "format_totals": format_totals,
    }


def check_forecast_vs_actuals(sheet_data, product, current_month):
    """Check 1: Forecast for current month cannot be lower than actuals."""
    errors = []
    actuals_val = sheet_data["actuals"].get(current_month)
    forecast_val = sheet_data["forecast"].get(current_month)

    if actuals_val is not None and forecast_val is not None:
        if forecast_val < actuals_val:
            errors.append(
                f"`{product}`: `{current_month}` — forecast {forecast_val:,.0f} vs actuals {actuals_val:,.0f}"
            )
    return errors


def check_format_mismatch(sheet_data, product):
    """Check 2: Format row totals must match Forecast row values when both exist."""
    errors = []
    all_periods = set(sheet_data["forecast"].keys()) | set(sheet_data["format_totals"].keys())

    for period in sorted(all_periods):
        forecast_val = sheet_data["forecast"].get(period)
        format_sum = sheet_data["format_totals"].get(period)

        has_forecast = forecast_val is not None
        has_formats = format_sum is not None

        if has_forecast and has_formats:
            if abs(forecast_val - format_sum) > 0.5:
                errors.append(
                    f"`{product}`: Mismatch in `{period}` — "
                    f"{forecast_val:,.0f} total vs {format_sum:,.0f} formats"
                )
    return errors


def check_coverage_gaps(sheet_data, product, current_month):
    """Check 3: No gaps within the forecasted month range."""
    errors = []
    forecast_na = sheet_data.get("forecast_na", set())

    filled_months = sorted([p for p in sheet_data["forecast"] if p >= current_month])
    na_months = sorted([p for p in forecast_na if p >= current_month])

    if not filled_months and not na_months:
        return errors

    if na_months and filled_months:
        na_fmt = ", ".join(f"`{m}`" for m in na_months)
        errors.append(f"`{product}`: Missing in {na_fmt}")

    if len(filled_months) >= 2:
        first = filled_months[0]
        last = filled_months[-1]
        full_range = []
        cursor = datetime.strptime(first, "%Y-%m")
        end = datetime.strptime(last, "%Y-%m")
        while cursor <= end:
            full_range.append(cursor.strftime("%Y-%m"))
            if cursor.month == 12:
                cursor = cursor.replace(year=cursor.year + 1, month=1)
            else:
                cursor = cursor.replace(month=cursor.month + 1)

        forecast_set = set(filled_months)
        structural_gaps = [m for m in full_range if m not in forecast_set and m not in na_months]
        if structural_gaps:
            gap_fmt = ", ".join(f"`{m}`" for m in structural_gaps)
            errors.append(f"`{product}`: Missing in {gap_fmt}")

    return errors


def validate_shop_file(file_bytes, current_month):
    """Run all 3 checks on a shop-level submit file.

    Returns (passed, errors_by_type, product_data) where errors_by_type is
    {"mismatch": [...], "gaps": [...], "actuals": [...]}.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    summary_ws = find_summary_sheet(wb)
    errors_by_type = {"mismatch": [], "gaps": [], "actuals": []}
    product_data = {}

    for ws in wb.worksheets:
        if ws.title == summary_ws.title:
            continue

        header = _find_header_row(ws)
        if header < 0:
            continue

        product = ws.title
        sheet_data = extract_sheet_data(ws, header)
        if not sheet_data:
            continue

        product_data[product] = sheet_data

        errors_by_type["actuals"].extend(
            check_forecast_vs_actuals(sheet_data, product, current_month))
        errors_by_type["mismatch"].extend(
            check_format_mismatch(sheet_data, product))
        errors_by_type["gaps"].extend(
            check_coverage_gaps(sheet_data, product, current_month))

    all_errors = errors_by_type["mismatch"] + errors_by_type["gaps"] + errors_by_type["actuals"]
    return (len(all_errors) == 0, errors_by_type, product_data)


def extract_shoptype_block(ws, dest_r, block_end, month_cols):
    """Extract forecast total + individual shop data from a destination block."""
    forecast_row = None
    shop_rows = []

    for r in range(dest_r + 1, block_end + 1):
        v = ws.cell(r, 1).value
        label = str(v).strip() if v is not None else ""
        if label == "Forecast":
            forecast_row = r
            continue
        if forecast_row is not None and label and label not in (
            "Actuals", "Actuals LY", "YoY %", "F-YoY %",
            "Calculations", "Last 6w % growth:", "Metric", "Product name",
        ) and not label.startswith("Destination:"):
            if any(is_blue(ws.cell(r, c)) for c in month_cols):
                shop_rows.append((r, label))

    forecast_total = {}
    forecast_na = set()
    if forecast_row:
        for col, period in month_cols.items():
            cell = ws.cell(forecast_row, col)
            if is_blue(cell):
                if is_na(cell.value):
                    forecast_na.add(period)
                else:
                    try:
                        forecast_total[period] = float(cell.value)
                    except (ValueError, TypeError):
                        pass

    shop_sum = {}
    shop_na = set()
    # Per-shop detail: {shop_name: {"filled": {period: val}, "na": set()}}
    shop_detail = {}
    for shop_r, shop_name in shop_rows:
        shop_filled = {}
        shop_na_set = set()
        for col, period in month_cols.items():
            cell = ws.cell(shop_r, col)
            if is_blue(cell):
                if is_na(cell.value):
                    shop_na.add(period)
                    shop_na_set.add(period)
                else:
                    try:
                        val = float(cell.value)
                        shop_sum[period] = shop_sum.get(period, 0) + val
                        shop_filled[period] = val
                    except (ValueError, TypeError):
                        pass
        shop_detail[shop_name] = {"filled": shop_filled, "na": shop_na_set}

    all_na = forecast_na | shop_na
    forecast = {}
    forecast.update(shop_sum)
    forecast.update(forecast_total)

    return forecast_total, shop_sum, all_na, forecast, shop_detail


def validate_shoptype_file(file_bytes, current_month):
    """Run validation checks on a shoptype-level submit file.

    Shoptype files have NO format breakdown — only shop-level blue cells.
    Structure per product sheet, per destination block:
      Destination: X  | months...
      Actuals         | non-blue
      Actuals LY      | non-blue
      YoY %           | non-blue
      F-YoY %         | non-blue (user-editable growth rates)
      Forecast        | non-blue (formula: (1+F-YoY)*LY)
      shop_name       | blue: na or value per forecast month
      Actuals LY      |
      F-YoY %         |
      ...more shops...
      Other [shoptype] | blue: residual forecast values

    Checks:
    1. Partial coverage: shops with some months filled but others 'na'
    2. Forecast >= current month actuals
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    errors_by_type = {"partial": [], "actuals": []}
    product_data = {}

    for ws in wb.worksheets:
        v11 = ws.cell(1, 1).value
        if isinstance(v11, str) and v11.strip() == "TOP Products":
            continue
        if ws.title.lower().startswith("shoptype "):
            continue

        product = ws.title

        dest_rows = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.strip().startswith("Destination:"):
                dest_rows.append(r)

        product_forecast = {}
        product_actuals_ly = {}
        product_shops = {}
        product_na = set()

        for idx, dest_r in enumerate(dest_rows):
            block_end = dest_rows[idx + 1] - 1 if idx + 1 < len(dest_rows) else ws.max_row

            dest_raw = str(ws.cell(dest_r, 1).value).strip()
            dest_value = dest_raw[len("Destination:"):].strip() if ":" in dest_raw else ""

            month_cols = {}
            for c in range(3, ws.max_column + 1):
                p = cell_to_period(ws.cell(dest_r, c).value)
                if p:
                    month_cols[c] = p
            if not month_cols:
                continue

            fc_total, shop_sum, block_na, block_forecast, shop_detail = (
                extract_shoptype_block(ws, dest_r, block_end, month_cols))

            label = f"{product} ({dest_value})" if dest_value else product

            # Check 1: Partial coverage — shops with some forecast months
            # filled but others still 'na'. Shops with ALL na are fine
            # (they delegate to Other). Shops with ALL filled are fine.
            for shop_name, detail in shop_detail.items():
                filled = sorted(p for p in detail["filled"] if p >= current_month)
                na = sorted(p for p in detail["na"] if p >= current_month)
                if filled and na:
                    na_fmt = ", ".join(f"`{m}`" for m in na)
                    errors_by_type["partial"].append(
                        f"`{label}` / `{shop_name}`: Filled {len(filled)}mo "
                        f"but missing {na_fmt}"
                    )

            # Check 2: Forecast >= actuals for current month
            actuals_row = find_row(ws, "Actuals", dest_r + 1, min(dest_r + 10, block_end))
            if actuals_row:
                for col, period in month_cols.items():
                    if period == current_month:
                        act_val = ws.cell(actuals_row, col).value
                        fc_val = block_forecast.get(period)
                        if act_val is not None and fc_val is not None:
                            try:
                                if float(fc_val) < float(act_val):
                                    errors_by_type["actuals"].append(
                                        f"`{label}`: `{period}` — forecast {float(fc_val):,.0f} "
                                        f"vs actuals {float(act_val):,.0f}"
                                    )
                            except (ValueError, TypeError):
                                pass

            # Read Actuals LY for YoY calculation in success report
            actuals_ly_row = find_row(ws, "Actuals LY", dest_r + 1, min(dest_r + 10, block_end))
            block_ly = {}
            if actuals_ly_row:
                for col, period in month_cols.items():
                    val = ws.cell(actuals_ly_row, col).value
                    if val is not None and not is_na(val):
                        try:
                            block_ly[period] = float(val)
                        except (ValueError, TypeError):
                            pass

            for p, v in block_forecast.items():
                product_forecast[p] = product_forecast.get(p, 0) + v
            for p, v in block_ly.items():
                product_actuals_ly[p] = product_actuals_ly.get(p, 0) + v
            product_na |= block_na

            # Store per-shop forecast for success report
            for shop_name, detail in shop_detail.items():
                if detail["filled"]:
                    product_shops.setdefault(shop_name, {})
                    for p, v in detail["filled"].items():
                        product_shops[shop_name][p] = (
                            product_shops[shop_name].get(p, 0) + v)

        if product_forecast:
            product_data[product] = {
                "forecast": product_forecast,
                "forecast_na": product_na,
                "actuals": {},
                "actuals_ly": product_actuals_ly,
                "month_cols": {},
                "format_totals": {},
                "shops": product_shops,
            }

    all_errors = errors_by_type["partial"] + errors_by_type["actuals"]
    return (len(all_errors) == 0, errors_by_type, product_data)


def build_failure_report(entity_name, errors_by_type):
    """Build a structured Slack error message grouped by check type."""
    lines = [f"❌ Submitted forecast for `{entity_name}` does not pass Quality check — please resubmit!"]

    if errors_by_type.get("partial"):
        lines.append("")
        lines.append("*Incomplete forecast period (some months filled, others missing):*")
        for e in errors_by_type["partial"]:
            lines.append(f"  • {e}")

    if errors_by_type.get("mismatch"):
        lines.append("")
        lines.append("*Format vs Total forecast mismatch:*")
        for e in errors_by_type["mismatch"]:
            lines.append(f"  • {e}")

    if errors_by_type.get("gaps"):
        lines.append("")
        lines.append("*Missing forecasts:*")
        for e in errors_by_type["gaps"]:
            lines.append(f"  • {e}")

    if errors_by_type.get("actuals"):
        lines.append("")
        lines.append("*Current month forecast is too low:*")
        for e in errors_by_type["actuals"]:
            lines.append(f"  • {e}")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════
#  SUCCESS REPORT
# ═══════════════════════════════════════════════════════════════════


def load_previous_forecasts(bucket, key, entity_name, current_month):
    """Load previous forecast totals per product from the combined CSV."""
    try:
        raw = s3_bytes(bucket, key).decode("utf-8", "replace")
    except Exception:
        return {}

    prev = {}
    for row in csv.DictReader(StringIO(raw)):
        shop = (row.get("forecasted_shop") or "").strip()
        period = (row.get("period") or "").strip()
        product = (row.get("product") or "").strip()
        if shop != entity_name or period < current_month:
            continue
        try:
            val = float(row.get("forecast") or 0)
        except (ValueError, TypeError):
            val = 0.0
        prev.setdefault(product, {})
        prev[product][period] = prev[product].get(period, 0) + val

    return prev


def _delta(new_val, old_val):
    """Return arrow + % change comparing new vs old forecast."""
    if old_val is None or old_val == 0:
        return "new"
    pct = (new_val / old_val - 1) * 100
    if abs(pct) < 0.5:
        return f"→  0%"
    arrow = "↑" if pct > 0 else "↓"
    return f"{arrow} {pct:+.0f}%"


def build_success_report(entity_name, product_data, current_month, prev_forecasts):
    """Build a Slack-formatted bullet list for a passing submission."""
    lines = [f"✅ Forecast for `{entity_name}` passed Quality check.\n"]

    rows = []
    period_label = ""
    for product, sd in product_data.items():
        forecast = sd["forecast"]
        actuals_ly = sd.get("actuals_ly", {})

        blue_months = sorted(m for m in forecast if m >= current_month)
        if not blue_months:
            continue

        if not period_label:
            period_label = f"{blue_months[0]} to {blue_months[-1]} ({len(blue_months)}mo)"

        fc_total = sum(forecast.get(m, 0) for m in blue_months)
        ly_total = sum(actuals_ly.get(m, 0) for m in blue_months)

        prev_prod = prev_forecasts.get(product, {})
        prev_total = sum(prev_prod.get(m, 0) for m in blue_months)
        has_prev = prev_total > 0

        ly_str = f"{int(ly_total):,}" if ly_total > 0 else "—"
        fyoy = f"{(fc_total / ly_total - 1) * 100:.0f}%" if ly_total > 0 else "na"
        delta = _delta(fc_total, prev_total if has_prev else None)

        # Collect shop-level detail for shoptype files
        shops = sd.get("shops", {})
        shop_rows = []
        for shop_name, shop_fc in shops.items():
            shop_total = sum(shop_fc.get(m, 0) for m in blue_months)
            if shop_total > 0:
                shop_rows.append((shop_total, shop_name))
        shop_rows.sort(key=lambda r: r[0], reverse=True)

        rows.append((fc_total, product, ly_str, fyoy, delta, shop_rows))

    rows.sort(key=lambda r: r[0], reverse=True)

    for fc_total, product, ly_str, fyoy, delta, shop_rows in rows:
        lines.append(
            f"  • `{product}` — Forecast: *{int(fc_total):,}* | "
            f"Actuals LY: {ly_str} | F-YoY %: {fyoy} | vs Last version: {delta}"
        )
        for shop_total, shop_name in shop_rows:
            lines.append(f"      ↳ {shop_name}: {int(shop_total):,}")

    if period_label:
        lines.append(f"\n_Period: {period_label}_")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════
#  HANDLER
# ═══════════════════════════════════════════════════════════════════


def lambda_handler(event, context):
    # --- Mode 1: Called from review-trigger with Slack user context ---
    user_id = event.get("user_id", "")
    channel_id = event.get("channel_id", "")
    response_url = event.get("response_url", "")

    # --- Mode 2: Direct call with s3_key (local testing / step function) ---
    s3_key = event.get("s3_key", "")
    local_file = event.get("local_file", "")
    entity_name = event.get("entity_name", "")
    file_type = event.get("file_type", "shop")
    current_month = event.get("current_month", "")

    file_bytes = None

    # If called from Slack (user_id present), download file first
    if user_id and channel_id:
        result = download_slack_file_to_s3(user_id, channel_id, response_url)
        s3_key, entity_name, file_bytes, file_type = result
        if not s3_key:
            return {"ok": False, "error": "file download failed"}
        slack_reply(response_url, f"Verifying `{entity_name}` — performing a quality check...")

    # Read current_month from config if not provided
    if not current_month:
        try:
            current_month = read_forecast_config(BUCKET, INPUT_XLSX_KEY)
        except Exception as e:
            print(f"[submit-verify] Failed to read config: {e}")
            return {"ok": False, "error": f"Cannot read current_month: {e}"}

    print(f"[submit-verify] Validating {entity_name} ({file_type}), "
          f"current_month={current_month}")

    # Load Excel bytes if not already loaded
    if file_bytes is None:
        if local_file:
            with open(local_file, "rb") as f:
                file_bytes = f.read()
        elif s3_key:
            file_bytes = s3_bytes(BUCKET, s3_key)
        else:
            return {"ok": False, "error": "No file to validate"}

    # Run validation
    try:
        if file_type == "shop":
            passed, errors_by_type, product_data = validate_shop_file(file_bytes, current_month)
        elif file_type == "shoptype":
            passed, errors_by_type, product_data = validate_shoptype_file(file_bytes, current_month)
        else:
            passed, errors_by_type, product_data = True, {}, {}
        total_errors = sum(len(v) for v in errors_by_type.values()) if errors_by_type else 0
        print(f"[submit-verify] Validation: passed={passed}, errors={total_errors}")
    except Exception as e:
        print(f"[submit-verify] Validation crashed: {e}")
        import traceback; traceback.print_exc()
        passed, errors_by_type, product_data = False, {"mismatch": [f"Internal error: {e}"]}, {}

    # Load previous forecasts BEFORE submit rebuilds the CSV
    prev_forecasts = {}
    try:
        prev_forecasts = load_previous_forecasts(BUCKET, COMBINED_KEY, entity_name, current_month)
        print(f"[submit-verify] Loaded prev forecasts: {len(prev_forecasts)} products")
    except Exception as e:
        print(f"[submit-verify] Could not load previous forecasts: {e}")

    # Build and send Slack message
    try:
        token = get_slack_token()
        if passed and file_type == "shop":
            # Only show success summary for shop files
            message = build_success_report(entity_name, product_data, current_month, prev_forecasts)
            print(f"[submit-verify] Posting success to Slack (channel={channel_id}, msg_len={len(message)})")
            send_to_slack(token, channel_id, response_url, message)
        elif not passed:
            message = build_failure_report(entity_name, errors_by_type)
            print(f"[submit-verify] Posting failure to Slack (channel={channel_id}, msg_len={len(message)})")
            send_to_slack(token, channel_id, response_url, message)
            # Delete the failed file from S3 — it should not be processed
            try:
                S3.delete_object(Bucket=BUCKET, Key=s3_key)
                print(f"[submit-verify] Deleted failed file from S3: {s3_key}")
            except Exception as del_err:
                print(f"[submit-verify] Could not delete failed file: {del_err}")
    except Exception as e:
        print(f"[submit-verify] Slack posting failed: {e}")
        import traceback; traceback.print_exc()

    # If passed, invoke FORECAST-submit to process the file
    if passed:
        try:
            send_to_slack(token, channel_id, response_url, f"Submitting `{entity_name}` — adding to seperate forecasts...")
            submit_payload = {
                "response_url": response_url,
                "channel_id": channel_id,
                "entity_name": entity_name,
            }
            boto3.client("lambda").invoke(
                FunctionName=SUBMIT_LAMBDA_ARN,
                InvocationType="Event",
                Payload=json.dumps(submit_payload).encode("utf-8"),
            )
            print(f"[submit-verify] Invoked FORECAST-submit for {entity_name}")
        except Exception as e:
            print(f"[submit-verify] Failed to invoke submit: {e}")
            import traceback; traceback.print_exc()

    return {
        "ok": passed,
        "entity_name": entity_name,
        "file_type": file_type,
        "s3_key": s3_key,
        "current_month": current_month,
        "errors": [e for errs in errors_by_type.values() for e in errs] if errors_by_type else [],
    }
