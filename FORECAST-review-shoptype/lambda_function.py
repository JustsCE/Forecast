import os
import json
import re
import urllib.request
import urllib.parse
import urllib.error
from dataclasses import dataclass
from io import BytesIO
from typing import List, Tuple, Optional

import boto3
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter


def lambda_handler(event, context):
    @dataclass(frozen=True)
    class Config:
        bucket: str = "bi-automations"
        in_key: str = "Forecast/actuals.csv"
        out_key: str = "Forecast/review.xlsx"

        date_col: str = "fulldate"
        actuals_col: str = "actuals"

        # ONLY shoptype scope is supported
        shoptype_col: str = "shoptype"

        product_col: str = "forecast_product"
        label_col: str = "TOP Products"
        week_col: str = "week"

        destination_col: str = "destination_region"

        # Column used for "Top 5 shops" in product sheets
        shop_col: str = "forecasted_shop"

        forecast_key: str = "Forecast/Seperate Forecasts/seperate_forecasts_combined.csv"

        excluded_products: Tuple[str, ...] = ("EXCLUDED PRODUCT", "NEW PRODUCT")
        top_n: int = 15

    CFG = Config()
    S3 = boto3.client("s3")

    _WEEK_RE = re.compile(r"^(?P<y>\d{4})-(?P<w>\d{2})$")

    WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
    FORECAST_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")

    LEFT_ALIGN = Alignment(horizontal="left")
    RIGHT_ALIGN = Alignment(horizontal="right")
    INDENT_ALIGN = Alignment(horizontal="left", indent=1)
    PERCENT_FMT = "0.0%"
    INTEGER_FMT = "#,##0"

    THIN = Side(style="thin", color="000000")
    DOUBLE = Side(style="double", color="000000")
    HEADER_BOTTOM_BORDER = Border(bottom=THIN)
    ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BOLD_FONT = Font(bold=True)

    DEST_USCA = "US+CA"
    DEST_EUROW = "EU+RoW"

    # ── Slack helpers ─────────────────────────────────────────────────

    def slack_api_post_form(method: str, token: str, fields: dict, timeout: int = 20) -> dict:
        body = urllib.parse.urlencode(fields).encode("utf-8")
        req = urllib.request.Request(
            f"https://slack.com/api/{method}",
            data=body,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/x-www-form-urlencoded; charset=utf-8",
            },
            method="POST",
        )
        try:
            resp = urllib.request.urlopen(req, timeout=timeout).read()
        except urllib.error.HTTPError as e:
            err_body = e.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Slack API HTTPError {e.code} for {method}: {err_body}") from e

        data = json.loads(resp.decode("utf-8"))
        if not data.get("ok"):
            raise RuntimeError(f"Slack API {method} failed: {data}")
        return data

    def slack_upload_xlsx(file_bytes: bytes, filename: str, message: str):
        token = os.environ["SLACK_BOT_TOKEN"]
        channels_env = os.environ["SLACK_CHANNEL_ID"]
        channel_ids = [c.strip() for c in channels_env.split(",") if c.strip()]
        if not channel_ids:
            raise ValueError("SLACK_CHANNEL_ID is empty")

        init = slack_api_post_form(
            "files.getUploadURLExternal",
            token,
            {"filename": filename, "length": str(len(file_bytes))},
        )
        upload_url = init["upload_url"]
        file_id = init["file_id"]

        upload_req = urllib.request.Request(
            upload_url,
            data=file_bytes,
            headers={"Content-Type": "application/octet-stream"},
            method="POST",
        )
        try:
            upload_resp = urllib.request.urlopen(upload_req, timeout=30)
            status = getattr(upload_resp, "status", None) or upload_resp.getcode()
            if int(status) != 200:
                raise RuntimeError(f"Upload to Slack upload_url failed: HTTP {status}")
        except urllib.error.HTTPError as e:
            err_body = e.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Upload to Slack upload_url HTTPError {e.code}: {err_body}") from e

        files_payload = json.dumps([{"id": file_id, "title": filename}])

        complete_fields = {"files": files_payload, "initial_comment": message}
        if len(channel_ids) == 1:
            complete_fields["channel_id"] = channel_ids[0]
        else:
            complete_fields["channels"] = ",".join(channel_ids)

        slack_api_post_form("files.completeUploadExternal", token, complete_fields)

    # ── Scope parsing ─────────────────────────────────────────────────

    def parse_scope_shoptype(event, cfg: Config) -> Tuple[str, str, str]:
        raw = (event.get("text") or event.get("shoptype") or "").strip()
        low = raw.lower()

        if low.startswith("shoptype "):
            val = raw[9:].strip()
        elif event.get("shoptype"):
            val = str(event.get("shoptype")).strip()
        else:
            val = raw  # bare text = shoptype name

        if val:
            return ("shoptype", val, cfg.shoptype_col)
        return ("", "", "")

    # ── Numeric / date helpers ────────────────────────────────────────

    def to_num(s: pd.Series) -> pd.Series:
        return pd.to_numeric(s, errors="coerce").fillna(0).astype(float)

    def safe_growth(curr: pd.Series, prev: pd.Series) -> pd.Series:
        curr = to_num(curr)
        prev = to_num(prev)
        out = np.where(prev != 0, (curr / prev) - 1.0, 0.0)
        out = np.where(np.isfinite(out), out, 0.0)
        return pd.Series(out, index=curr.index)

    def normalize_ratio(s: pd.Series) -> pd.Series:
        s = to_num(s)
        mask = s.abs() > 5
        s.loc[mask] = s.loc[mask] / 100.0
        return s

    def parse_week_key(s: str) -> Tuple[int, int]:
        if not isinstance(s, str):
            return (-1, -1)
        m = _WEEK_RE.match(s.strip())
        if not m:
            return (-1, -1)
        return (int(m.group("y")), int(m.group("w")))

    def shift_week_year(week_key: str, new_year: int) -> str:
        if not isinstance(week_key, str) or "-" not in week_key:
            return f"{new_year}-00"
        _, ww = week_key.split("-", 1)
        return f"{new_year}-{ww.zfill(2)}"

    def last_3_weeks_keys_any(df: pd.DataFrame, t: int, week_col: str) -> List[str]:
        d = df[df[week_col].astype(str).str.startswith(f"{t}-")]
        keys = sorted(d[week_col].dropna().astype(str).unique().tolist(), key=parse_week_key)
        if not keys:
            return []
        keys = keys[:-1]
        return keys[-3:] if len(keys) >= 3 else keys

    def last_3_full_months_labels(max_dt_t: pd.Timestamp) -> List[str]:
        month_start = pd.Timestamp(year=max_dt_t.year, month=max_dt_t.month, day=1)
        last_full_month_end = month_start - pd.Timedelta(days=1)
        last_full_month = last_full_month_end.to_period("M")
        months = pd.period_range(end=last_full_month, periods=3, freq="M")
        return [str(p) for p in months]

    def shift_month_label_ly(m: str) -> str:
        return f"{int(m[:4]) - 1}{m[4:]}"

    def destination_group(v) -> str:
        s = str(v or "").strip().upper()
        if s in ("US", "CA", "US+CA", "CA+US"):
            return DEST_USCA
        return DEST_EUROW

    # ── Data cleaning / masking ───────────────────────────────────────

    def clean_input(df: pd.DataFrame, cfg: Config) -> pd.DataFrame:
        df = df.copy()
        df = df[~df[cfg.product_col].astype(str).isin(cfg.excluded_products)]
        df[cfg.date_col] = pd.to_datetime(df[cfg.date_col], errors="coerce")
        df = df[df[cfg.date_col].notna()].copy()
        df[cfg.actuals_col] = to_num(df[cfg.actuals_col])
        df["_month"] = df[cfg.date_col].dt.to_period("M").astype(str)
        if cfg.destination_col in df.columns:
            df["_dest_group"] = df[cfg.destination_col].apply(destination_group)
        else:
            df["_dest_group"] = DEST_EUROW
        return df

    def mask_for_scope(df: pd.DataFrame, scope_col: str, scope_value: str) -> pd.Series:
        if scope_value:
            return df[scope_col].astype(str) == str(scope_value)
        return pd.Series(True, index=df.index)

    def mask_for_dest_group(df: pd.DataFrame, grp: str) -> pd.Series:
        if "_dest_group" in df.columns and grp:
            return df["_dest_group"].astype(str) == str(grp)
        return pd.Series(True, index=df.index)

    def agg_sum(df: pd.DataFrame, cfg: Config, mask: pd.Series, label: str) -> pd.DataFrame:
        return (
            df[mask]
            .groupby(cfg.product_col, as_index=False)[cfg.actuals_col]
            .sum()
            .rename(columns={cfg.product_col: cfg.label_col, cfg.actuals_col: label})
        )

    def ensure_col(df: pd.DataFrame, col: str) -> pd.Series:
        if col in df.columns:
            return to_num(df[col])
        return pd.Series(0.0, index=df.index)

    # ── compute_review_table ──────────────────────────────────────────

    def compute_review_table(
        df_raw: pd.DataFrame,
        scope_value: str,
        scope_col: str,
        cfg: Config,
        dest_group_filter: Optional[str] = None,
    ) -> Tuple[int, pd.DataFrame]:
        df = clean_input(df_raw, cfg)
        scope_value = (scope_value or "").strip()

        scope_mask = mask_for_scope(df, scope_col, scope_value)
        if dest_group_filter:
            scope_mask = scope_mask & mask_for_dest_group(df, dest_group_filter)

        df_scope = df[scope_mask] if scope_value else df[scope_mask]
        if df_scope.empty:
            df_scope = df[scope_mask]

        t = int(df_scope[cfg.date_col].dt.year.max()) if not df_scope.empty else int(df[cfg.date_col].dt.year.max())
        max_dt_t = df_scope.loc[df_scope[cfg.date_col].dt.year == t, cfg.date_col].max()
        if pd.isna(max_dt_t):
            max_dt_t = df.loc[df[cfg.date_col].dt.year == t, cfg.date_col].max()

        cutoff_t1 = max_dt_t - pd.DateOffset(years=1)

        base_t = scope_mask & (df[cfg.date_col].dt.year == t) & (df[cfg.date_col] <= max_dt_t)
        base_t1_ytd = scope_mask & (df[cfg.date_col].dt.year == t - 1) & (df[cfg.date_col] <= cutoff_t1)
        base_t1_fy = scope_mask & (df[cfg.date_col].dt.year == t - 1)
        base_t2_fy = scope_mask & (df[cfg.date_col].dt.year == t - 2)

        out = (
            agg_sum(df, cfg, base_t, f"YTD {t}")
            .merge(agg_sum(df, cfg, base_t1_ytd, f"YTD {t-1}"), on=cfg.label_col, how="outer")
            .merge(agg_sum(df, cfg, base_t1_fy, f"FY {t-1}"), on=cfg.label_col, how="outer")
            .merge(agg_sum(df, cfg, base_t2_fy, f"FY {t-2}"), on=cfg.label_col, how="outer")
            .fillna(0)
        )

        last3w_t_keys = last_3_weeks_keys_any(df, t, cfg.week_col)
        last3w_t1_keys = [shift_week_year(k, t - 1) for k in last3w_t_keys]

        out = (
            out.merge(
                agg_sum(df, cfg, scope_mask & df[cfg.week_col].astype(str).isin(last3w_t_keys), f"Last 3W {t}"),
                on=cfg.label_col,
                how="outer",
            )
            .merge(
                agg_sum(df, cfg, scope_mask & df[cfg.week_col].astype(str).isin(last3w_t1_keys), f"Last 3W {t-1}"),
                on=cfg.label_col,
                how="outer",
            )
            .fillna(0)
        )

        last3m_t = last_3_full_months_labels(max_dt_t)
        last3m_t1 = [shift_month_label_ly(m) for m in last3m_t]

        for m_t, m_t1 in zip(last3m_t, last3m_t1):
            msum_t = agg_sum(df, cfg, scope_mask & (df["_month"] == m_t), "_mt")
            msum_t1 = agg_sum(df, cfg, scope_mask & (df["_month"] == m_t1), "_mt1")
            tmp = msum_t.merge(msum_t1, on=cfg.label_col, how="outer").fillna(0)
            tmp[m_t] = safe_growth(tmp["_mt"], tmp["_mt1"])
            out = out.merge(tmp[[cfg.label_col, m_t]], on=cfg.label_col, how="outer").fillna(0)

        out["YTD % Growth"] = safe_growth(ensure_col(out, f"YTD {t}"), ensure_col(out, f"YTD {t-1}"))
        out["FY % Growth"] = safe_growth(ensure_col(out, f"FY {t-1}"), ensure_col(out, f"FY {t-2}"))
        out["Last 3W % Growth"] = safe_growth(ensure_col(out, f"Last 3W {t}"), ensure_col(out, f"Last 3W {t-1}"))
        out["% of Total YTD"] = 0.0

        out = out.sort_values(by=f"YTD {t}", ascending=False).reset_index(drop=True)

        out["_rank_ytd"] = np.arange(1, len(out) + 1, dtype=int)
        fy_col = f"FY {t-1}"
        out["_fy_t1"] = ensure_col(out, fy_col)

        keep_mask = (out["_rank_ytd"] <= cfg.top_n) & (out["_fy_t1"] >= 1000.0)

        top_rows = out.loc[keep_mask].copy()
        other_rows = out.loc[~keep_mask].copy()

        helper_cols = {cfg.label_col, "_rank_ytd", "_fy_t1"}
        numeric_cols = [c for c in out.columns if c not in helper_cols and pd.api.types.is_numeric_dtype(out[c])]

        include_other = False
        other_df = None

        if not other_rows.empty:
            other_agg = other_rows[numeric_cols].sum(numeric_only=True).to_dict()
        else:
            other_agg = {c: 0.0 for c in numeric_cols}

        if any(float(other_agg.get(c, 0.0) or 0.0) != 0.0 for c in numeric_cols):
            include_other = True
            other_agg[cfg.label_col] = "Other products"
            other_df = pd.DataFrame([other_agg])

        combined = top_rows.drop(columns=["_rank_ytd", "_fy_t1"], errors="ignore")
        if include_other and other_df is not None:
            combined = pd.concat([combined, other_df], ignore_index=True)

        total_vals = combined[numeric_cols].sum(numeric_only=True).to_dict()
        total_label = f"TOTAL {scope_value}".strip()
        total_row = pd.DataFrame([{cfg.label_col: total_label, **total_vals}])

        out_final = pd.concat([combined, total_row], ignore_index=True)

        out_final["YTD % Growth"] = safe_growth(ensure_col(out_final, f"YTD {t}"), ensure_col(out_final, f"YTD {t-1}"))
        out_final["FY % Growth"] = safe_growth(ensure_col(out_final, f"FY {t-1}"), ensure_col(out_final, f"FY {t-2}"))
        out_final["Last 3W % Growth"] = safe_growth(
            ensure_col(out_final, f"Last 3W {t}"), ensure_col(out_final, f"Last 3W {t-1}")
        )

        total_mask = out_final[cfg.label_col].astype(str).str.startswith("TOTAL")
        non_total_mask = ~total_mask

        total_ytd = float(to_num(out_final.loc[non_total_mask, f"YTD {t}"]).sum())
        if total_ytd != 0:
            out_final.loc[non_total_mask, "% of Total YTD"] = to_num(out_final.loc[non_total_mask, f"YTD {t}"]) / total_ytd
            out_final.loc[total_mask, "% of Total YTD"] = 1.0
        else:
            out_final.loc[non_total_mask, "% of Total YTD"] = 0.0
            out_final.loc[total_mask, "% of Total YTD"] = 0.0

        for m_t, m_t1 in zip(last3m_t, last3m_t1):
            tot_mt = float(df[scope_mask & (df["_month"] == m_t)][cfg.actuals_col].sum())
            tot_mt1 = float(df[scope_mask & (df["_month"] == m_t1)][cfg.actuals_col].sum())
            out_final.loc[total_mask, m_t] = (tot_mt / tot_mt1 - 1.0) if tot_mt1 != 0 else 0.0

        out_final = out_final.rename(
            columns={
                "YTD % Growth": "YTD % growth",
                "FY % Growth": "FY % growth",
                "Last 3W % Growth": "Last 3w % growth",
            }
        )

        percent_cols = ["% of Total YTD", "YTD % growth", "FY % growth", "Last 3w % growth", *last3m_t]
        for c in percent_cols:
            if c in out_final.columns:
                out_final[c] = normalize_ratio(out_final[c])

        final_cols = [
            cfg.label_col,
            "% of Total YTD",
            f"YTD {t}",
            f"YTD {t-1}",
            "YTD % growth",
            f"FY {t-1}",
            f"FY {t-2}",
            "FY % growth",
            "Last 3w % growth",
            *last3m_t,
        ]
        for c in final_cols:
            if c not in out_final.columns:
                out_final[c] = 0.0

        out_final = out_final[final_cols].fillna(0)

        out_final["_is_total"] = out_final[cfg.label_col].astype(str).str.startswith("TOTAL")
        out_final = (
            out_final.sort_values(["_is_total", f"YTD {t}"], ascending=[True, False])
            .drop(columns=["_is_total"])
            .reset_index(drop=True)
        )

        return t, out_final

    # ── Sheet-name helpers ────────────────────────────────────────────

    def safe_sheet_base(name: str) -> str:
        bad = r'[:\\/?*\[\]]'
        s = re.sub(bad, " ", str(name)).strip()
        s = re.sub(r"\s+", " ", s)
        return (s[:31] if s else "Sheet")

    def dedupe_sheet_name(base: str, used: set) -> str:
        name = base
        i = 2
        while name in used:
            suffix = f" {i}"
            name = (base[: (31 - len(suffix))] + suffix) if len(base) + len(suffix) > 31 else (base + suffix)
            i += 1
        used.add(name)
        return name

    # ── Month helpers ─────────────────────────────────────────────────

    def month_labels_m2_to_p11(max_dt_t: pd.Timestamp) -> List[str]:
        cur = max_dt_t.to_period("M")
        start = cur - 2
        end = cur + 11
        months = pd.period_range(start=start, end=end, freq="M")
        return [str(p) for p in months]

    def get_agg_value(agg: pd.Series, key: tuple) -> float:
        try:
            return float(agg.loc[key])
        except Exception:
            return 0.0

    def region_has_actuals(m_agg: pd.Series, scope_key: str, prod_key: str, dest_group: str) -> bool:
        try:
            v = float(m_agg.loc[(scope_key, prod_key, slice(None), dest_group)].sum())
            return v != 0.0
        except Exception:
            return False

    # ── Product-sheet: yyyy-mm block builder ──────────────────────────

    def build_product_block_df(
        cfg: Config,
        scope_value: str,
        product: str,
        months: List[str],
        cur_month_label: str,
        max_dt_t: pd.Timestamp,
        m_agg: pd.Series,
        d_agg: pd.Series,
        dest_group: str,
        fc_product_totals: pd.Series,
    ) -> pd.DataFrame:
        scope_key = (scope_value or "").strip()
        prod_key = str(product)
        cur_p = pd.Period(cur_month_label, freq="M")

        actuals, actuals_ly, yoy, f_yoy, forecast = [], [], [], [], []

        for m in months:
            m_ly = shift_month_label_ly(m)
            mp = pd.Period(m, freq="M")

            if str(m) == str(cur_month_label):
                cutoff_t = pd.to_datetime(max_dt_t)
                cutoff_t1 = cutoff_t - pd.DateOffset(years=1)

                month_start_t = pd.Timestamp(year=cutoff_t.year, month=cutoff_t.month, day=1)
                month_start_t1 = pd.Timestamp(year=cutoff_t1.year, month=cutoff_t1.month, day=1)

                try:
                    s = d_agg.loc[(scope_key, prod_key, dest_group)]
                    v = float(s.loc[month_start_t:cutoff_t].sum()) if hasattr(s, "loc") else 0.0
                except Exception:
                    v = 0.0

                try:
                    s = d_agg.loc[(scope_key, prod_key, dest_group)]
                    vly_mtd = float(s.loc[month_start_t1:cutoff_t1].sum()) if hasattr(s, "loc") else 0.0
                except Exception:
                    vly_mtd = 0.0

                vly_full = get_agg_value(m_agg, (scope_key, prod_key, m_ly, dest_group))

                actuals.append(v)
                actuals_ly.append(vly_full)

                y = 0.0 if (v == 0.0 or vly_mtd == 0.0) else (v / vly_mtd - 1.0)
                yoy.append(y)

                if mp < cur_p:
                    f_yoy.append(y)
                    forecast.append((1.0 + y) * vly_full)
                else:
                    fc_val = get_agg_value(fc_product_totals, (dest_group, prod_key, m))
                    fy = (fc_val / vly_full - 1.0) if (fc_val != 0 and vly_full != 0) else 0.0
                    f_yoy.append(fy)
                    forecast.append(fc_val if fc_val != 0 else "na")
            else:
                v = get_agg_value(m_agg, (scope_key, prod_key, m, dest_group))
                vly = get_agg_value(m_agg, (scope_key, prod_key, m_ly, dest_group))

                actuals.append(v)
                actuals_ly.append(vly)

                y = 0.0 if (v == 0.0 or vly == 0.0) else (v / vly - 1.0)
                yoy.append(y)

                if mp < cur_p:
                    f_yoy.append(y)
                    forecast.append((1.0 + y) * vly)
                else:
                    fc_val = get_agg_value(fc_product_totals, (dest_group, prod_key, m))
                    fy = (fc_val / vly - 1.0) if (fc_val != 0 and vly != 0) else 0.0
                    f_yoy.append(fy)
                    forecast.append(fc_val if fc_val != 0 else "na")

        cols = ["Product name", "", *months]

        return pd.DataFrame(
            [
                [f"Destination: {dest_group}", "", *months],
                ["Actuals", "", *actuals],
                ["Actuals LY", "", *actuals_ly],
                ["YoY %", "", *yoy],
                ["", "", *([None] * len(months))],
                ["F-YoY %", "", *f_yoy],
                ["Forecast", "", *forecast],
            ],
            columns=cols,
        )

    # ── Formatting helpers ────────────────────────────────────────────

    def is_total_value(v) -> bool:
        return isinstance(v, str) and v.startswith("TOTAL")

    def _set_right_border(cell, side: Side):
        b = cell.border or Border()
        cell.border = Border(left=b.left, right=side, top=b.top, bottom=b.bottom)

    def fill_white_rows(ws, start_row: int, end_row: int):
        if end_row < start_row:
            return
        max_col = ws.max_column
        for r in range(start_row, end_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = WHITE_FILL

    def apply_formatting_summary_range(ws, label_header: str, start_row: int, end_row: int):
        if end_row < start_row:
            return
        max_col = ws.max_column

        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.fill = WHITE_FILL

        for c in range(1, max_col + 1):
            hcell = ws.cell(row=start_row, column=c)
            hcell.font = BOLD_FONT
            hcell.border = HEADER_BOTTOM_BORDER

        headers = {ws.cell(row=start_row, column=c).value: c for c in range(1, max_col + 1)}

        rb_names = {"YTD % growth", "FY % growth", "Last 3w % growth"}
        rb_cols = []
        for name in rb_names:
            idx = headers.get(name)
            if idx:
                rb_cols.append(idx)

        if rb_cols:
            for r in range(start_row, end_row + 1):
                for c in rb_cols:
                    _set_right_border(ws.cell(row=r, column=c), THIN)

        label_col_idx = headers.get(label_header)
        if label_col_idx:
            ws.cell(row=start_row, column=label_col_idx).alignment = LEFT_ALIGN
            for r in range(start_row + 1, end_row + 1):
                v = ws.cell(row=r, column=label_col_idx).value
                if not is_total_value(v):
                    ws.cell(row=r, column=label_col_idx).alignment = INDENT_ALIGN

        percent_headers = {"% of Total YTD", "YTD % growth", "FY % growth", "Last 3w % growth"}
        percent_col_idxs = []
        for h, c in headers.items():
            if h is None:
                continue
            hs = str(h).strip()
            if hs in percent_headers or hs.count("-") == 1:
                percent_col_idxs.append(c)
                for r in range(start_row + 1, end_row + 1):
                    ws.cell(row=r, column=c).number_format = PERCENT_FMT

        total_row_idx = None
        for r in range(start_row + 1, end_row + 1):
            if is_total_value(ws.cell(row=r, column=1).value):
                total_row_idx = r
                break

        if total_row_idx:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=total_row_idx, column=c)
                cell.font = BOLD_FONT
                b = cell.border or Border()
                cell.border = Border(left=b.left, right=b.right, bottom=b.bottom, top=THIN)

        for c in range(1, max_col + 1):
            is_percent = c in percent_col_idxs
            max_len = 0
            for r in range(start_row, end_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    s = ""
                elif r == start_row:
                    s = str(v)
                elif is_percent:
                    try:
                        s = f"{float(v) * 100:.1f}%"
                    except Exception:
                        s = "0.0%"
                else:
                    s = str(v)
                max_len = max(max_len, len(s))
            ws.column_dimensions[ws.cell(row=start_row, column=c).column_letter].width = max(8, min(60, max_len + 2))

    # ── write_df_to_ws ────────────────────────────────────────────────

    def write_df_to_ws(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> Tuple[int, int]:
        for j, col in enumerate(df.columns, start=start_col):
            ws.cell(row=start_row, column=j, value=col)
        for i in range(len(df)):
            r = start_row + 1 + i
            for j, col in enumerate(df.columns, start=start_col):
                ws.cell(row=r, column=j, value=df.iloc[i][col])
        end_row = start_row + len(df)
        end_col = start_col + len(df.columns) - 1
        return end_row, end_col

    def write_df_to_ws_no_header(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
        for i in range(len(df)):
            r = start_row + i
            for j, col in enumerate(df.columns, start=start_col):
                ws.cell(row=r, column=j, value=df.iloc[i][col])
        return start_row + len(df) - 1 if len(df) > 0 else start_row

    # ── Summary-sheet destination label ───────────────────────────────

    def add_destination_label_under_table(ws, label_row: int, dest_text: str):
        c1 = ws.cell(row=label_row, column=1, value="Destination:")
        c2 = ws.cell(row=label_row, column=2, value=dest_text)
        for c in (c1, c2):
            c.font = BOLD_FONT
            c.alignment = LEFT_ALIGN
            c.fill = WHITE_FILL
        for col in range(3, ws.max_column + 1):
            ws.cell(row=label_row, column=col).fill = WHITE_FILL

    # ── Calculations block ────────────────────────────────────────────

    def write_product_calcs_from_df(
        ws, summary_df: pd.DataFrame, cfg: Config, product: str,
        start_row: int = 1, start_col: int = 1,
    ) -> int:
        prod_mask = summary_df[cfg.label_col].astype(str) == str(product)
        if not prod_mask.any():
            return start_row

        prod_series = summary_df[prod_mask].iloc[0]
        metrics = [c for c in summary_df.columns if c != cfg.label_col]

        underline_metrics = {"YTD % growth", "FY % growth", "Last 3w % growth"}
        percent_metrics = {"% of Total YTD", "YTD % growth", "FY % growth", "Last 3w % growth"}

        r = start_row
        ws.cell(row=r, column=start_col, value="Calculations").font = BOLD_FONT
        r += 1

        ws.cell(row=r, column=start_col, value="Metric").font = BOLD_FONT
        ws.cell(row=r, column=start_col + 1, value="Value").font = BOLD_FONT
        ws.cell(row=r, column=start_col).border = HEADER_BOTTOM_BORDER
        ws.cell(row=r, column=start_col + 1).border = HEADER_BOTTOM_BORDER
        r += 1

        for m in metrics:
            ms = str(m).strip()
            mcell = ws.cell(row=r, column=start_col, value=m)
            tgt = ws.cell(row=r, column=start_col + 1, value=prod_series[m])

            if ms in percent_metrics or ms.count("-") == 1:
                tgt.number_format = PERCENT_FMT

            if ms in underline_metrics:
                for cell in (mcell, tgt):
                    b = cell.border or Border()
                    cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=THIN)

            r += 1

        return r

    # ── Last 6w block ─────────────────────────────────────────────────

    def write_product_last6w_block(
        ws, scope_value: str, product: str,
        start_row: int, start_col: int,
        w_agg: pd.Series,
        dest_group: Optional[str] = None,
    ) -> int:
        scope_key = (scope_value or "").strip()
        prod_key = str(product)

        try:
            if dest_group is not None:
                sub = w_agg.loc[(scope_key, prod_key)]
                if isinstance(sub.index, pd.MultiIndex):
                    sub_dest = sub.xs(dest_group, level=-1)
                    keys = sorted(sub_dest.index.astype(str).unique().tolist(), key=parse_week_key)
                else:
                    keys = []
            else:
                wk = w_agg.loc[(scope_key, prod_key)]
                keys = sorted(wk.index.astype(str).unique().tolist(), key=parse_week_key)
        except Exception:
            return start_row

        if not keys:
            return start_row

        keys = keys[:-1]
        if not keys:
            return start_row

        last6 = keys[-6:] if len(keys) >= 6 else keys

        actuals = []
        actuals_ly = []
        yoy = []

        for k in last6:
            if dest_group is not None:
                v = get_agg_value(w_agg, (scope_key, prod_key, str(k), dest_group))
                y_val, w_val = parse_week_key(str(k))
                k_ly = f"{y_val - 1}-{str(w_val).zfill(2)}" if y_val != -1 else "0000-00"
                vly = get_agg_value(w_agg, (scope_key, prod_key, k_ly, dest_group))
            else:
                v = get_agg_value(w_agg, (scope_key, prod_key, str(k)))
                y_val, w_val = parse_week_key(str(k))
                k_ly = f"{y_val - 1}-{str(w_val).zfill(2)}" if y_val != -1 else "0000-00"
                vly = get_agg_value(w_agg, (scope_key, prod_key, k_ly))

            actuals.append(v)
            actuals_ly.append(vly)

            yv = 0.0 if (v == 0.0 or vly == 0.0) else (v / vly - 1.0)
            yoy.append(yv)

        r = start_row

        ws.cell(row=r, column=start_col, value="Last 6w % growth:").font = BOLD_FONT
        r += 1

        ws.cell(row=r, column=start_col, value="").border = HEADER_BOTTOM_BORDER
        for i, k in enumerate(last6, start=1):
            c = start_col + i
            cell = ws.cell(row=r, column=c, value=str(k))
            cell.font = BOLD_FONT
            cell.border = HEADER_BOTTOM_BORDER
            cell.alignment = LEFT_ALIGN
        r += 1

        labels = ["Actuals", "Actuals LY", "YoY %"]
        rows = [actuals, actuals_ly, yoy]

        for lbl, vals in zip(labels, rows):
            ws.cell(row=r, column=start_col, value=lbl).font = BOLD_FONT
            ws.cell(row=r, column=start_col).alignment = LEFT_ALIGN
            for i, v in enumerate(vals, start=1):
                c = start_col + i
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = RIGHT_ALIGN
                if lbl == "YoY %":
                    cell.number_format = PERCENT_FMT
            r += 1

        return r

    # ── Top 5 shops block ─────────────────────────────────────────────

    def write_top5_shops_block(
        ws,
        cfg: Config,
        scope_value: str,
        product: str,
        dest_group: str,
        months: List[str],
        cur_month_label: str,
        start_row: int,
        shop_month_agg: pd.Series,
        fc_shop_lookup: pd.Series,
    ) -> int:
        scope_key = (scope_value or "").strip()
        prod_key = str(product)

        try:
            shop_data = shop_month_agg.loc[(scope_key, prod_key, dest_group)]
            shop_totals = shop_data.groupby(level=0).sum().sort_values(ascending=False)
            top5 = shop_totals.head(5).index.tolist()
        except Exception:
            return start_row, []

        if not top5:
            return start_row, []

        month_col_map = {m: i for i, m in enumerate(months, start=3)}
        blue_start = month_col_map.get(str(cur_month_label).strip())

        max_c = 2 + len(months)
        for gap_c in range(1, max_c + 1):
            ws.cell(row=start_row - 1, column=gap_c).fill = WHITE_FILL

        cur_p = pd.Period(cur_month_label, freq="M")

        def _get_shop_value(shop_keys, month):
            v = 0.0
            for sk in shop_keys:
                try:
                    v += float(shop_data.loc[(sk, month)])
                except Exception:
                    pass
            return v

        def _get_forecast_value(shop_keys, month):
            v = 0.0
            for sk in shop_keys:
                try:
                    v += float(fc_shop_lookup.loc[(sk, dest_group, prod_key, month)])
                except Exception:
                    pass
            return v

        # ── FIX 1: no_blue_fill suppresses forecast-input blue styling.
        # ── FIX 2: fc_override_key lets "Other [shoptype]" pull its own
        #    forecast row from the CSV rather than summing residual shops.
        def _write_shop_group(r, label, shop_keys, fc_override_keys=None,
                              no_blue_fill=False, skip_fyoy=False):
            """Write 3-4 rows for a shop: Forecast formula, Actuals LY, [F-YoY %], blank.

            fc_override_keys – list of keys for fc_shop_lookup on forecast months
                               instead of iterating shop_keys.
            no_blue_fill     – skip the blue FORECAST_FILL / ALL_THIN border on
                               forecast cells.
            skip_fyoy        – omit the F-YoY % row entirely (for Other [shoptype]
                               where the ratio is meaningless).
            """
            forecast_row = r  # capture for caller
            # ── Row 1: shop name / actuals or forecast formula ──
            ws.cell(row=r, column=1, value=label).fill = WHITE_FILL
            ws.cell(row=r, column=2).fill = WHITE_FILL
            actuals_vals = {}
            for i, m in enumerate(months, start=3):
                mp = pd.Period(m, freq="M")
                if mp >= cur_p:
                    if fc_override_keys is not None:
                        fc_val = _get_forecast_value(fc_override_keys, m)
                    else:
                        fc_val = _get_forecast_value(shop_keys, m)
                    actuals_vals[m] = fc_val
                    if fc_val != 0:
                        if skip_fyoy:
                            # No F-YoY% row — write raw forecast value
                            cell = ws.cell(row=r, column=i, value=fc_val)
                        else:
                            m_ly = shift_month_label_ly(m)
                            ly_val = _get_shop_value(shop_keys, m_ly)
                            if ly_val == 0:
                                # Actuals LY is 0 — formula would give 0 regardless.
                                # Write the absolute forecast value instead.
                                cell = ws.cell(row=r, column=i, value=fc_val)
                            else:
                                # Write formula: =IFERROR((1+F-YoY%)*Actuals LY, 0)
                                cl = get_column_letter(i)
                                ly_ref = f"{cl}{r + 1}"    # Actuals LY row
                                fyoy_ref = f"{cl}{r + 2}"  # F-YoY % row
                                cell = ws.cell(row=r, column=i,
                                               value=f"=IFERROR((1+{fyoy_ref})*{ly_ref},0)")
                    else:
                        cell = ws.cell(row=r, column=i, value="na")
                else:
                    v = _get_shop_value(shop_keys, m)
                    actuals_vals[m] = v
                    cell = ws.cell(row=r, column=i, value=v)
                # Blue highlight only for individual named shops, not the Other row
                if (not no_blue_fill
                        and blue_start is not None
                        and blue_start <= i <= blue_start + 5):
                    cell.border = ALL_THIN
                    cell.fill = FORECAST_FILL
                else:
                    cell.fill = WHITE_FILL
                cell.alignment = RIGHT_ALIGN
                cell.number_format = INTEGER_FMT
            r += 1

            # ── Row 2: Actuals LY (always actuals residual) ──
            ws.cell(row=r, column=1, value="Actuals LY").fill = WHITE_FILL
            ws.cell(row=r, column=2).fill = WHITE_FILL
            ly_vals = {}
            for i, m in enumerate(months, start=3):
                m_ly = shift_month_label_ly(m)
                v = _get_shop_value(shop_keys, m_ly)
                ly_vals[m] = v
                cell = ws.cell(row=r, column=i, value=v)
                cell.fill = WHITE_FILL
                cell.alignment = RIGHT_ALIGN
            r += 1

            # ── Row 3: F-YoY % (skipped for Other [shoptype]) ──
            if not skip_fyoy:
                ws.cell(row=r, column=1, value="F-YoY %").fill = WHITE_FILL
                ws.cell(row=r, column=2).fill = WHITE_FILL
                for i, m in enumerate(months, start=3):
                    act = actuals_vals.get(m, 0.0)
                    act_ly = ly_vals.get(m, 0.0)
                    yoy_val = (act / act_ly - 1.0) if (act != 0.0 and act_ly != 0.0) else 0.0
                    cell = ws.cell(row=r, column=i, value=yoy_val)
                    cell.fill = WHITE_FILL
                    cell.number_format = PERCENT_FMT
                    cell.alignment = RIGHT_ALIGN
                for c in range(1, max_c + 1):
                    cell = ws.cell(row=r, column=c)
                    b = cell.border or Border()
                    cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=THIN)
                r += 1

            # ── Row 4: Blank ──
            for c in range(1, max_c + 1):
                ws.cell(row=r, column=c).fill = WHITE_FILL
            r += 1

            return r, forecast_row

        r = start_row
        shop_forecast_rows = []

        # ── Write Top-5 shops from actuals ──
        for shop in top5:
            r, frow = _write_shop_group(r, str(shop), [shop])
            shop_forecast_rows.append(frow)

        # ── FIX 2: additional shops present in forecast CSV but not in top5 ──
        # These shops have a forecast row in seperate_forecasts_combined but did
        # not rank in the top-5 actuals.  Show them individually (no blue fill
        # because they are not direct user-input cells).
        try:
            fc_sub = fc_shop_lookup.xs(
                key=(dest_group, prod_key),
                level=("destination", "product"),
            )
            # fc_sub is indexed by (forecasted_shop, period)
            fc_shop_sums = fc_sub.groupby(level=0).sum()
            extra_fc_shops = [
                s for s in fc_shop_sums[fc_shop_sums > 0].index
                if s not in top5
                and not str(s).startswith("Other ")
                and str(s) != str(scope_value)   # exclude the shoptype-level rollup row
            ]
        except Exception:
            extra_fc_shops = []

        for shop in extra_fc_shops:
            r, frow = _write_shop_group(r, str(shop), [shop])
            shop_forecast_rows.append(frow)

        # ── "Other [shoptype]" residual row – no blue fill (FIX 1) ──
        # Also includes shoptype-level forecasts (forecasted_shop == scope_value)
        all_shops = shop_totals.index.tolist()
        # Exclude extra_fc_shops from the residual so they don't double-count
        shown_shops = set(top5) | set(extra_fc_shops)
        other_shops = [s for s in all_shops if s not in shown_shops]
        if other_shops:
            other_fc_key = f"Other {scope_value}".strip()
            r, frow = _write_shop_group(
                r, other_fc_key, other_shops,
                fc_override_keys=[other_fc_key, scope_value.strip()],
                skip_fyoy=True,
            )
            shop_forecast_rows.append(frow)

        return r, shop_forecast_rows

    # ── Product-sheet formatting ──────────────────────────────────────

    def apply_formatting_product(ws, cur_month_label: str):
        max_row = ws.max_row
        max_col = ws.max_column

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.fill = WHITE_FILL

        for r in range(1, max_row + 1):
            v = str(ws.cell(row=r, column=1).value or "").strip()
            if v.startswith("Destination:"):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = BOLD_FONT
                    cell.border = HEADER_BOTTOM_BORDER
                    cell.alignment = LEFT_ALIGN

        for r in range(1, max_row + 1):
            v = str(ws.cell(row=r, column=1).value or "").strip()
            if v in ("YoY %", "F-YoY %"):
                for c in range(3, max_col + 1):
                    ws.cell(row=r, column=c).number_format = PERCENT_FMT
                ws.cell(row=r, column=1).font = BOLD_FONT

        header_r = None
        for r in range(1, max_row + 1):
            v = str(ws.cell(row=r, column=1).value or "").strip()
            if v.startswith("Destination:"):
                header_r = r
                break

        month_to_col = {}
        if header_r is not None:
            for c in range(3, max_col + 1):
                h = ws.cell(row=header_r, column=c).value
                if h is not None:
                    month_to_col[str(h).strip()] = c

        start_c = month_to_col.get(str(cur_month_label).strip())
        if start_c is not None:
            end_c = min(max_col, start_c + 5)

            for r in range(1, max_row + 1):
                v = str(ws.cell(row=r, column=1).value or "").strip()
                if v == "Forecast":
                    for c in range(start_c, end_c + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = ALL_THIN
                        cell.fill = FORECAST_FILL

        for r in range(1, max_row + 1):
            v = str(ws.cell(row=r, column=1).value or "").strip()
            if v == "Forecast" or v.startswith("Destination:"):
                ws.cell(row=r, column=1).font = BOLD_FONT

        for r in range(1, max_row + 1):
            v = str(ws.cell(row=r, column=1).value or "").strip()
            if v == "Forecast":
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    b = cell.border or Border()
                    cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=DOUBLE)
                    if isinstance(cell.value, (int, float, np.integer, np.floating)):
                        cell.number_format = INTEGER_FMT

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str) and cell.value.strip().lower() == "na":
                    cell.alignment = RIGHT_ALIGN

        def display_len(cell) -> int:
            v = cell.value
            if v is None:
                return 0
            if isinstance(v, str):
                return len(v)
            if cell.number_format == PERCENT_FMT and isinstance(v, (int, float, np.integer, np.floating)):
                try:
                    return len(f"{float(v) * 100:.1f}%")
                except Exception:
                    return len(str(v))
            if isinstance(v, (int, float, np.integer, np.floating)):
                try:
                    return len(f"{float(v):,.0f}")
                except Exception:
                    return len(str(v))
            return len(str(v))

        max_len = 0
        for r in range(1, max_row + 1):
            max_len = max(max_len, display_len(ws.cell(row=r, column=1)))
        ws.column_dimensions[get_column_letter(1)].width = max(14, min(60, max_len + 2))

        max_len = 0
        for r in range(1, max_row + 1):
            max_len = max(max_len, display_len(ws.cell(row=r, column=2)))
        ws.column_dimensions[get_column_letter(2)].width = max(10, min(35, max_len + 2))

        ref_r = header_r or 1
        for c in range(3, max_col + 1):
            header = ws.cell(row=ref_r, column=c).value
            header_len = len(str(header)) if header is not None else 0
            ws.column_dimensions[get_column_letter(c)].width = max(10, min(18, header_len + 2))

    # ── Main Excel builder ────────────────────────────────────────────

    def build_excel_bytes(
        sheet_name: str,
        df_raw: pd.DataFrame,
        cfg: Config,
        scope_value: str,
        scope_col: str,
    ) -> bytes:
        df_clean = clean_input(df_raw, cfg)

        scope_value = (scope_value or "").strip()
        scope_mask = mask_for_scope(df_clean, scope_col, scope_value)
        df_scope = df_clean[scope_mask] if scope_value else df_clean

        t = int(df_scope[cfg.date_col].dt.year.max())
        max_dt_t = df_scope.loc[df_scope[cfg.date_col].dt.year == t, cfg.date_col].max()

        months = month_labels_m2_to_p11(max_dt_t)
        cur_month_label = str(max_dt_t.to_period("M"))

        _, out_eu = compute_review_table(df_raw, scope_value, scope_col, cfg, dest_group_filter=DEST_EUROW)
        _, out_us = compute_review_table(df_raw, scope_value, scope_col, cfg, dest_group_filter=DEST_USCA)

        regional_summaries = {DEST_EUROW: out_eu, DEST_USCA: out_us}

        top_products = (
            out_eu.loc[~out_eu[cfg.label_col].astype(str).str.startswith(("Other products", "TOTAL")), cfg.label_col]
            .astype(str)
            .head(cfg.top_n)
            .tolist()
        )

        dfc = df_clean.copy()
        dfc["_scope"] = dfc[scope_col].astype(str)
        dfc[cfg.product_col] = dfc[cfg.product_col].astype(str)
        dfc[cfg.week_col] = dfc[cfg.week_col].astype(str)

        w_agg_dest = (
            dfc.groupby(["_scope", cfg.product_col, cfg.week_col, "_dest_group"], dropna=False)[cfg.actuals_col]
            .sum()
            .astype(float)
        )

        m_agg_dest = (
            dfc.groupby(["_scope", cfg.product_col, "_month", "_dest_group"], dropna=False)[cfg.actuals_col]
            .sum()
            .astype(float)
        )

        d_agg_dest = (
            dfc.groupby(["_scope", cfg.product_col, "_dest_group", cfg.date_col], dropna=False)[cfg.actuals_col]
            .sum()
            .astype(float)
        ).reorder_levels([0, 1, 2, 3]).sort_index()

        has_shop_col = cfg.shop_col in dfc.columns
        shop_month_agg = None
        if has_shop_col:
            dfc[cfg.shop_col] = dfc[cfg.shop_col].astype(str)
            shop_month_agg = (
                dfc
                .groupby(["_scope", cfg.product_col, "_dest_group", cfg.shop_col, "_month"], dropna=False)[cfg.actuals_col]
                .sum()
                .astype(float)
            )

        try:
            fc_obj = S3.get_object(Bucket=cfg.bucket, Key=cfg.forecast_key)
            fc_df = pd.read_csv(BytesIO(fc_obj["Body"].read()))
            fc_df["forecast"] = pd.to_numeric(fc_df["forecast"], errors="coerce").fillna(0.0)
        except Exception:
            fc_df = pd.DataFrame(columns=["forecasted_shop", "shoptype", "destination", "product", "period", "forecast"])

        if "shoptype" in fc_df.columns:
            fc_df = fc_df[fc_df["shoptype"].astype(str) == str(scope_value)].copy()

        fc_shop_lookup = (
            fc_df.groupby(["forecasted_shop", "destination", "product", "period"], dropna=False)["forecast"]
            .sum().astype(float)
        )
        fc_product_totals = (
            fc_df.groupby(["destination", "product", "period"], dropna=False)["forecast"]
            .sum().astype(float)
        )

        buf = BytesIO()
        used = set()

        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            summary_sheet = dedupe_sheet_name(safe_sheet_base(sheet_name), used)
            wb = writer.book
            ws = wb.create_sheet(summary_sheet)
            if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
                try:
                    wb.remove(wb["Sheet"])
                except Exception:
                    pass

            r1 = 1
            end_r_1, _ = write_df_to_ws(ws, out_eu, start_row=r1, start_col=1)
            add_destination_label_under_table(ws, end_r_1 + 2, DEST_EUROW)
            fill_white_rows(ws, end_r_1 + 1, end_r_1 + 4)
            apply_formatting_summary_range(ws, cfg.label_col, start_row=r1, end_row=end_r_1)

            r2 = end_r_1 + 5
            end_r_2, _ = write_df_to_ws(ws, out_us, start_row=r2, start_col=1)
            add_destination_label_under_table(ws, end_r_2 + 2, DEST_USCA)
            fill_white_rows(ws, end_r_2 + 1, end_r_2 + 3)
            apply_formatting_summary_range(ws, cfg.label_col, start_row=r2, end_row=end_r_2)

            scope_key = scope_value.strip()

            for prod in top_products:
                sname = dedupe_sheet_name(safe_sheet_base(prod), used)
                pws = wb.create_sheet(sname)
                writer.sheets[sname] = pws

                current_row = 1
                first_region = True
                top5_tasks = []

                for dest in (DEST_EUROW, DEST_USCA):
                    if not region_has_actuals(m_agg_dest, scope_key, str(prod), dest):
                        continue

                    if not first_region:
                        fill_white_rows(pws, current_row, current_row + 14)
                        current_row += 15
                    first_region = False

                    calcs_end = write_product_calcs_from_df(
                        pws, regional_summaries[dest], cfg, prod,
                        start_row=current_row, start_col=1,
                    )

                    l6w_end = write_product_last6w_block(
                        pws, scope_value, prod,
                        start_row=current_row, start_col=4,
                        w_agg=w_agg_dest, dest_group=dest,
                    )

                    top_block_end = max(calcs_end, l6w_end)
                    yyyy_start = top_block_end + 1

                    block_df = build_product_block_df(
                        cfg=cfg,
                        scope_value=scope_value,
                        product=prod,
                        months=months,
                        cur_month_label=cur_month_label,
                        max_dt_t=max_dt_t,
                        m_agg=m_agg_dest,
                        d_agg=d_agg_dest,
                        dest_group=dest,
                        fc_product_totals=fc_product_totals,
                    )
                    yyyy_end = write_df_to_ws_no_header(pws, block_df, start_row=yyyy_start)

                    top5_start = yyyy_end + 2
                    if has_shop_col and shop_month_agg is not None:
                        top5_tasks.append((dest, top5_start, yyyy_start))
                        current_row = top5_start + 25
                    else:
                        current_row = yyyy_end + 2

                apply_formatting_product(pws, cur_month_label)

                cur_p = pd.Period(cur_month_label, freq="M")
                for dest, t5row, ystart in top5_tasks:
                    _, shop_fc_rows = write_top5_shops_block(
                        pws, cfg, scope_value, prod, dest,
                        months, cur_month_label, t5row, shop_month_agg,
                        fc_shop_lookup,
                    )

                    if shop_fc_rows:
                        product_forecast_row = ystart + 6
                        product_actuals_ly_row = ystart + 2
                        product_fyoy_row = ystart + 5

                        for i, m in enumerate(months, start=3):
                            mp = pd.Period(m, freq="M")
                            if mp >= cur_p:
                                cl = get_column_letter(i)
                                # Forecast = SUM of individual shop forecast rows
                                refs = ",".join(f"{cl}{sr}" for sr in shop_fc_rows)
                                fc_cell = pws.cell(row=product_forecast_row, column=i,
                                                   value=f"=SUM({refs})")
                                # Remove blue fill — derived total, not user input
                                fc_cell.fill = WHITE_FILL
                                # F-YoY% = IFERROR(Forecast / Actuals LY - 1, 0)
                                fc_ref = f"{cl}{product_forecast_row}"
                                ly_ref = f"{cl}{product_actuals_ly_row}"
                                pws.cell(row=product_fyoy_row, column=i,
                                         value=f"=IFERROR({fc_ref}/{ly_ref}-1,0)")

        return buf.getvalue()

    # ── S3 I/O ────────────────────────────────────────────────────────

    def read_csv_from_s3(cfg: Config) -> pd.DataFrame:
        obj = S3.get_object(Bucket=cfg.bucket, Key=cfg.in_key)
        return pd.read_csv(obj["Body"])


# ── Entrypoint ────────────────────────────────────────────────────

    mode, scope_value, scope_col = parse_scope_shoptype(event, CFG)
    if not scope_col:
        raise ValueError("Invalid command. Use: 'shoptype <type>'.")

    df = read_csv_from_s3(CFG)

    sheet_name = safe_sheet_base(f"{mode} {scope_value}")
    file_label = f"{mode}_{scope_value}".strip()
    excel_bytes = build_excel_bytes(
        sheet_name=sheet_name,
        df_raw=df,
        cfg=CFG,
        scope_value=scope_value,
        scope_col=scope_col,
    )

    slack_upload_xlsx(
        excel_bytes,
        filename=f"review_{file_label}.xlsx",
        message=f"Finished reviewing {sheet_name}",
    )

    return {"ok": True, "mode": mode, "scope_value": scope_value}