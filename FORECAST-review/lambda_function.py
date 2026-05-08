import os
import json
import re
import urllib.request
import urllib.parse
import urllib.error
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import List, Tuple, Dict, Any, Optional

import boto3
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter


COMBINED_KEY = "Forecast/Seperate Forecasts/seperate_forecasts_combined.csv"


def lambda_handler(event, context):
    @dataclass(frozen=True)
    class Config:
        bucket: str = "bi-automations"
        in_key: str = "Forecast/actuals.csv"
        out_key_prefix: str = "Forecast/"

        date_col: str = "fulldate"
        actuals_col: str = "actuals"
        shop_col: str = "forecasted_shop"
        product_col: str = "forecast_product"
        label_col: str = "TOP Products"
        week_col: str = "week"
        format_col: str = "format_clean"
        destination_col: str = "destination_region"
        revenue_col: str = "revenue"

        excluded_products: Tuple[str, ...] = ("EXCLUDED PRODUCT", "NEW PRODUCT")
        top_n: int = 15

    CFG = Config()
    S3 = boto3.client("s3")
    _WEEK_RE = re.compile(r"^(?P<y>\d{4})-(?P<w>\d{2})$")

    # ----------------------------
    # Read combined forecasts lookup
    # ----------------------------
    def read_combined_forecasts(bucket, key):
        """Read seperate_forecasts_combined.csv.
        Returns dict: (forecasted_shop, product) -> {(format, frame, period): forecast_value}
        """
        try:
            raw = S3.get_object(Bucket=bucket, Key=key)["Body"].read().decode("utf-8", "replace")
        except Exception:
            return {}
        df = pd.read_csv(StringIO(raw))
        lookup = {}
        for _, row in df.iterrows():
            shop = str(row.get("forecasted_shop", "")).strip()
            product = str(row.get("product", "")).strip()
            fmt = str(row.get("format", "")).strip()
            frame = str(row.get("frame", "")).strip()
            period = str(row.get("period", "")).strip()
            try:
                val = float(row.get("forecast", 0))
            except (ValueError, TypeError):
                val = 0.0
            sp_key = (shop, product)
            fpp_key = (fmt, frame, period)
            lookup.setdefault(sp_key, {})
            lookup[sp_key][fpp_key] = lookup[sp_key].get(fpp_key, 0.0) + val
        return lookup

    # ----------------------------
    # Robust request parsing (SHOP ONLY)
    # ----------------------------
    def _strip_leading_keyword(val: str, keyword: str) -> str:
        if not isinstance(val, str):
            return ""
        s = val.strip()
        kw = keyword.strip().lower()
        if not s:
            return ""
        if s.lower() == kw:
            return ""
        if s.lower().startswith(kw + " "):
            return s[len(keyword) :].strip()
        return s

    def parse_shop_request(evt: dict) -> str:
        text = (evt.get("text") or "").strip()
        if text:
            parts = text.split(None, 1)
            cmd = parts[0].lower()
            val = parts[1].strip() if len(parts) > 1 else ""
            if cmd == "shop":
                return val

        if "shop" in evt:
            return _strip_leading_keyword(str(evt.get("shop") or ""), "shop")

        return ""

    # ----------------------------
    # Helpers
    # ----------------------------
    def to_num(s: pd.Series) -> pd.Series:
        return pd.to_numeric(s, errors="coerce").fillna(0).astype(float)

    def safe_growth(curr: pd.Series, prev: pd.Series) -> pd.Series:
        curr = to_num(curr)
        prev = to_num(prev)
        out = np.where(prev != 0, (curr / prev) - 1.0, 0.0)
        out = np.where(np.isfinite(out), out, 0.0)
        return pd.Series(out, index=curr.index)

    def normalize_ratio(s: pd.Series) -> pd.Series:
        s = to_num(s)
        mask = s.abs() > 5
        s.loc[mask] = s.loc[mask] / 100.0
        return s

    def parse_week_key(s: str) -> Tuple[int, int]:
        if not isinstance(s, str):
            return (-1, -1)
        m = _WEEK_RE.match(s.strip())
        if not m:
            return (-1, -1)
        return (int(m.group("y")), int(m.group("w")))

    def shift_week_year(week_key: str, new_year: int) -> str:
        if not isinstance(week_key, str) or "-" not in week_key:
            return f"{new_year}-00"
        _, ww = week_key.split("-", 1)
        return f"{new_year}-{ww.zfill(2)}"

    def last_3_weeks_keys_any(df: pd.DataFrame, t: int, week_col: str) -> List[str]:
        d = df[df[week_col].astype(str).str.startswith(f"{t}-")]
        keys = sorted(d[week_col].dropna().astype(str).unique().tolist(), key=parse_week_key)
        if not keys:
            return []
        keys = keys[:-1]
        return keys[-3:] if len(keys) >= 3 else keys

    def last_3_full_months_labels(max_dt_t: pd.Timestamp) -> List[str]:
        month_start = pd.Timestamp(year=max_dt_t.year, month=max_dt_t.month, day=1)
        last_full_month_end = month_start - pd.Timedelta(days=1)
        last_full_month = last_full_month_end.to_period("M")
        months = pd.period_range(end=last_full_month, periods=3, freq="M")
        return [str(p) for p in months]

    def shift_month_label_ly(m: str) -> str:
        return f"{int(m[:4]) - 1}{m[4:]}"

    def month_labels_m2_to_p11(max_dt_t: pd.Timestamp) -> List[str]:
        cur = max_dt_t.to_period("M")
        start = cur - 2
        end = cur + 11
        months = pd.period_range(start=start, end=end, freq="M")
        return [str(p) for p in months]

    def clean_input(df: pd.DataFrame, cfg: Config) -> pd.DataFrame:
        df = df.copy()
        df = df[~df[cfg.product_col].astype(str).isin(cfg.excluded_products)]
        df[cfg.date_col] = pd.to_datetime(df[cfg.date_col], errors="coerce")
        df = df[df[cfg.date_col].notna()].copy()
        df[cfg.actuals_col] = to_num(df[cfg.actuals_col])
        if cfg.revenue_col in df.columns:
            df[cfg.revenue_col] = to_num(df[cfg.revenue_col])
        df["_month"] = df[cfg.date_col].dt.to_period("M").astype(str)

        df[cfg.shop_col] = df[cfg.shop_col].astype(str)
        df[cfg.product_col] = df[cfg.product_col].astype(str)
        df[cfg.week_col] = df[cfg.week_col].astype(str)

        if cfg.format_col in df.columns:
            df[cfg.format_col] = df[cfg.format_col].astype(str)
        if "frame_color" in df.columns:
            df["frame_color"] = df["frame_color"].astype(str)
        else:
            df["frame_color"] = ""

        if cfg.destination_col in df.columns:
            df[cfg.destination_col] = df[cfg.destination_col].astype(str)
        return df

    def ensure_col(df: pd.DataFrame, col: str) -> pd.Series:
        if col in df.columns:
            return to_num(df[col])
        return pd.Series(0.0, index=df.index)

    def safe_sheet_base(name: str) -> str:
        bad = r'[:\\/?*\[\]]'
        s = re.sub(bad, " ", str(name)).strip()
        s = re.sub(r"\s+", " ", s)
        return (s[:31] if s else "Sheet")

    def dedupe_sheet_name(base: str, used: set) -> str:
        name = base
        i = 2
        while name in used:
            suffix = f" {i}"
            name = (base[: (31 - len(suffix))] + suffix) if len(base) + len(suffix) > 31 else (base + suffix)
            i += 1
        used.add(name)
        return name

    # ----------------------------
    # Pre-aggregation (groupby once)
    # ----------------------------
    def build_aggregates(df_clean: pd.DataFrame, cfg: Config) -> Dict[str, Any]:
        m_agg = (
            df_clean.groupby([cfg.shop_col, cfg.product_col, "_month"], dropna=False)[cfg.actuals_col]
            .sum()
            .astype(float)
        )
        d_agg = (
            df_clean.groupby([cfg.shop_col, cfg.product_col, cfg.date_col], dropna=False)[cfg.actuals_col]
            .sum()
            .astype(float)
        )
        w_agg = (
            df_clean.groupby([cfg.shop_col, cfg.product_col, cfg.week_col], dropna=False)[cfg.actuals_col]
            .sum()
            .astype(float)
        )
        fmt_agg = None
        if cfg.format_col in df_clean.columns:
            fmt_agg = (
                df_clean.groupby([cfg.shop_col, cfg.product_col, cfg.format_col, "frame_color", "_month"], dropna=False)[
                    cfg.actuals_col
                ]
                .sum()
                .astype(float)
            )
        dest_agg = None
        if cfg.destination_col in df_clean.columns:
            dest_agg = (
                df_clean.groupby([cfg.shop_col, cfg.destination_col], dropna=False)[cfg.actuals_col]
                .sum()
                .astype(float)
            )
        return {"m_agg": m_agg, "d_agg": d_agg, "w_agg": w_agg, "fmt_agg": fmt_agg, "dest_agg": dest_agg}

    def get_agg_value(agg: pd.Series, key: tuple) -> float:
        try:
            return float(agg.loc[key])
        except Exception:
            return 0.0

    # ----------------------------
    # Summary table
    # ----------------------------
    def mask_for_shop(df: pd.DataFrame, cfg: Config, shop: str) -> pd.Series:
        if shop:
            return df[cfg.shop_col].astype(str) == str(shop)
        return pd.Series(True, index=df.index)

    def agg_sum(df: pd.DataFrame, cfg: Config, mask: pd.Series, label: str) -> pd.DataFrame:
        return (
            df.loc[mask]
            .groupby(cfg.product_col, as_index=False)[cfg.actuals_col]
            .sum()
            .rename(columns={cfg.product_col: cfg.label_col, cfg.actuals_col: label})
        )

    def compute_review_table(df_raw: pd.DataFrame, shop: str, cfg: Config) -> Tuple[int, pd.DataFrame]:
        df = clean_input(df_raw, cfg)
        shop = (shop or "").strip()

        shop_mask = mask_for_shop(df, cfg, shop)
        df_scope = df.loc[shop_mask] if shop else df

        t = int(df_scope[cfg.date_col].dt.year.max())
        max_dt_t = df_scope.loc[df_scope[cfg.date_col].dt.year == t, cfg.date_col].max()
        cutoff_t1 = max_dt_t - pd.DateOffset(years=1)

        base_t = shop_mask & (df[cfg.date_col].dt.year == t) & (df[cfg.date_col] <= max_dt_t)
        base_t1_ytd = shop_mask & (df[cfg.date_col].dt.year == t - 1) & (df[cfg.date_col] <= cutoff_t1)
        base_t1_fy = shop_mask & (df[cfg.date_col].dt.year == t - 1)
        base_t2_fy = shop_mask & (df[cfg.date_col].dt.year == t - 2)

        out = (
            agg_sum(df, cfg, base_t, f"YTD {t}")
            .merge(agg_sum(df, cfg, base_t1_ytd, f"YTD {t-1}"), on=cfg.label_col, how="outer")
            .merge(agg_sum(df, cfg, base_t1_fy, f"FY {t-1}"), on=cfg.label_col, how="outer")
            .merge(agg_sum(df, cfg, base_t2_fy, f"FY {t-2}"), on=cfg.label_col, how="outer")
            .fillna(0)
        )

        last3w_t_keys = last_3_weeks_keys_any(df, t, cfg.week_col)
        last3w_t1_keys = [shift_week_year(k, t - 1) for k in last3w_t_keys]

        out = (
            out.merge(
                agg_sum(df, cfg, shop_mask & df[cfg.week_col].astype(str).isin(last3w_t_keys), f"Last 3W {t}"),
                on=cfg.label_col,
                how="outer",
            )
            .merge(
                agg_sum(df, cfg, shop_mask & df[cfg.week_col].astype(str).isin(last3w_t1_keys), f"Last 3W {t-1}"),
                on=cfg.label_col,
                how="outer",
            )
            .fillna(0)
        )

        last3m_t = last_3_full_months_labels(max_dt_t)
        last3m_t1 = [shift_month_label_ly(m) for m in last3m_t]

        for m_t, m_t1 in zip(last3m_t, last3m_t1):
            msum_t = agg_sum(df, cfg, shop_mask & (df["_month"] == m_t), "_mt")
            msum_t1 = agg_sum(df, cfg, shop_mask & (df["_month"] == m_t1), "_mt1")
            tmp = msum_t.merge(msum_t1, on=cfg.label_col, how="outer").fillna(0)
            tmp[m_t] = safe_growth(tmp["_mt"], tmp["_mt1"])
            out = out.merge(tmp[[cfg.label_col, m_t]], on=cfg.label_col, how="outer").fillna(0)

        out["YTD % Growth"] = safe_growth(ensure_col(out, f"YTD {t}"), ensure_col(out, f"YTD {t-1}"))
        out["FY % Growth"] = safe_growth(ensure_col(out, f"FY {t-1}"), ensure_col(out, f"FY {t-2}"))
        out["Last 3W % Growth"] = safe_growth(ensure_col(out, f"Last 3W {t}"), ensure_col(out, f"Last 3W {t-1}"))
        out["% of Total YTD"] = 0.0

        out = out.sort_values(by=f"YTD {t}", ascending=False).reset_index(drop=True)

        out["_rank_ytd"] = np.arange(1, len(out) + 1, dtype=int)
        fy_col = f"FY {t-1}"
        out["_fy_t1"] = ensure_col(out, fy_col)

        keep_mask = (out["_rank_ytd"] <= cfg.top_n) & (out["_fy_t1"] >= 1000.0)

        helper_cols = {cfg.label_col, "_rank_ytd", "_fy_t1"}
        numeric_cols = [c for c in out.columns if c not in helper_cols and pd.api.types.is_numeric_dtype(out[c])]

        if keep_mask.sum() == 0:
            combined = out.drop(columns=["_rank_ytd", "_fy_t1"], errors="ignore")
        else:
            top_rows = out.loc[keep_mask].copy()
            other_rows = out.loc[~keep_mask].copy()

            include_other = False
            other_df = None

            if not other_rows.empty:
                other_agg = other_rows[numeric_cols].sum(numeric_only=True).to_dict()
            else:
                other_agg = {c: 0.0 for c in numeric_cols}

            if any(float(other_agg.get(c, 0.0) or 0.0) != 0.0 for c in numeric_cols):
                include_other = True
                other_agg[cfg.label_col] = "Other products"
                other_df = pd.DataFrame([other_agg])

            combined = top_rows.drop(columns=["_rank_ytd", "_fy_t1"], errors="ignore")
            if include_other and other_df is not None:
                combined = pd.concat([combined, other_df], ignore_index=True)

        total_vals = combined[numeric_cols].sum(numeric_only=True).to_dict()
        total_label = f"TOTAL {shop}".strip()
        total_row = pd.DataFrame([{cfg.label_col: total_label, **total_vals}])

        out_final = pd.concat([combined, total_row], ignore_index=True)

        out_final["YTD % Growth"] = safe_growth(ensure_col(out_final, f"YTD {t}"), ensure_col(out_final, f"YTD {t-1}"))
        out_final["FY % Growth"] = safe_growth(ensure_col(out_final, f"FY {t-1}"), ensure_col(out_final, f"FY {t-2}"))
        out_final["Last 3W % Growth"] = safe_growth(
            ensure_col(out_final, f"Last 3W {t}"), ensure_col(out_final, f"Last 3W {t-1}")
        )

        total_mask = out_final[cfg.label_col].astype(str).str.startswith("TOTAL")
        other_mask = out_final[cfg.label_col].astype(str).str.startswith("Other products")
        non_total_mask = ~total_mask

        if cfg.revenue_col in df.columns:
            rev_ytd = (
                df.loc[base_t]
                .groupby(cfg.product_col, as_index=False)[cfg.revenue_col]
                .sum()
                .rename(columns={cfg.product_col: cfg.label_col, cfg.revenue_col: "_rev_ytd"})
            )
            out_final = out_final.merge(rev_ytd, on=cfg.label_col, how="left")
            out_final["_rev_ytd"] = to_num(out_final["_rev_ytd"])

            top_rev = float(out_final.loc[non_total_mask & ~other_mask, "_rev_ytd"].sum())
            total_rev = float(rev_ytd["_rev_ytd"].sum())
            out_final.loc[other_mask, "_rev_ytd"] = total_rev - top_rev
            out_final.loc[total_mask, "_rev_ytd"] = total_rev

            if total_rev != 0:
                out_final.loc[non_total_mask, "% of Total YTD"] = out_final.loc[non_total_mask, "_rev_ytd"] / total_rev
                out_final.loc[total_mask, "% of Total YTD"] = 1.0
            else:
                out_final["% of Total YTD"] = 0.0
            out_final = out_final.drop(columns=["_rev_ytd"])
        else:
            total_ytd = float(to_num(out_final.loc[non_total_mask, f"YTD {t}"]).sum())
            if total_ytd != 0:
                out_final.loc[non_total_mask, "% of Total YTD"] = to_num(out_final.loc[non_total_mask, f"YTD {t}"]) / total_ytd
                out_final.loc[total_mask, "% of Total YTD"] = 1.0
            else:
                out_final.loc[non_total_mask, "% of Total YTD"] = 0.0
                out_final.loc[total_mask, "% of Total YTD"] = 0.0

        for m_t, m_t1 in zip(last3m_t, last3m_t1):
            tot_mt = float(df.loc[shop_mask & (df["_month"] == m_t), cfg.actuals_col].sum())
            tot_mt1 = float(df.loc[shop_mask & (df["_month"] == m_t1), cfg.actuals_col].sum())
            out_final.loc[total_mask, m_t] = (tot_mt / tot_mt1 - 1.0) if tot_mt1 != 0 else 0.0

        out_final = out_final.rename(
            columns={
                "YTD % Growth": "YTD % growth",
                "FY % Growth": "FY % growth",
                "Last 3W % Growth": "Last 3w % growth",
            }
        )

        percent_cols = ["% of Total YTD", "YTD % growth", "FY % growth", "Last 3w % growth", *last3m_t]
        for c in percent_cols:
            if c in out_final.columns:
                out_final[c] = normalize_ratio(out_final[c])

        final_cols = [
            cfg.label_col,
            "% of Total YTD",
            f"YTD {t}",
            f"YTD {t-1}",
            "YTD % growth",
            f"FY {t-1}",
            f"FY {t-2}",
            "FY % growth",
            "Last 3w % growth",
            *last3m_t,
        ]
        for c in final_cols:
            if c not in out_final.columns:
                out_final[c] = 0.0

        out_final = out_final[final_cols].fillna(0)

        out_final["_is_total"] = out_final[cfg.label_col].astype(str).str.startswith("TOTAL")
        out_final = (
            out_final.sort_values(["_is_total", f"YTD {t}"], ascending=[True, False])
            .drop(columns=["_is_total"])
            .reset_index(drop=True)
        )

        return t, out_final

    # ----------------------------
    # Product sheet builder (with forecast lookup)
    # ----------------------------
    def build_product_sheet_df_fast(
        cfg: Config,
        shop: str,
        product: str,
        months: List[str],
        cur_month_label: str,
        max_dt_t: pd.Timestamp,
        m_agg: pd.Series,
        d_agg: pd.Series,
        fmt_agg: Optional[pd.Series],
        fc_data: Optional[Dict] = None,
    ) -> pd.DataFrame:
        shop_key = str(shop)
        prod_key = str(product)
        cur_p = pd.Period(cur_month_label, freq="M")

        # Get forecast lookup for this shop/product
        fc = {}
        if fc_data:
            fc = fc_data.get((shop_key, prod_key), {})

        actuals, actuals_ly, yoy, f_yoy, forecast = [], [], [], [], []

        for m in months:
            m_ly = shift_month_label_ly(m)
            mp = pd.Period(m, freq="M")

            if str(m) == str(cur_month_label):
                cutoff_t = pd.to_datetime(max_dt_t)
                cutoff_t1 = cutoff_t - pd.DateOffset(years=1)

                month_start_t = pd.Timestamp(year=cutoff_t.year, month=cutoff_t.month, day=1)
                month_start_t1 = pd.Timestamp(year=cutoff_t1.year, month=cutoff_t1.month, day=1)

                try:
                    s = d_agg.loc[(shop_key, prod_key)]
                    v = float(s.loc[month_start_t:cutoff_t].sum()) if hasattr(s, "loc") else 0.0
                except Exception:
                    v = 0.0

                try:
                    s = d_agg.loc[(shop_key, prod_key)]
                    vly_mtd = float(s.loc[month_start_t1:cutoff_t1].sum()) if hasattr(s, "loc") else 0.0
                except Exception:
                    vly_mtd = 0.0

                vly_full = get_agg_value(m_agg, (shop_key, prod_key, m_ly))

                actuals.append(v)
                actuals_ly.append(vly_full)

                y = 0.0 if (v == 0.0 or vly_mtd == 0.0) else (v / vly_mtd - 1.0)
                yoy.append(y)

                # Current month and future: use combined forecast if available
                if mp >= cur_p:
                    fc_total = sum(val for (f, fr, p), val in fc.items() if p == m)
                    if fc_total > 0:
                        forecast.append(fc_total)
                        f_yoy.append((fc_total / vly_full - 1.0) if vly_full > 0 else 0.0)
                    else:
                        f_yoy.append("na")
                        forecast.append("na")
                else:
                    f_yoy.append(y)
                    forecast.append((1.0 + y) * vly_full)
            else:
                v = get_agg_value(m_agg, (shop_key, prod_key, m))
                vly = get_agg_value(m_agg, (shop_key, prod_key, m_ly))

                actuals.append(v)
                actuals_ly.append(vly)

                y = 0.0 if (v == 0.0 or vly == 0.0) else (v / vly - 1.0)
                yoy.append(y)

                if mp < cur_p:
                    f_yoy.append(y)
                    forecast.append((1.0 + y) * vly)
                else:
                    # Future month: use combined forecast
                    fc_total = sum(val for (f, fr, p), val in fc.items() if p == m)
                    if fc_total > 0:
                        forecast.append(fc_total)
                        f_yoy.append((fc_total / vly - 1.0) if vly > 0 else 0.0)
                    else:
                        f_yoy.append("na")
                        forecast.append("na")

        cols = [product, "", *months]

        base_rows = [
            ["Actuals", "", *actuals],
            ["Actuals LY", "", *actuals_ly],
            ["YoY %", "", *yoy],
            ["", "", *([None] * len(months))],
            ["F-YoY %", "", *f_yoy],
            ["Forecast", "", *forecast],
            ["", "", *([None] * len(months))],
            ["", "", *([None] * len(months))],
        ]
        base_df = pd.DataFrame(base_rows, columns=cols)

        # Available formats section
        if fmt_agg is not None or fc:
            # --- Historical format data ---
            hist_fmt_df = None
            if fmt_agg is not None:
                try:
                    s = fmt_agg.loc[(shop_key, prod_key)]
                except Exception:
                    s = None

                if s is not None and not (isinstance(s, pd.Series) and s.empty):
                    hist_fmt_df = s.rename("v").reset_index()
                    hist_fmt_df = hist_fmt_df[hist_fmt_df["_month"].astype(str).isin(months)].copy()
                    if hist_fmt_df.empty:
                        hist_fmt_df = None

            # --- Build pivot from historical ---
            if hist_fmt_df is not None:
                fmt_pivot = (
                    hist_fmt_df.pivot_table(
                        index=[cfg.format_col, "frame_color"],
                        columns="_month",
                        values="v",
                        aggfunc="sum",
                        fill_value=0.0,
                        dropna=False,
                    )
                    .reset_index()
                )
                for m in months:
                    if m not in fmt_pivot.columns:
                        fmt_pivot[m] = 0.0
                fmt_pivot = fmt_pivot[fmt_pivot[[m for m in months if pd.Period(m, freq="M") < cur_p]].sum(axis=1) > 0] if any(pd.Period(m, freq="M") < cur_p for m in months) else fmt_pivot
                fmt_pivot = fmt_pivot[[cfg.format_col, "frame_color", *months]]
                fmt_pivot = fmt_pivot.rename(columns={cfg.format_col: product, "frame_color": ""})
            else:
                fmt_pivot = pd.DataFrame(columns=[product, "", *months])

            # --- Overlay forecast values for future months ---
            if fc:
                # Collect all (format, frame) combos from forecast data
                fc_combos = {}
                for (f, fr, p), val in fc.items():
                    if f and f not in ("Unknown", ""):
                        fc_combos.setdefault((f, fr), {})[p] = fc_combos.get((f, fr), {}).get(p, 0) + val

                # Get existing combos from historical pivot
                existing_combos = set()
                for _, row in fmt_pivot.iterrows():
                    existing_combos.add((str(row[product]).strip(), str(row[""]).strip()))

                # Update existing rows with forecast values
                for idx in fmt_pivot.index:
                    fmt_name = str(fmt_pivot.loc[idx, product]).strip()
                    frame_name = str(fmt_pivot.loc[idx, ""]).strip()
                    combo_fc = fc_combos.get((fmt_name, frame_name), {})
                    for m in months:
                        mp = pd.Period(m, freq="M")
                        if mp >= cur_p:
                            val = combo_fc.get(m, 0)
                            fmt_pivot.loc[idx, m] = val if val > 0 else "na"

                # Add new format rows that only exist in forecast
                new_rows = []
                for (f, fr), period_vals in fc_combos.items():
                    if (f, fr) not in existing_combos:
                        row_data = {product: f, "": fr}
                        for m in months:
                            mp = pd.Period(m, freq="M")
                            if mp >= cur_p:
                                val = period_vals.get(m, 0)
                                row_data[m] = val if val > 0 else "na"
                            else:
                                row_data[m] = 0
                        new_rows.append(row_data)

                if new_rows:
                    fmt_pivot = pd.concat(
                        [fmt_pivot, pd.DataFrame(new_rows, columns=[product, "", *months])],
                        ignore_index=True,
                    )
            else:
                # No forecast data: set future months to "na" (original behavior)
                for m in months:
                    if pd.Period(m, freq="M") >= cur_p:
                        fmt_pivot[m] = "na"

            # Remove rows that are all zero/na across all months
            def row_has_data(row):
                for m in months:
                    v = row[m]
                    if v != "na" and v != 0 and v is not None:
                        try:
                            if float(v) != 0:
                                return True
                        except (ValueError, TypeError):
                            pass
                return False

            if not fmt_pivot.empty:
                keep_mask = fmt_pivot.apply(row_has_data, axis=1)
                fmt_pivot = fmt_pivot[keep_mask].reset_index(drop=True)

            if not fmt_pivot.empty:
                section_header = pd.DataFrame(
                    [["Available formats:", "frames:", *([None] * len(months))]],
                    columns=cols,
                )
                return pd.concat([base_df, section_header, fmt_pivot], ignore_index=True)

        return base_df

    # ----------------------------
    # Excel formatting
    # ----------------------------
    WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
    FORECAST_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")

    LEFT_ALIGN = Alignment(horizontal="left")
    RIGHT_ALIGN = Alignment(horizontal="right")
    INDENT_ALIGN = Alignment(horizontal="left", indent=1)
    PERCENT_FMT = "0.0%"
    INTEGER_FMT = "#,##0"

    THIN = Side(style="thin", color="000000")
    HEADER_BOTTOM_BORDER = Border(bottom=THIN)
    ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BOLD_FONT = Font(bold=True)

    def is_total_value(v) -> bool:
        return isinstance(v, str) and v.startswith("TOTAL")

    def _set_right_border(cell, side: Side):
        b = cell.border or Border()
        cell.border = Border(left=b.left, right=side, top=b.top, bottom=b.bottom)

    def apply_formatting_summary(ws, label_header: str):
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 1 or max_col < 1:
            return

        total_row_idx = None
        for r in range(2, max_row + 1):
            if is_total_value(ws.cell(row=r, column=1).value):
                total_row_idx = r
                break

        end_row = total_row_idx if total_row_idx is not None else max_row

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.fill = WHITE_FILL
                cell.border = Border()

        headers = {}
        for c in range(1, max_col + 1):
            hcell = ws.cell(row=1, column=c)
            headers[hcell.value] = c
            hcell.font = BOLD_FONT
            hcell.alignment = LEFT_ALIGN
            hcell.border = HEADER_BOTTOM_BORDER

        rb_names = {"YTD % growth", "FY % growth", "Last 3w % growth"}
        rb_cols = [headers.get(n) for n in rb_names if headers.get(n)]
        if rb_cols:
            for r in range(1, end_row + 1):
                for c in rb_cols:
                    cell = ws.cell(row=r, column=c)
                    b = cell.border or Border()
                    cell.border = Border(left=b.left, right=THIN, top=b.top, bottom=b.bottom)

        label_col_idx = headers.get(label_header)
        if label_col_idx:
            ws.cell(row=1, column=label_col_idx).alignment = LEFT_ALIGN
            for r in range(2, end_row + 1):
                v = ws.cell(row=r, column=label_col_idx).value
                ws.cell(row=r, column=label_col_idx).alignment = (
                    INDENT_ALIGN if not is_total_value(v) else LEFT_ALIGN
                )

        percent_headers = {"% of Total YTD", "YTD % growth", "FY % growth", "Last 3w % growth"}
        percent_col_idxs = []
        for h, c in headers.items():
            if h is None:
                continue
            hs = str(h).strip()
            if hs in percent_headers or hs.count("-") == 1:
                percent_col_idxs.append(c)
                for r in range(2, end_row + 1):
                    ws.cell(row=r, column=c).number_format = PERCENT_FMT

        if total_row_idx is not None:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=total_row_idx, column=c)
                cell.font = BOLD_FONT
                b = cell.border or Border()
                cell.border = Border(left=b.left, right=b.right, top=THIN, bottom=b.bottom)

        for c in range(1, max_col + 1):
            is_percent = c in percent_col_idxs
            max_len = 0
            for r in range(1, max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    s = ""
                elif r == 1:
                    s = str(v)
                elif is_percent and r <= end_row:
                    try:
                        s = f"{float(v) * 100:.1f}%"
                    except Exception:
                        s = "0.0%"
                else:
                    s = str(v)
                max_len = max(max_len, len(s))

            ws.column_dimensions[
                ws.cell(row=1, column=c).column_letter
            ].width = max(8, min(60, max_len + 2))

    def add_destination_row_after_total_from_agg(ws, cfg: Config, shop: str, dest_agg: Optional[pd.Series]):
        if dest_agg is None:
            return

        total_row_idx = None
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if is_total_value(v):
                total_row_idx = r
                break
        if total_row_idx is None:
            return

        ws.insert_rows(total_row_idx + 1, amount=2)

        shop_key = str(shop)
        top_dest = "Unknown"
        try:
            s = dest_agg.loc[shop_key]
            if hasattr(s, "sort_values"):
                s2 = s.sort_values(ascending=False)
                if len(s2) > 0:
                    top_dest = s2.index[0]
                    if top_dest is None or str(top_dest).strip() == "" or str(top_dest).lower() == "nan":
                        top_dest = "Unknown"
        except Exception:
            top_dest = "Unknown"

        dest_row = total_row_idx + 3
        ws.cell(row=dest_row, column=1, value="Destination:").font = BOLD_FONT
        ws.cell(row=dest_row, column=1).alignment = LEFT_ALIGN
        ws.cell(row=dest_row, column=2, value=str(top_dest)).font = BOLD_FONT
        ws.cell(row=dest_row, column=2).alignment = LEFT_ALIGN

    def write_product_last6w_block_from_agg(
        ws, cfg: Config, shop: str, product: str, start_row: int, start_col: int, w_agg: pd.Series,
    ):
        shop_key = str(shop)
        prod_key = str(product)

        try:
            s = w_agg.loc[(shop_key, prod_key)]
            keys = sorted(s.index.astype(str).unique().tolist(), key=parse_week_key)
        except Exception:
            return

        if not keys:
            return

        keys = keys[:-1]
        if not keys:
            return

        last6 = keys[-6:] if len(keys) >= 6 else keys

        actuals, actuals_ly, yoy = [], [], []
        for k in last6:
            v = get_agg_value(w_agg, (shop_key, prod_key, str(k)))

            y, w = parse_week_key(str(k))
            k_ly = f"{y-1}-{str(w).zfill(2)}" if y != -1 else "0000-00"
            vly = get_agg_value(w_agg, (shop_key, prod_key, k_ly))

            actuals.append(v)
            actuals_ly.append(vly)
            yv = 0.0 if (v == 0.0 or vly == 0.0) else (v / vly - 1.0)
            yoy.append(yv)

        r = start_row
        ws.cell(row=r, column=start_col, value="Last 6w % growth:").font = BOLD_FONT
        r += 1

        ws.cell(row=r, column=start_col, value="").border = HEADER_BOTTOM_BORDER
        for i, k in enumerate(last6, start=1):
            c = start_col + i
            cell = ws.cell(row=r, column=c, value=str(k))
            cell.font = BOLD_FONT
            cell.border = HEADER_BOTTOM_BORDER
            cell.alignment = LEFT_ALIGN
        r += 1

        labels = ["Actuals", "Actuals LY", "YoY %"]
        rows = [actuals, actuals_ly, yoy]
        for lbl, vals in zip(labels, rows):
            ws.cell(row=r, column=start_col, value=lbl).font = BOLD_FONT
            ws.cell(row=r, column=start_col).alignment = LEFT_ALIGN
            for i, v in enumerate(vals, start=1):
                c = start_col + i
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = RIGHT_ALIGN
                if lbl == "YoY %":
                    cell.number_format = PERCENT_FMT
            r += 1

    def write_product_calcs_transposed(ws, summary_ws, cfg: Config, product: str, start_row: int = 1, start_col: int = 1):
        prod_row = None
        for r in range(2, summary_ws.max_row + 1):
            if str(summary_ws.cell(row=r, column=1).value) == str(product):
                prod_row = r
                break
        if prod_row is None:
            return

        header_map = {}
        for c in range(1, summary_ws.max_column + 1):
            h = summary_ws.cell(row=1, column=c).value
            if h is not None:
                header_map[str(h).strip()] = c

        metrics = [h for h in header_map.keys() if h != cfg.label_col]
        underline_metrics = {"YTD % growth", "FY % growth", "Last 3w % growth"}

        r = start_row
        ws.cell(row=r, column=start_col).value = "Calculations"
        ws.cell(row=r, column=start_col).font = BOLD_FONT
        r += 1

        ws.cell(row=r, column=start_col).value = "Metric"
        ws.cell(row=r, column=start_col + 1).value = "Value"
        ws.cell(row=r, column=start_col).font = BOLD_FONT
        ws.cell(row=r, column=start_col + 1).font = BOLD_FONT
        r += 1

        for m in metrics:
            src_c = header_map[m]
            src = summary_ws.cell(row=prod_row, column=src_c)

            mcell = ws.cell(row=r, column=start_col)
            mcell.value = m

            tgt = ws.cell(row=r, column=start_col + 1)
            tgt.value = src.value
            tgt.number_format = src.number_format

            if str(m).strip() in underline_metrics:
                for cell in (mcell, tgt):
                    b = cell.border or Border()
                    cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=THIN)

            r += 1

    def apply_formatting_product(ws, cur_month_label: str, row_offset: int = 0):
        max_row = ws.max_row
        max_col = ws.max_column

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.fill = WHITE_FILL

        header_r = 1 + row_offset

        for c in range(1, max_col + 1):
            cell = ws.cell(row=header_r, column=c)
            cell.border = HEADER_BOTTOM_BORDER
            cell.font = BOLD_FONT
            cell.alignment = LEFT_ALIGN

        for r in range(header_r + 1, max_row + 1):
            ws.cell(row=r, column=1).alignment = LEFT_ALIGN
            ws.cell(row=r, column=2).alignment = LEFT_ALIGN

        yoy_row = 4 + row_offset
        f_yoy_row = 6 + row_offset
        for c in range(3, max_col + 1):
            ws.cell(row=yoy_row, column=c).number_format = PERCENT_FMT
            ws.cell(row=f_yoy_row, column=c).number_format = PERCENT_FMT

        ws.cell(row=yoy_row, column=1).font = BOLD_FONT
        ws.cell(row=f_yoy_row, column=1).font = BOLD_FONT

        forecast_row = None
        avail_row = None
        for r in range(header_r, max_row + 1):
            v = str(ws.cell(row=r, column=1).value or "").strip()
            if v == "Forecast":
                forecast_row = r
            elif v == "Available formats:":
                avail_row = r

        if forecast_row is not None:
            ws.cell(row=forecast_row, column=1).font = BOLD_FONT
            # Integer format for numeric forecast cells
            for c in range(3, max_col + 1):
                cell = ws.cell(row=forecast_row, column=c)
                if isinstance(cell.value, (int, float, np.integer, np.floating)):
                    cell.number_format = INTEGER_FMT

        month_to_col = {}
        for c in range(3, max_col + 1):
            h = ws.cell(row=header_r, column=c).value
            if h is not None:
                month_to_col[str(h).strip()] = c

        start_c = month_to_col.get(str(cur_month_label).strip())

        if forecast_row is not None and start_c is not None:
            end_c = min(max_col, start_c + 5)
            for c in range(start_c, end_c + 1):
                cell = ws.cell(row=forecast_row, column=c)
                cell.border = ALL_THIN
                cell.fill = FORECAST_FILL

        if avail_row is not None:
            ws.cell(row=avail_row, column=1).font = BOLD_FONT
            ws.cell(row=avail_row, column=1).alignment = LEFT_ALIGN
            ws.cell(row=avail_row, column=2).font = BOLD_FONT
            ws.cell(row=avail_row, column=2).alignment = LEFT_ALIGN

            for c in range(1, max_col + 1):
                cell = ws.cell(row=avail_row, column=c)
                b = cell.border or Border()
                cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=THIN)

            if start_c is not None:
                end_c = min(max_col, start_c + 5)
                r = avail_row + 1
                while r <= max_row:
                    v1 = ws.cell(row=r, column=1).value
                    v2 = ws.cell(row=r, column=2).value
                    if (v1 is None or str(v1).strip() == "") and (v2 is None or str(v2).strip() == ""):
                        break
                    ws.cell(row=r, column=1).alignment = INDENT_ALIGN
                    ws.cell(row=r, column=2).alignment = INDENT_ALIGN
                    for c in range(start_c, end_c + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = ALL_THIN
                        cell.fill = FORECAST_FILL
                        # Integer format for numeric format-level forecast cells
                    for c in range(3, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        if isinstance(cell.value, (int, float, np.integer, np.floating)):
                            cell.number_format = INTEGER_FMT
                    r += 1

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str) and cell.value.strip().lower() == "na":
                    cell.alignment = RIGHT_ALIGN

        def display_len(cell) -> int:
            v = cell.value
            if v is None:
                return 0
            if isinstance(v, str):
                return len(v)
            if cell.number_format == PERCENT_FMT and isinstance(v, (int, float, np.integer, np.floating)):
                try:
                    return len(f"{float(v) * 100:.1f}%")
                except Exception:
                    return len(str(v))
            if isinstance(v, (int, float, np.integer, np.floating)):
                try:
                    return len(f"{float(v):,.0f}")
                except Exception:
                    return len(str(v))
            return len(str(v))

        max_len = 0
        for r in range(1, ws.max_row + 1):
            max_len = max(max_len, display_len(ws.cell(row=r, column=1)))
        ws.column_dimensions[get_column_letter(1)].width = max(14, min(60, max_len + 2))

        max_len = 0
        for r in range(1, ws.max_row + 1):
            max_len = max(max_len, display_len(ws.cell(row=r, column=2)))
        ws.column_dimensions[get_column_letter(2)].width = max(10, min(35, max_len + 2))

        for c in range(3, ws.max_column + 1):
            header = ws.cell(row=header_r, column=c).value
            header_len = len(str(header)) if header is not None else 0

            scan_rows = set([header_r, header_r + 1, header_r + 2, yoy_row, f_yoy_row])
            if forecast_row is not None:
                scan_rows.add(forecast_row)
            if avail_row is not None:
                scan_rows.add(avail_row)
                scan_rows.add(avail_row + 1)

            max_len2 = header_len
            for rr in scan_rows:
                if 1 <= rr <= ws.max_row:
                    max_len2 = max(max_len2, display_len(ws.cell(row=rr, column=c)))

            ws.column_dimensions[get_column_letter(c)].width = max(10, min(18, max_len2 + 2))

    # ----------------------------
    # Excel build (uses aggregates + forecast lookup)
    # ----------------------------
    def build_excel_bytes(
        sheet_name: str, out: pd.DataFrame, df_raw: pd.DataFrame,
        shop: str, cfg: Config, fc_data: Optional[Dict] = None,
    ) -> bytes:
        df_clean = clean_input(df_raw, cfg)
        agg = build_aggregates(df_clean, cfg)

        shop = (shop or "").strip()
        shop_key = str(shop)

        df_scope = df_clean[df_clean[cfg.shop_col] == shop_key] if shop_key else df_clean
        t = int(df_scope[cfg.date_col].dt.year.max())
        max_dt_t = df_scope.loc[df_scope[cfg.date_col].dt.year == t, cfg.date_col].max()

        months = month_labels_m2_to_p11(max_dt_t)
        cur_month_label = str(max_dt_t.to_period("M"))

        top_products_series = (
            out.loc[~out[cfg.label_col].astype(str).str.startswith(("Other products", "TOTAL")), cfg.label_col]
            .astype(str)
        )
        has_other = out[cfg.label_col].astype(str).str.startswith("Other products").any()
        top_products = (top_products_series.head(cfg.top_n) if has_other else top_products_series).tolist()

        buf = BytesIO()
        used = set()

        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            base = safe_sheet_base(sheet_name)
            summary_sheet = dedupe_sheet_name(base, used)

            out.to_excel(writer, sheet_name=summary_sheet, index=False)
            summary_ws = writer.book[summary_sheet]

            apply_formatting_summary(summary_ws, cfg.label_col)
            add_destination_row_after_total_from_agg(summary_ws, cfg, shop_key, agg.get("dest_agg"))
            apply_formatting_summary(summary_ws, cfg.label_col)

            for prod in top_products:
                prod_df = build_product_sheet_df_fast(
                    cfg=cfg,
                    shop=shop_key,
                    product=str(prod),
                    months=months,
                    cur_month_label=cur_month_label,
                    max_dt_t=max_dt_t,
                    m_agg=agg["m_agg"],
                    d_agg=agg["d_agg"],
                    fmt_agg=agg.get("fmt_agg"),
                    fc_data=fc_data,
                )

                sname = dedupe_sheet_name(safe_sheet_base(prod), used)
                prod_df.to_excel(writer, sheet_name=sname, index=False)

                ws = writer.book[sname]
                ws.insert_rows(1, 15)
                write_product_calcs_transposed(ws, summary_ws, cfg, prod, start_row=1, start_col=1)
                write_product_last6w_block_from_agg(ws, cfg, shop_key, prod, start_row=1, start_col=4, w_agg=agg["w_agg"])
                apply_formatting_product(ws, cur_month_label, row_offset=15)

        return buf.getvalue()

    # ----------------------------
    # IO / Slack
    # ----------------------------
    def read_csv_from_s3(cfg: Config) -> pd.DataFrame:
        obj = S3.get_object(Bucket=cfg.bucket, Key=cfg.in_key)
        return pd.read_csv(obj["Body"])



    def slack_api_post_form(method: str, token: str, fields: dict, timeout: int = 20) -> dict:
        body = urllib.parse.urlencode(fields).encode("utf-8")
        req = urllib.request.Request(
            f"https://slack.com/api/{method}",
            data=body,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/x-www-form-urlencoded; charset=utf-8",
            },
            method="POST",
        )
        try:
            resp = urllib.request.urlopen(req, timeout=timeout).read()
        except urllib.error.HTTPError as e:
            err_body = e.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Slack API HTTPError {e.code} for {method}: {err_body}") from e

        data = json.loads(resp.decode("utf-8"))
        if not data.get("ok"):
            raise RuntimeError(f"Slack API {method} failed: {data}")
        return data

    def slack_upload_xlsx(file_bytes: bytes, filename: str, message: str):
        token = os.environ["SLACK_BOT_TOKEN"]
        channels_env = os.environ["SLACK_CHANNEL_ID"]
        channel_ids = [c.strip() for c in channels_env.split(",") if c.strip()]
        if not channel_ids:
            raise ValueError("SLACK_CHANNEL_ID is empty")

        init = slack_api_post_form(
            "files.getUploadURLExternal",
            token,
            {"filename": filename, "length": str(len(file_bytes))},
        )
        upload_url = init["upload_url"]
        file_id = init["file_id"]

        upload_req = urllib.request.Request(
            upload_url,
            data=file_bytes,
            headers={"Content-Type": "application/octet-stream"},
            method="POST",
        )
        try:
            upload_resp = urllib.request.urlopen(upload_req, timeout=30)
            status = getattr(upload_resp, "status", None) or upload_resp.getcode()
            if int(status) != 200:
                raise RuntimeError(f"Upload to Slack upload_url failed: HTTP {status}")
        except urllib.error.HTTPError as e:
            err_body = e.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Upload to Slack upload_url HTTPError {e.code}: {err_body}") from e

        files_payload = json.dumps([{"id": file_id, "title": filename}])

        complete_fields = {"files": files_payload, "initial_comment": message}
        if len(channel_ids) == 1:
            complete_fields["channel_id"] = channel_ids[0]
        else:
            complete_fields["channels"] = ",".join(channel_ids)

        slack_api_post_form("files.completeUploadExternal", token, complete_fields)

# ----------------------------
    # MAIN (SHOP ONLY)
    # ----------------------------
    shop = parse_shop_request(event).strip()
    if not shop:
        raise ValueError('Invalid request. Use Slack: "/review shop <shopname>"')

    df = read_csv_from_s3(CFG)

    fc_data = read_combined_forecasts(CFG.bucket, COMBINED_KEY)

    t, out = compute_review_table(df, shop, CFG)

    sheet_name = shop or "review"
    excel_bytes = build_excel_bytes(sheet_name, out, df, shop, CFG, fc_data=fc_data)

    slack_upload_xlsx(
        excel_bytes,
        filename=f"review_shop_{shop}.xlsx",
        message=f"Finished reviewing shop {sheet_name}",
    )

    return {"ok": True, "mode": "shop", "shop": shop, "year_t": t}