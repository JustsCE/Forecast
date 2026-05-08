import os
import json
import urllib3

http = urllib3.PoolManager()

CLIENT_ID = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]
TENANT_ID = os.environ["TENANT_ID"]
SITE_HOSTNAME = os.environ["SITE_HOSTNAME"]
SITE_PATH = os.environ["SITE_PATH"]
FOLDER_PATH = os.environ["FOLDER_PATH"]

import boto3
import numpy as np
import pandas as pd
from io import BytesIO
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

def slack_reply(response_url, text):
    if not response_url:
        return
    http.request("POST", response_url,
        body=json.dumps({"response_type": "in_channel", "text": text}).encode("utf-8"),
        headers={"Content-Type": "application/json"})


def get_access_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    resp = http.request(
        "POST", url,
        body=(f"client_id={CLIENT_ID}&client_secret={CLIENT_SECRET}"
              f"&scope=https://graph.microsoft.com/.default"
              f"&grant_type=client_credentials"),
        headers={"Content-Type": "application/x-www-form-urlencoded"})
    return json.loads(resp.data.decode("utf-8"))["access_token"]


def upload_to_sharepoint(token, file_bytes, file_name, content_type="text/csv"):
    h = {"Authorization": f"Bearer {token}"}
    site_url = f"https://graph.microsoft.com/v1.0/sites/{SITE_HOSTNAME}:{SITE_PATH}"
    site_id = json.loads(
        http.request("GET", site_url, headers=h).data.decode("utf-8"))["id"]
    drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    drives = json.loads(
        http.request("GET", drives_url, headers=h).data.decode("utf-8"))
    drive_id = next(d["id"] for d in drives["value"] if d["name"] == "Documents")
    path = f"{FOLDER_PATH}/{file_name}".replace("//", "/")
    url = (f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
           f"/root:/{path}:/content")
    resp = http.request("PUT", url, body=file_bytes,
                        headers={**h, "Content-Type": content_type})
    return json.loads(resp.data.decode("utf-8"))


def lambda_handler(event, context):
    BUCKET = "bi-automations"
    ACTUALS_KEY = "Forecast/actuals.csv"
    COMBINED_KEY = "Forecast/Seperate Forecasts/seperate_forecasts_combined.csv"
    INPUT_KEY = "Forecast/Forecast_Input.xlsx"
    OUTPUT_KEY = "Forecast/Seperate Forecasts/remaining_forecasts.xlsx"
    FORECAST_KEY = "Forecast/forecast.csv"
    EXCLUDED_PRODUCTS = {"EXCLUDED PRODUCT", "NEW PRODUCT"}
    N_MONTHS = 3

    S3 = boto3.client("s3")

    # ── Helpers ───────────────────────────────────────────────────────
    def to_num(s):
        return pd.to_numeric(s, errors="coerce").fillna(0).astype(float)

    def safe_growth(curr, prev):
        curr, prev = to_num(curr), to_num(prev)
        out = np.where(prev != 0, (curr / prev) - 1.0, 0.0)
        return pd.Series(np.where(np.isfinite(out), out, 0.0), index=curr.index)

    def normalize_ratio(s):
        s = to_num(s)
        mask = s.abs() > 5
        s.loc[mask] = s.loc[mask] / 100.0
        return s

    def last_3_full_months(max_date):
        month_start = pd.Timestamp(year=max_date.year, month=max_date.month, day=1)
        last_full = (month_start - pd.Timedelta(days=1)).to_period("M")
        return [str(p) for p in pd.period_range(end=last_full, periods=3, freq="M")]

    def shift_month_ly(m):
        return f"{int(m[:4]) - 1}{m[4:]}"

    def compute_shares(df, *, groupby_dims, total_by_dims,
                    months=None, month_col=None,
                    excluded=None, exclude_col=None):
        required_cols = set(groupby_dims) | set(total_by_dims) | {"actuals"}
        missing = required_cols - set(df.columns)
        if missing:
            raise KeyError(f"df is missing required columns: {sorted(missing)}")
        data = df.copy()
        if months is not None:
            if not month_col:
                raise ValueError("month_col required when months is specified.")
            data = data[data[month_col].isin(months)].copy()
        if excluded is not None:
            if not exclude_col:
                raise ValueError("exclude_col required when excluded is specified.")
            data = data[~data[exclude_col].isin(excluded)].copy()
        g = data.groupby(groupby_dims, as_index=False, dropna=False).agg({"actuals": "sum"})
        totals = g.groupby(total_by_dims, dropna=False)["actuals"].transform("sum")
        g["denominator_actuals"] = totals
        g["share"] = np.divide(g["actuals"], totals,
                            out=np.zeros_like(g["actuals"], dtype=float),
                            where=totals.to_numpy() > 0)
        return g

    def make_combo_key(df, cols, new_col="combo_key"):
        df[new_col] = df[cols[0]].astype(str).str.cat(
            [df[c].astype(str) for c in cols[1:]], sep="_")
        return df

    def share_check(df, key_cols, avg_col="avg_type",
                    last_n_label="last N months", q4_label="last Q4"):
        last_n_df = df.loc[df[avg_col] == last_n_label].copy()
        q4_df = df.loc[df[avg_col] == q4_label].copy()
        presence_df = (
            last_n_df[key_cols].assign(in_last_n=True)
            .merge(q4_df[key_cols].assign(in_q4=True), on=key_cols, how="outer")
            .fillna({"in_last_n": False, "in_q4": False}))
        presence_df["status"] = np.select([
            presence_df["in_last_n"] & presence_df["in_q4"],
            presence_df["in_last_n"] & ~presence_df["in_q4"],
            ~presence_df["in_last_n"] & presence_df["in_q4"]],
            ["both", "only_last_n", "only_q4"], default="none")
        last_n_shares = last_n_df[key_cols + ["share"]].rename(columns={"share": "share_last_n"})
        q4_shares = q4_df[key_cols + ["share"]].rename(columns={"share": "share_q4"})
        presence_df = (presence_df
                    .merge(last_n_shares, on=key_cols, how="left")
                    .merge(q4_shares, on=key_cols, how="left"))
        presence_df = presence_df.drop(columns=["in_last_n", "in_q4"])
        return presence_df.loc[presence_df["status"] != "both"].reset_index(drop=True)

    def get_week_num(w):
        return int(w[-2:]) if w[-2:].isdigit() else int(w.split("-")[-1])

    def iso_weeks_clamped(dates, year):
        """Build '{year}-WW' strings, clamping boundary dates to year's range."""
        iso = dates.dt.isocalendar()
        iy = iso.year.astype(int).values
        iw = iso.week.astype(int).values
        mask_t = iy == year
        max_wk = int(iw[mask_t].max()) if mask_t.any() else 52
        return [
            f"{year}-{max_wk:02d}" if y > year else
            f"{year}-01" if y < year else
            f"{year}-{w:02d}"
            for y, w in zip(iy, iw)
        ], max_wk

    # ── Load actuals ──────────────────────────────────────────────────
    act_full_raw = S3.get_object(Bucket=BUCKET, Key=ACTUALS_KEY)["Body"].read()
    act_full = pd.read_csv(BytesIO(act_full_raw))
    del act_full_raw  # raw bytes no longer needed

    act_full["fulldate"] = pd.to_datetime(act_full["fulldate"], errors="coerce")
    act_full = act_full[act_full["fulldate"].notna()]
    act_full = act_full[~act_full["forecast_product"].astype(str).isin(EXCLUDED_PRODUCTS)]
    act_full["actuals"] = to_num(act_full["actuals"])
    act_full["_month"] = act_full["fulldate"].dt.to_period("M").astype(str)
    act_full["product"] = act_full["forecast_product"].astype(str).str.strip()
    act_full["week"] = act_full["week"].astype(str)
    act_full["region"] = act_full["destination_region"].astype(str).str.strip()

    # ── Load Forecast_Input.xlsx ──────────────────────────────────────
    input_bytes = BytesIO(S3.get_object(Bucket=BUCKET, Key=INPUT_KEY)["Body"].read())
    new_formats_eu = pd.read_excel(input_bytes, sheet_name="EU formats"); input_bytes.seek(0)
    new_formats_us = pd.read_excel(input_bytes, sheet_name="US formats"); input_bytes.seek(0)
    new_formats = pd.concat([new_formats_eu, new_formats_us], ignore_index=True)
    del new_formats_eu, new_formats_us  # merged into new_formats

    ignore_formats = pd.read_excel(input_bytes, sheet_name="adhoc", header=9, usecols="E").dropna()
    ignore_formats.columns = ["ignore_format"]

    input_bytes.seek(0)
    pcs_eu = pd.read_excel(input_bytes, sheet_name="EU pcs"); input_bytes.seek(0)
    pcs_us = pd.read_excel(input_bytes, sheet_name="US pcs")

    def load_pcs(df, destination_region):
        df = df.copy()
        df.columns = [str(c).strip() for c in df.columns]
        long = df.melt(id_vars=["PRODUCT"], var_name="col", value_name="pcs_share")
        long[["year_month", "pcs"]] = long["col"].str.split("_", n=1, expand=True)
        long["pcs"] = long["pcs"].str.strip()
        long = long[long["pcs"].notna() & (long["pcs"] != "")].copy()
        long["destination_region"] = destination_region
        return long[["PRODUCT", "year_month", "pcs", "destination_region", "pcs_share"]]

    pcs_shares = pd.concat([load_pcs(pcs_eu, "EU+RoW"), load_pcs(pcs_us, "US+CA")], ignore_index=True)
    del pcs_eu, pcs_us  # merged into pcs_shares

    input_bytes.seek(0)
    anchor_weeks = pd.read_excel(input_bytes, sheet_name="Anchor Weeks")
    anchor_weeks.columns = [str(c).strip() for c in anchor_weeks.columns]

    input_bytes.seek(0)
    update_info = pd.read_excel(input_bytes, sheet_name="info", header=None)
    input_bytes.seek(0)
    substitute_distribution = pd.read_excel(input_bytes, sheet_name="adhoc", header=9, usecols="A:C").dropna()
    del input_bytes  # all sheets loaded

    current_month = str(update_info.iat[3, 2])   # cell C4
    cutoff_week = str(update_info.iat[1, 2])      # cell C2
    del update_info  # only needed for those two cells
    actuals_end = int(cutoff_week[-2:]) if cutoff_week[-2:].isdigit() else int(cutoff_week.split("-")[-1])
    t = int(current_month[:4])

    # ── Build calc sheet for a region ─────────────────────────────────
    def build_calc_sheet(df):
        date_col, actuals_col, product_col, week_col = "fulldate", "actuals", "product", "week"
        yr = int(df[date_col].dt.year.max())
        max_dt = df.loc[df[date_col].dt.year == yr, date_col].max()
        cutoff_ly = max_dt - pd.DateOffset(years=1)

        def agg(m):
            return df.loc[m].groupby(product_col, as_index=False)[actuals_col].sum()

        ytd_t = agg(df[date_col].dt.year.eq(yr) & df[date_col].le(max_dt)).rename(columns={actuals_col: f"YTD {yr}"})
        ytd_ly = agg(df[date_col].dt.year.eq(yr - 1) & df[date_col].le(cutoff_ly)).rename(columns={actuals_col: f"YTD {yr-1}"})
        out = ytd_t.merge(ytd_ly, on=product_col, how="outer").fillna(0)

        def week_keys_for_year(y):
            keys = sorted(df.loc[df[week_col].str.startswith(f"{y}-"), week_col].dropna().unique())
            trimmed = keys[:-1]
            return trimmed[-3:] if len(trimmed) >= 3 else trimmed

        w_t = week_keys_for_year(yr)
        w_ly = [f"{yr-1}-{k.split('-', 1)[1]}" for k in w_t]
        l3w_t = agg(df[week_col].isin(w_t)).rename(columns={actuals_col: f"Last 3W {yr}"})
        l3w_ly = agg(df[week_col].isin(w_ly)).rename(columns={actuals_col: f"Last 3W {yr-1}"})
        out = out.merge(l3w_t, on=product_col, how="outer").merge(l3w_ly, on=product_col, how="outer").fillna(0)

        last3m = last_3_full_months(max_dt)
        for m in last3m:
            m_ly = shift_month_ly(m)
            mt = agg(df["_month"].eq(m)).rename(columns={actuals_col: "_mt"})
            mly = agg(df["_month"].eq(m_ly)).rename(columns={actuals_col: "_mly"})
            tmp = mt.merge(mly, on=product_col, how="outer").fillna(0)
            tmp[m] = safe_growth(tmp["_mt"], tmp["_mly"])
            out = out.merge(tmp[[product_col, m]], on=product_col, how="outer").fillna(0)

        out["YTD % growth"] = safe_growth(out[f"YTD {yr}"], out[f"YTD {yr-1}"])
        out["Last 3w % growth"] = safe_growth(out[f"Last 3W {yr}"], out[f"Last 3W {yr-1}"])
        out = out.sort_values(f"YTD {yr}", ascending=False).reset_index(drop=True)

        num_cols = [c for c in out.columns if c != product_col and pd.api.types.is_numeric_dtype(out[c])]
        total_vals = out[num_cols].sum()
        total_row = pd.DataFrame([{product_col: "TOTAL", **total_vals.to_dict()}])
        final = pd.concat([out, total_row], ignore_index=True)

        final["YTD % growth"] = safe_growth(final[f"YTD {yr}"], final[f"YTD {yr-1}"])
        final["Last 3w % growth"] = safe_growth(final[f"Last 3W {yr}"], final[f"Last 3W {yr-1}"])

        non_total = ~final[product_col].str.startswith("TOTAL")
        total_ytd = final.loc[non_total, f"YTD {yr}"].sum()
        final["% of Total YTD"] = np.where(non_total, final[f"YTD {yr}"] / total_ytd if total_ytd else 0, 1.0)

        total_mask = final[product_col].str.startswith("TOTAL")
        for m in last3m:
            m_ly = shift_month_ly(m)
            tot_mt = float(df.loc[df["_month"] == m, actuals_col].sum())
            tot_mly = float(df.loc[df["_month"] == m_ly, actuals_col].sum())
            final.loc[total_mask, m] = (tot_mt / tot_mly - 1.0) if tot_mly != 0 else 0.0

        pct_cols = ["% of Total YTD", "YTD % growth", "Last 3w % growth", *last3m]
        for c in pct_cols:
            if c in final.columns:
                final[c] = normalize_ratio(final[c])

        final_cols = [product_col, "% of Total YTD", f"YTD {yr}", f"YTD {yr-1}", "YTD % growth",
                    "Last 3w % growth", *last3m]
        for c in final_cols:
            if c not in final.columns:
                final[c] = 0.0
        final = final[final_cols].fillna(0)
        final["_is_total"] = final[product_col].str.startswith("TOTAL")
        final = (final.sort_values(["_is_total", f"YTD {yr}"], ascending=[True, False])
                .drop(columns=["_is_total"]).reset_index(drop=True))
        return final, last3m

    # ── Coverage analysis ─────────────────────────────────────────────
    act = act_full[act_full["fulldate"] >= "2025-01-01"].copy()
    act["shoptype"] = act["shoptype"].astype(str).str.strip()
    act["shop"] = act["forecasted_shop"].astype(str).str.strip()

    actuals_combos = (
        act.groupby(["shoptype", "shop", "product", "region"], dropna=False)
        .size().reset_index(name="actuals_rows"))
    del act  # only needed for actuals_combos

    fc = pd.read_csv(S3.get_object(Bucket=BUCKET, Key=COMBINED_KEY)["Body"])
    fc["region"] = fc["destination"].astype(str).str.strip()
    fc["shoptype"] = fc["shoptype"].astype(str).str.strip()
    fc["shop"] = fc["forecasted_shop"].astype(str).str.strip()
    fc["product"] = fc["product"].astype(str).str.strip()

    fully_covered, shop_covered = set(), set()
    for (st, prod, reg), grp in fc.groupby(["shoptype", "product", "region"]):
        shops_in_fc = set(grp["shop"].unique())
        if st in shops_in_fc:
            fully_covered.add((st, prod, reg)); continue
        other_name = f"Other {st}"
        if other_name in shops_in_fc:
            fully_covered.add((st, prod, reg)); continue
        for shop in shops_in_fc:
            shop_covered.add((st, shop, prod, reg))

    def check_forecasted(row):
        if (row["shoptype"], row["product"], row["region"]) in fully_covered:
            return "Yes"
        if (row["shoptype"], row["shop"], row["product"], row["region"]) in shop_covered:
            return "Yes"
        return "No"

    actuals_combos["forecasted"] = actuals_combos.apply(check_forecasted, axis=1)
    del fully_covered, shop_covered  # sets no longer needed after apply
    actuals_combos = actuals_combos.sort_values(
        ["forecasted", "shoptype", "product", "region", "shop"]).reset_index(drop=True)

    not_fc = actuals_combos[actuals_combos["forecasted"] == "No"]

    unfc_combos = not_fc[["region", "product", "shop"]].drop_duplicates().reset_index(drop=True)
    unfc_combos["_key"] = unfc_combos["region"] + "||" + unfc_combos["product"] + "||" + unfc_combos["shop"]

    _act_keyed = act_full.copy()
    _act_keyed["_key"] = (
        _act_keyed["region"] + "||" +
        _act_keyed["product"].astype(str).str.strip() + "||" +
        _act_keyed["forecasted_shop"].astype(str).str.strip()
    )

    df_eu = _act_keyed[
        (_act_keyed["region"] == "EU+RoW") &
        (_act_keyed["_key"].isin(unfc_combos["_key"]))
    ].drop(columns=["_key"]).copy()

    df_us = _act_keyed[
        (_act_keyed["region"] == "US+CA") &
        (_act_keyed["_key"].isin(unfc_combos["_key"]))
    ].drop(columns=["_key"]).copy()
    del _act_keyed  # keyed copy no longer needed

    _ly_keyed = act_full.copy()
    _ly_keyed["_key"] = (
        _ly_keyed["region"] + "||" +
        _ly_keyed["product"].astype(str).str.strip() + "||" +
        _ly_keyed["forecasted_shop"].astype(str).str.strip()
    )
    ly_agg = (_ly_keyed[
        (_ly_keyed["fulldate"].dt.year == t - 1) &
        (_ly_keyed["_key"].isin(unfc_combos["_key"]))
    ].groupby(["region", "product"], as_index=False)["actuals"].sum())
    del _ly_keyed  # keyed copy no longer needed

    def build_sg_region(calc_df, months, region_name, tag):
        if calc_df.empty or len(months) < 3:
            return pd.DataFrame(columns=["forecast_product", f"{tag}_LY", f"YoY_{tag}", f"Forecast_{tag}"])
        src = calc_df[~calc_df["product"].str.startswith("TOTAL")].reset_index(drop=True)
        ly = ly_agg.loc[ly_agg["region"] == region_name, ["product", "actuals"]].rename(
            columns={"actuals": f"{tag}_LY"})
        out = src[["product"]].merge(ly, on="product", how="left").fillna(0)
        out[f"YoY_{tag}"] = (
            0.30 * src["YTD % growth"].values
            + 0.20 * src["Last 3w % growth"].values
            + 0.25 * src.iloc[:, -1].values
            + 0.15 * src.iloc[:, -2].values
            + 0.10 * src.iloc[:, -3].values)
        out[f"Forecast_{tag}"] = (1 + out[f"YoY_{tag}"]) * out[f"{tag}_LY"]
        return out.rename(columns={"product": "forecast_product"})[
            ["forecast_product", f"{tag}_LY", f"YoY_{tag}", f"Forecast_{tag}"]]

    calc_eu, months_eu = build_calc_sheet(df_eu) if not df_eu.empty else (pd.DataFrame(), [])
    calc_us, months_us = build_calc_sheet(df_us) if not df_us.empty else (pd.DataFrame(), [])
    del df_eu, df_us  # only needed for calc sheets

    sg_eu = build_sg_region(calc_eu, months_eu, "EU+RoW", "EU")
    sg_us = build_sg_region(calc_us, months_us, "US+CA", "US")
    set_growth = sg_eu.merge(sg_us, on="forecast_product", how="outer").fillna(0)
    set_growth = set_growth.sort_values("forecast_product").reset_index(drop=True)
    del sg_eu, sg_us  # merged into set_growth

    act_fmt = act_full.copy()
    act_fmt["forecast_product"] = act_fmt["product"]
    act_fmt["destination_region"] = act_fmt["region"]
    act_fmt["forecasted_shop"] = act_fmt["forecasted_shop"].astype(str).str.strip()
    act_fmt["shoptype"] = act_fmt["shoptype"].astype(str).str.strip()
    act_fmt["format_clean"] = act_fmt["format_clean"].astype(str).str.strip()
    act_fmt["frame"] = act_fmt["frame"].astype(str).str.strip()
    act_fmt["yyyy-mm"] = act_fmt["_month"]
    act_fmt["_key"] = act_fmt["destination_region"] + "||" + act_fmt["forecast_product"] + "||" + act_fmt["forecasted_shop"]

    base_actuals = act_fmt[act_fmt["_key"].isin(unfc_combos["_key"])].copy()
    base_actuals.drop(columns=["_key"], inplace=True)

    # ── Compute shares ────────────────────────────────────────────────
    format_shares = pd.DataFrame()
    shoptype_shares = pd.DataFrame()
    dow_shares = pd.DataFrame()

    if not base_actuals.empty:
        max_date = base_actuals["fulldate"].max()
        current_period = max_date.to_period("M")
        last_full = current_period - 1
        last_n_months = [str(last_full - i) for i in range(N_MONTHS - 1, -1, -1)]
        q4_year = max_date.year - 1 if max_date.month <= 12 else max_date.year
        last_q4 = [f"{q4_year}-{m:02d}" for m in (10, 11, 12)]

        CHILD_DIMS = ["destination_region", "forecast_product", "frame", "format_clean"]
        PARENT_DIMS = ["forecast_product", "destination_region"]
        EXCLUDED_FMTS = ["EXCLUDED FORMAT", "NEW FORMAT"]

        format_actuals = base_actuals[
            ~base_actuals["forecast_product"].isin(ignore_formats["ignore_format"])].copy()

        format_shares_n = compute_shares(
            format_actuals, months=last_n_months, month_col="yyyy-mm",
            groupby_dims=CHILD_DIMS, total_by_dims=PARENT_DIMS,
            excluded=EXCLUDED_FMTS, exclude_col="format_clean",
        ).assign(avg_type="last N months")

        format_shares_q4 = compute_shares(
            format_actuals, months=last_q4, month_col="yyyy-mm",
            groupby_dims=CHILD_DIMS, total_by_dims=PARENT_DIMS,
            excluded=EXCLUDED_FMTS, exclude_col="format_clean",
        ).assign(avg_type="last Q4")
        del format_actuals  # no longer needed after shares computed

        format_shares = pd.concat([format_shares_n, format_shares_q4], ignore_index=True)
        del format_shares_n, format_shares_q4  # merged into format_shares

        if ("expected_share" in new_formats.columns
                and pd.to_numeric(new_formats["expected_share"], errors="coerce").notna().any()):
            new_fmts = new_formats[new_formats["expected_share"] != "x"].copy()
            new_fmts["expected_share"] = to_num(new_fmts["expected_share"])
            nf_n = new_fmts.assign(avg_type="last N months")
            nf_q4 = new_fmts.assign(avg_type="last Q4")
            new_fmts = (pd.concat([nf_n, nf_q4], ignore_index=True)
                        .drop_duplicates(["destination", "product", "format_clean", "frame", "avg_type"]))
            del nf_n, nf_q4  # merged into new_fmts

            format_shares = make_combo_key(format_shares,
                cols=["destination_region", "forecast_product", "format_clean", "frame", "avg_type"],
                new_col="match_key_new_formats")
            new_fmts = make_combo_key(new_fmts,
                cols=["destination", "product", "format_clean", "frame", "avg_type"],
                new_col="match_key_new_formats")
            format_shares = format_shares[
                ~format_shares["match_key_new_formats"].isin(new_fmts["match_key_new_formats"])]

            format_shares = make_combo_key(format_shares,
                cols=["destination_region", "forecast_product", "avg_type"],
                new_col="match_key_denominator_actuals")
            new_fmts = make_combo_key(new_fmts,
                cols=["destination", "product", "avg_type"],
                new_col="match_key_denominator_actuals")
            denom = (format_shares[["match_key_denominator_actuals", "denominator_actuals"]]
                    .drop_duplicates(subset=["match_key_denominator_actuals"]))
            new_fmts = new_fmts.merge(denom, how="left", on="match_key_denominator_actuals")
            del denom  # merged into new_fmts
            new_fmts["expected_actuals"] = new_fmts["expected_share"] * new_fmts["denominator_actuals"]

            format_shares = (format_shares
                            .groupby(CHILD_DIMS + ["avg_type"], as_index=False)
                            .agg({"actuals": "sum"}))
            nf_agg = (new_fmts
                    .groupby(["destination", "product", "format_clean", "frame", "avg_type"], as_index=False)
                    .agg({"expected_actuals": "sum"})
                    .rename(columns={"destination": "destination_region",
                                    "product": "forecast_product",
                                    "expected_actuals": "actuals"}))
            del new_fmts  # aggregated into nf_agg
            format_shares = pd.concat([format_shares, nf_agg], ignore_index=True)
            del nf_agg  # merged into format_shares

            fs_n = compute_shares(
                format_shares[format_shares["avg_type"] == "last N months"],
                groupby_dims=CHILD_DIMS, total_by_dims=PARENT_DIMS,
            ).assign(avg_type="last N months")
            fs_q4 = compute_shares(
                format_shares[format_shares["avg_type"] == "last Q4"],
                groupby_dims=CHILD_DIMS, total_by_dims=PARENT_DIMS,
            ).assign(avg_type="last Q4")
            format_shares = pd.concat([fs_n, fs_q4], ignore_index=True)
            del fs_n, fs_q4  # merged into format_shares

        forced_format, forced_frame = "na", "na"
        ignored_base = (base_actuals
                        .groupby(["destination_region", "forecast_product", "format_clean", "frame"],
                                as_index=False).agg({"actuals": "sum"}))
        ignored_rows = ignored_base[ignored_base["forecast_product"].isin(ignore_formats["ignore_format"])]
        del ignored_base  # only needed for ignored_rows filter
        if not ignored_rows.empty:
            for avg_type in ("last N months", "last Q4"):
                block = ignored_rows.assign(frame=forced_frame, format_clean=forced_format, share=1.0, avg_type=avg_type)
                format_shares = pd.concat([format_shares, block], ignore_index=True)
            format_shares = (format_shares
                            .groupby(CHILD_DIMS + ["avg_type"], as_index=False)
                            .agg({"share": "sum"}))
        del ignored_rows  # used only in the if block above

        KEY_COLS = ["destination_region", "forecast_product", "format_clean", "frame"]
        sc = share_check(format_shares, key_cols=KEY_COLS)
        if not sc.empty:
            only_last_n = sc["status"] == "only_last_n"
            only_q4 = sc["status"] == "only_q4"
            to_add_q4 = sc.loc[only_last_n, KEY_COLS].assign(
                avg_type="last Q4", share=sc.loc[only_last_n, "share_last_n"].values)
            to_add_n = sc.loc[only_q4, KEY_COLS].assign(
                avg_type="last N months", share=sc.loc[only_q4, "share_q4"].values)
            to_apply = pd.concat([to_add_q4, to_add_n], ignore_index=True)
            del to_add_q4, to_add_n  # merged into to_apply
            format_shares = (
                pd.concat([format_shares, to_apply], ignore_index=True)
                .groupby(CHILD_DIMS + ["avg_type"], as_index=False, dropna=False)
                .agg(share=("share", "max")))
            del to_apply
        del sc

        out_cols = ["destination_region", "forecast_product", "frame", "format_clean", "share", "avg_type"]
        out_cols = [c for c in out_cols if c in format_shares.columns]
        format_shares = (format_shares[out_cols]
                        .sort_values(["destination_region", "forecast_product", "format_clean", "avg_type"])
                        .reset_index(drop=True))

        shoptype_shares = compute_shares(
            df=base_actuals, months=last_n_months, month_col="yyyy-mm",
            groupby_dims=["destination_region", "forecast_product", "shoptype"],
            total_by_dims=["destination_region", "forecast_product"])
        out_cols = ["destination_region", "forecast_product", "shoptype", "share"]
        out_cols = [c for c in out_cols if c in shoptype_shares.columns]
        shoptype_shares = (shoptype_shares[out_cols]
                        .sort_values(["destination_region", "forecast_product", "shoptype"])
                        .reset_index(drop=True))

        daily = base_actuals[["fulldate", "actuals"]].copy()
        daily["iso_dow"] = daily["fulldate"].dt.isocalendar().day.astype(int)
        daily["dow_name"] = daily["fulldate"].dt.day_name()
        iso = daily["fulldate"].dt.isocalendar()
        daily["week"] = [f"{y}-{int(w):02d}" for y, w in zip(iso.year, iso.week)]

        daily_agg = (daily.groupby(["week", "iso_dow", "dow_name"], as_index=False)
                    .agg({"actuals": "sum"}))
        del daily  # aggregated into daily_agg
        daily_agg["week_total"] = daily_agg.groupby("week")["actuals"].transform("sum")
        daily_agg["dow_share"] = np.where(daily_agg["week_total"] > 0,
                                        daily_agg["actuals"] / daily_agg["week_total"], 0.0)

        aw = anchor_weeks[["week", "anchor_week"]].copy()
        aw["anchor_week"] = aw.apply(
            lambda row: row["week"].replace("2026", "2025")
            if str(row["anchor_week"]).strip() == "x" else row["anchor_week"], axis=1)

        anchor_dow = (daily_agg[["week", "iso_dow", "dow_name", "dow_share"]]
                    .rename(columns={"week": "anchor_week"}))
        del daily_agg  # only needed for anchor_dow
        dow_shares = (aw[["week", "anchor_week"]]
                    .merge(anchor_dow, on="anchor_week", how="inner"))
        del anchor_dow  # merged into dow_shares
        dow_shares = (dow_shares[["week", "anchor_week", "iso_dow", "dow_name", "dow_share"]]
                    .sort_values(["week", "iso_dow"]).reset_index(drop=True))

    # ── Weekly forecast ───────────────────────────────────────────────
    growth_long = []
    for tag, reg in [("EU", "EU+RoW"), ("US", "US+CA")]:
        sub = set_growth[["forecast_product", f"Forecast_{tag}", f"YoY_{tag}"]].rename(
            columns={f"Forecast_{tag}": "fc_yyyy", f"YoY_{tag}": "yoy"})
        sub["destination_region"] = reg
        growth_long.append(sub)
    growth_df = pd.concat(growth_long, ignore_index=True)
    del growth_long  # merged into growth_df
    growth_df = growth_df[growth_df["fc_yyyy"] != 0].reset_index(drop=True)

    all_weekly_actuals = (act_fmt
        .groupby(["destination_region", "forecast_product", "week"], as_index=False)
        .agg({"actuals": "sum"}))

    remaining_weekly_actuals = (base_actuals
        .groupby(["destination_region", "forecast_product", "week"], as_index=False)
        .agg({"actuals": "sum"}))

    weekly_actuals = remaining_weekly_actuals[
        remaining_weekly_actuals["forecast_product"].isin(growth_df["forecast_product"].unique())
    ].copy()
    del remaining_weekly_actuals  # filtered into weekly_actuals

    weeks_list = [f"{t}-{i:02d}" for i in range(1, 54)]
    scaffold = (pd.DataFrame({"week": weeks_list})
                .merge(growth_df[["forecast_product", "destination_region"]].drop_duplicates(), how="cross"))

    aw = anchor_weeks[["week", "anchor_week"]].copy()
    scaffold = scaffold.merge(aw, on="week", how="left")
    scaffold["week_classified"] = np.where(scaffold["anchor_week"] != "x", "fc week", "dist. week")
    scaffold["anchor_week"] = np.where(
        scaffold["anchor_week"] == "x",
        scaffold["week"].str.replace(str(t), str(t - 1), n=1),
        scaffold["anchor_week"])

    wa_r = weekly_actuals.rename(columns={"week": "anchor_week"})
    scaffold = make_combo_key(scaffold, ["forecast_product", "destination_region", "anchor_week"], "mk_act")
    wa_r = make_combo_key(wa_r, ["forecast_product", "destination_region", "anchor_week"], "mk_act")
    scaffold = scaffold.merge(wa_r[["mk_act", "actuals"]], on="mk_act", how="left").fillna({"actuals": 0})
    scaffold.drop(columns=["mk_act"], inplace=True)
    del wa_r  # merged into scaffold

    scaffold = scaffold.merge(growth_df, on=["forecast_product", "destination_region"], how="left").fillna(0)

    scaffold["FQTY_type"] = np.where(scaffold["week"].apply(get_week_num) < actuals_end, "actuals", "forecast")

    is_fc = scaffold["week_classified"] == "fc week"
    anchor_is_ly = scaffold["anchor_week"].str.startswith(str(t - 1))
    anchor_is_ty = scaffold["anchor_week"].str.startswith(str(t))
    has_actuals = scaffold["actuals"] > 0

    scaffold["_ly_week"] = scaffold["week"].str.replace(str(t), str(t - 1), n=1)
    scaffold = make_combo_key(scaffold, ["forecast_product", "destination_region", "_ly_week"], "_mk_ly")
    wa_ly = weekly_actuals.rename(columns={"week": "_ly_week"})
    wa_ly = make_combo_key(wa_ly, ["forecast_product", "destination_region", "_ly_week"], "_mk_ly")
    scaffold = scaffold.merge(
        wa_ly[["_mk_ly", "actuals"]].rename(columns={"actuals": "_ly_actuals"}),
        on="_mk_ly", how="left").fillna({"_ly_actuals": 0})
    scaffold.drop(columns=["_mk_ly", "_ly_week"], inplace=True)
    del wa_ly, weekly_actuals  # merged into scaffold; weekly_actuals fully consumed

    scaffold["yoy_adj"] = np.where(is_fc & anchor_is_ly, scaffold["yoy"], 0.0)
    needs_fallback = is_fc & anchor_is_ty & ~has_actuals & (scaffold["FQTY_type"] == "forecast")
    scaffold["yoy_adj"] = np.where(needs_fallback, scaffold["yoy"], scaffold["yoy_adj"])
    scaffold["actuals"] = np.where(needs_fallback, scaffold["_ly_actuals"], scaffold["actuals"])
    scaffold.drop(columns=["_ly_actuals"], inplace=True)

    scaffold["weekly_fc"] = np.where(
        scaffold["week_classified"] == "fc week",
        scaffold["actuals"] * (1 + scaffold["yoy_adj"]), 0.0)
    scaffold["sum_weekly_fc"] = scaffold.groupby(
        ["forecast_product", "destination_region"])["weekly_fc"].transform("sum")
    scaffold["remaining_fc"] = scaffold["fc_yyyy"] - scaffold["sum_weekly_fc"]

    scaffold["weekly_dist"] = np.where(scaffold["week_classified"] == "dist. week", scaffold["actuals"], 0.0)
    scaffold["sum_weekly_dist"] = scaffold.groupby(
        ["forecast_product", "destination_region"])["weekly_dist"].transform("sum")
    scaffold["remaining_fc_dist"] = np.where(
        scaffold["sum_weekly_dist"] > 0,
        scaffold["weekly_dist"] / scaffold["sum_weekly_dist"], 0.0)

    scaffold["FQTY_weekly"] = scaffold["weekly_fc"] + scaffold["remaining_fc"] * scaffold["remaining_fc_dist"]

    scaffold_fc = scaffold[scaffold["FQTY_type"] == "forecast"].copy()
    scaffold_fc["_fc_total"] = scaffold_fc.groupby(
        ["forecast_product", "destination_region"])["FQTY_weekly"].transform("sum")
    scaffold_fc["week_share"] = np.where(
        scaffold_fc["_fc_total"] > 0,
        scaffold_fc["FQTY_weekly"] / scaffold_fc["_fc_total"], 0.0)

    weekly_fc_shares = scaffold_fc[
        ["week", "destination_region", "forecast_product", "week_share"]
    ].copy()
    del scaffold_fc  # extracted into weekly_fc_shares

    date_range = pd.date_range(f"{t}-01-01", f"{t}-12-31", freq="D")
    daily_cal = pd.DataFrame({"date": date_range})
    daily_cal["yyyy-mm"] = daily_cal["date"].dt.to_period("M").astype(str)
    iso_cal = daily_cal["date"].dt.isocalendar()
    daily_cal["week"], _max_wk_t = iso_weeks_clamped(daily_cal["date"], t)
    daily_cal["iso_dow"] = iso_cal.day.astype(int)
    daily_cal["dow_name"] = daily_cal["date"].dt.day_name()

    week_to_month = (daily_cal.groupby("week")["yyyy-mm"]
                    .agg(lambda x: x.mode().iloc[0])
                    .reset_index()
                    .rename(columns={"yyyy-mm": "week_month"}))

    sep_fc = fc[["region", "product", "shoptype", "forecasted_shop",
                "format", "frame", "period", "forecast"]].copy()
    del fc  # all needed columns extracted into sep_fc
    sep_fc.columns = ["destination_region", "forecast_product", "shoptype",
                    "forecasted_shop", "format_clean", "frame", "yyyy-mm", "FQTY"]
    sep_fc["FQTY"] = to_num(sep_fc["FQTY"])
    sep_fc["forecasted_shop"] = sep_fc["forecasted_shop"].astype(str).str.strip()
    sep_fc["format_clean"] = sep_fc["format_clean"].astype(str).str.strip()
    sep_fc["frame"] = sep_fc["frame"].astype(str).str.strip()
    sep_fc = (sep_fc
            .groupby(["destination_region", "forecast_product", "shoptype",
                        "forecasted_shop", "format_clean", "frame", "yyyy-mm"], as_index=False)
            .agg({"FQTY": "sum"}))

    sub_dist = substitute_distribution.copy()
    sub_dist.columns = [str(c).strip() for c in sub_dist.columns]
    sub_mapping = (sub_dist[["Product", "Destination", "Used_distribution"]]
                .drop_duplicates()
                .rename(columns={"Product": "forecast_product",
                                    "Destination": "destination_region"}))
    del sub_dist  # only needed for sub_mapping

    scaffold_products = set(
        growth_df[["forecast_product", "destination_region"]]
        .apply(lambda r: (r["forecast_product"], r["destination_region"]), axis=1))
    missing_subs = set()
    for _, row in sub_mapping.iterrows():
        sub_key = (row["Used_distribution"], row["destination_region"])
        if sub_key not in scaffold_products:
            missing_subs.add(sub_key)
    del scaffold_products  # only needed to populate missing_subs

    if missing_subs:
        missing_sub_df = pd.DataFrame(list(missing_subs),
                                    columns=["forecast_product", "destination_region"])

        aw_miss = anchor_weeks[["week", "anchor_week"]].copy()
        aw_miss["anchor_week"] = np.where(
            aw_miss["anchor_week"].astype(str).str.strip() == "x",
            aw_miss["week"].str.replace(str(t), str(t - 1), n=1),
            aw_miss["anchor_week"])
        aw_miss = aw_miss.merge(week_to_month, on="week", how="left")
        aw_miss = aw_miss.merge(missing_sub_df, how="cross")

        aw_miss = aw_miss.merge(
            all_weekly_actuals.rename(columns={"week": "anchor_week", "actuals": "ly_weekly"}),
            on=["anchor_week", "destination_region", "forecast_product"],
            how="left")
        aw_miss["ly_weekly"] = aw_miss["ly_weekly"].fillna(0)

        aw_miss["month_total"] = aw_miss.groupby(
            ["week_month", "destination_region", "forecast_product"])["ly_weekly"].transform("sum")
        n_wk = aw_miss.groupby(
            ["week_month", "destination_region", "forecast_product"])["week"].transform("count")
        no_data = aw_miss["month_total"] == 0
        aw_miss["wk_share_in_month"] = np.where(
            ~no_data, aw_miss["ly_weekly"] / aw_miss["month_total"], 1.0 / n_wk)

        sep_fc_monthly = (sep_fc[sep_fc["forecast_product"].isin(missing_sub_df["forecast_product"])]
            .groupby(["destination_region", "forecast_product", "yyyy-mm"], as_index=False)
            .agg({"FQTY": "sum"}))

        aw_miss = aw_miss.merge(
            sep_fc_monthly,
            left_on=["week_month", "destination_region", "forecast_product"],
            right_on=["yyyy-mm", "destination_region", "forecast_product"],
            how="left")
        del sep_fc_monthly  # merged into aw_miss
        aw_miss["FQTY"] = aw_miss["FQTY"].fillna(0)
        aw_miss["FQTY_weekly"] = aw_miss["FQTY"] * aw_miss["wk_share_in_month"]

        aw_miss["is_fc"] = aw_miss["week"].apply(get_week_num) > actuals_end
        fc_miss = aw_miss[aw_miss["is_fc"]].copy()
        del aw_miss  # filtered into fc_miss
        fc_miss["_total"] = fc_miss.groupby(
            ["forecast_product", "destination_region"])["FQTY_weekly"].transform("sum")
        fc_miss["week_share"] = np.where(
            fc_miss["_total"] > 0, fc_miss["FQTY_weekly"] / fc_miss["_total"], 0.0)

        mini_shares = fc_miss[["week", "destination_region", "forecast_product", "week_share"]].copy()
        del fc_miss  # extracted into mini_shares
        weekly_fc_shares = pd.concat([weekly_fc_shares, mini_shares], ignore_index=True)
        del mini_shares  # merged into weekly_fc_shares

    special_products = sub_mapping.merge(
        growth_df[["forecast_product", "destination_region", "fc_yyyy"]],
        on=["forecast_product", "destination_region"], how="inner")

    if not special_products.empty:
        sp_keys = set(
            special_products.apply(
                lambda r: (r["forecast_product"], r["destination_region"]), axis=1))

        sp_scaffold = scaffold[scaffold.apply(
            lambda r: (r["forecast_product"], r["destination_region"]) in sp_keys, axis=1
        )].copy()

        sp_actuals_rows = sp_scaffold[sp_scaffold["FQTY_type"] == "actuals"].copy()
        sp_actuals_rows["FQTY"] = sp_actuals_rows["FQTY_weekly"]

        sp_act_sum = (sp_actuals_rows
            .groupby(["forecast_product", "destination_region"], as_index=False)
            ["FQTY"].sum()
            .rename(columns={"FQTY": "actuals_total"}))
        sp_fc = special_products.merge(
            sp_act_sum, on=["forecast_product", "destination_region"], how="left")
        del sp_act_sum  # merged into sp_fc
        sp_fc["actuals_total"] = sp_fc["actuals_total"].fillna(0)
        sp_fc["remaining_fc"] = sp_fc["fc_yyyy"] - sp_fc["actuals_total"]

        sub_shares = weekly_fc_shares.rename(columns={
            "forecast_product": "Used_distribution",
            "week_share": "sub_week_share"})

        sp_forecast_rows = sp_fc.merge(
            sub_shares, on=["Used_distribution", "destination_region"], how="inner")
        del sp_fc, sub_shares  # consumed into sp_forecast_rows
        sp_forecast_rows["FQTY"] = sp_forecast_rows["remaining_fc"] * sp_forecast_rows["sub_week_share"]
        sp_forecast_rows["FQTY_type"] = "forecast"

        sp_weekly_final = pd.concat([
            sp_actuals_rows[["week", "destination_region", "forecast_product", "FQTY_type", "FQTY"]],
            sp_forecast_rows[["week", "destination_region", "forecast_product", "FQTY_type", "FQTY"]],
        ], ignore_index=True)
        del sp_scaffold, sp_actuals_rows, sp_forecast_rows  # merged into sp_weekly_final
    else:
        sp_weekly_final = pd.DataFrame(columns=["week", "destination_region",
                                                "forecast_product", "FQTY_type", "FQTY"])

    special_keys = set()
    if not special_products.empty:
        special_keys = set(
            special_products.apply(
                lambda r: (r["forecast_product"], r["destination_region"]), axis=1))
    del special_products  # only needed for special_keys

    scaffold["_is_special"] = scaffold.apply(
        lambda r: (r["forecast_product"], r["destination_region"]) in special_keys, axis=1)

    non_special_weekly = (
        scaffold[~scaffold["_is_special"]]
        .groupby(["week", "destination_region", "forecast_product", "FQTY_type"], as_index=False)
        .agg({"FQTY_weekly": "sum"})
        .rename(columns={"FQTY_weekly": "FQTY"}))

    weekly_fc_final = pd.concat([non_special_weekly, sp_weekly_final], ignore_index=True)
    del scaffold, non_special_weekly, sp_weekly_final  # all merged into weekly_fc_final

    # ── Daily forecast ────────────────────────────────────────────────
    wf_forecast = weekly_fc_final[weekly_fc_final["FQTY_type"] == "forecast"].copy()
    del weekly_fc_final  # split into wf_forecast (actuals path unused downstream)
    wf_forecast = wf_forecast[wf_forecast["FQTY"] != 0]

    base_fc_daily = daily_cal.merge(
        wf_forecast[["week", "destination_region", "forecast_product", "FQTY"]],
        on="week", how="inner")
    del wf_forecast  # merged into base_fc_daily
    base_fc_daily = base_fc_daily.rename(columns={"FQTY": "FQTY_Weekly"})

    if not dow_shares.empty:
        base_fc_daily = base_fc_daily.merge(
            dow_shares[["week", "iso_dow", "dow_share"]],
            on=["week", "iso_dow"], how="left")
    else:
        base_fc_daily["dow_share"] = 1.0 / 7
    base_fc_daily["dow_share"] = base_fc_daily["dow_share"].fillna(1.0 / 7)
    base_fc_daily["daily_FQTY"] = base_fc_daily["FQTY_Weekly"] * base_fc_daily["dow_share"]

    if not pcs_shares.empty:
        base_fc_daily = base_fc_daily.merge(
            pcs_shares.rename(columns={"PRODUCT": "forecast_product", "year_month": "yyyy-mm"}),
            on=["destination_region", "forecast_product", "yyyy-mm"], how="left")
        no_pcs = base_fc_daily["pcs_share"].isna()
        base_fc_daily.loc[no_pcs, "pcs"] = "na"
        base_fc_daily.loc[no_pcs, "pcs_share"] = 1.0
        base_fc_daily["pcs_share"] = to_num(base_fc_daily["pcs_share"])
        base_fc_daily["FQTY_pcs"] = base_fc_daily["daily_FQTY"] * base_fc_daily["pcs_share"]
        base_fc_daily = base_fc_daily.loc[base_fc_daily["FQTY_pcs"] != 0].reset_index(drop=True)
    else:
        base_fc_daily["pcs"] = "na"
        base_fc_daily["FQTY_pcs"] = base_fc_daily["daily_FQTY"]

    if not format_shares.empty:
        base_fc_daily["distribution_type"] = np.where(
            base_fc_daily["date"].dt.quarter == 4, "last Q4", "last N months")
        base_fc_daily = make_combo_key(base_fc_daily,
            cols=["destination_region", "forecast_product", "distribution_type"],
            new_col="match_key_format")
        fs = format_shares.copy()
        fs = make_combo_key(fs,
            cols=["destination_region", "forecast_product", "avg_type"],
            new_col="match_key_format")
        base_fc_daily = base_fc_daily.merge(
            fs[["match_key_format", "share", "format_clean", "frame"]],
            on="match_key_format", how="left")
        del fs  # merged into base_fc_daily
        base_fc_daily.rename(columns={"share": "format_share"}, inplace=True)
        base_fc_daily["format_share"] = base_fc_daily["format_share"].fillna(1.0)
        base_fc_daily["FQTY_format"] = base_fc_daily["FQTY_pcs"] * base_fc_daily["format_share"]
    else:
        base_fc_daily["format_clean"] = "na"
        base_fc_daily["frame"] = "na"
        base_fc_daily["FQTY_format"] = base_fc_daily["FQTY_pcs"]

    if not shoptype_shares.empty:
        base_fc_daily = base_fc_daily.merge(
            shoptype_shares.rename(columns={"share": "shoptype_share"}),
            on=["destination_region", "forecast_product"], how="left")
        base_fc_daily["shoptype_share"] = base_fc_daily["shoptype_share"].fillna(1.0)
        base_fc_daily["FQTY"] = base_fc_daily["FQTY_format"] * base_fc_daily["shoptype_share"]
    else:
        base_fc_daily["shoptype"] = "na"
        base_fc_daily["FQTY"] = base_fc_daily["FQTY_format"]

    base_fc_daily = (base_fc_daily
        .groupby(["date", "yyyy-mm", "destination_region", "pcs", "shoptype",
                "forecast_product", "format_clean", "frame"], as_index=False)
        .agg({"FQTY": "sum"}))

    _last_actual = base_actuals["fulldate"].max()
    base_fc_daily = base_fc_daily[base_fc_daily["date"] > _last_actual].reset_index(drop=True)
    min_fc_date = base_fc_daily["date"].min()
    year_start = pd.to_datetime(f"{t-1}-01-01")

    base_actuals["fulldate"] = pd.to_datetime(base_actuals["fulldate"])
    base_actuals_ytd = base_actuals[
        (base_actuals["fulldate"] >= year_start) &
        (base_actuals["fulldate"] < min_fc_date)
    ].groupby(["fulldate", "yyyy-mm", "destination_region", "pcs_region",
            "forecast_product", "format_clean", "frame", "shoptype"]
            )["actuals"].sum().reset_index()
    base_actuals_ytd = base_actuals_ytd.rename(columns={
        "fulldate": "date", "actuals": "FQTY", "pcs_region": "pcs"})

    if "shoptype" not in base_actuals_ytd.columns:
        base_actuals_ytd["shoptype"] = "na"
    base_actuals_ytd = base_actuals_ytd[["date", "yyyy-mm", "destination_region", "pcs",
                                        "shoptype", "forecast_product", "format_clean", "frame", "FQTY"]]
    base_fc_daily = pd.concat([base_actuals_ytd, base_fc_daily], ignore_index=True)
    del base_actuals_ytd  # prepended into base_fc_daily

    # ── Separate forecast daily distribution ──────────────────────────
    aw_sep = anchor_weeks[["week", "anchor_week"]].copy()
    aw_sep["anchor_week"] = np.where(
        aw_sep["anchor_week"].astype(str).str.strip() == "x",
        aw_sep["week"].str.replace(str(t), str(t - 1), n=1),
        aw_sep["anchor_week"])

    aw_sep = aw_sep.merge(week_to_month, on="week", how="left")

    sep_prods = sep_fc[["destination_region", "forecast_product"]].drop_duplicates()
    aw_sep = aw_sep.merge(sep_prods, how="cross")
    del sep_prods  # merged into aw_sep

    # ── LY actuals per anchor week (full-week totals) ──────────────────
    all_prod_weekly = (act_fmt
        .groupby(["destination_region", "forecast_product", "week"], as_index=False)
        .agg({"actuals": "sum"}))
    del act_fmt  # all groupby aggregations done

    # ── Day counts per (week, month) from the forecast-year calendar ──
    # For boundary weeks this produces 2 rows; for non-boundary weeks, 1.
    _day_counts = (daily_cal.groupby(["week", "yyyy-mm"]).size()
                   .reset_index(name="n_days")
                   .rename(columns={"yyyy-mm": "cal_month"}))
    _day_counts["week_days"] = _day_counts.groupby("week")["n_days"].transform("sum")
    _day_counts["day_frac"] = _day_counts["n_days"] / _day_counts["week_days"]

    # ── Expand aw_sep by forecast-year calendar months ────────────────
    # Each week gets one row per month it spans. Boundary weeks get 2 rows.
    aw_sep = aw_sep.drop(columns=["week_month"])
    aw_sep = aw_sep.merge(_day_counts[["week", "cal_month", "n_days", "day_frac"]],
                          on="week", how="inner")

    # Join full-week LY actuals via anchor_week, then split by day_frac.
    # This avoids the anchor-year-shift problem: even if the anchor week
    # spans different months than the forecast week, we split the LY total
    # using the forecast week's calendar boundaries.
    aw_sep = aw_sep.merge(
        all_prod_weekly.rename(columns={"week": "anchor_week", "actuals": "ly_weekly"}),
        on=["anchor_week", "destination_region", "forecast_product"],
        how="left")
    del all_prod_weekly
    aw_sep["ly_weekly"] = aw_sep["ly_weekly"].fillna(0)
    aw_sep["ly_portion"] = aw_sep["ly_weekly"] * aw_sep["day_frac"]

    aw_sep["week_month"] = aw_sep["cal_month"]

    del _day_counts

    # Compute week_share per calendar month: ly_portion / month_total.
    # Now month_total only includes portions of weeks that fall in that month.
    aw_sep["month_total"] = aw_sep.groupby(
        ["week_month", "destination_region", "forecast_product"])["ly_portion"].transform("sum")
    n_rows = aw_sep.groupby(
        ["week_month", "destination_region", "forecast_product"])["week"].transform("count")
    no_data = aw_sep["month_total"] == 0
    aw_sep["week_share"] = np.where(
        ~no_data,
        aw_sep["ly_portion"] / aw_sep["month_total"],
        aw_sep["n_days"] / aw_sep.groupby(
            ["week_month", "destination_region", "forecast_product"])["n_days"].transform("sum").clip(lower=1))

    sep_weekly = sep_fc.merge(
        aw_sep[["week", "week_month", "destination_region", "forecast_product", "week_share"]],
        left_on=["yyyy-mm", "destination_region", "forecast_product"],
        right_on=["week_month", "destination_region", "forecast_product"],
        how="inner")
    del aw_sep  # merged into sep_weekly
    sep_weekly["FQTY_weekly"] = sep_weekly["FQTY"] * sep_weekly["week_share"]

    sep_daily = sep_weekly.merge(
        daily_cal[["date", "week", "iso_dow"]],
        on="week", how="inner")
    del sep_weekly  # merged into sep_daily

    if not dow_shares.empty:
        sep_daily = sep_daily.merge(
            dow_shares[["week", "iso_dow", "dow_share"]], on=["week", "iso_dow"], how="left")
    else:
        # FIX (Issue 1): default to 1/7 to evenly split weekly FQTY across 7 days.
        # Previous default 1.0 caused 7x inflation when summed back to monthly,
        # because each of 7 days got the full weekly value.
        sep_daily["dow_share"] = 1.0 / 7
    sep_daily["dow_share"] = sep_daily["dow_share"].fillna(1.0 / 7)
    sep_daily["FQTY_daily"] = sep_daily["FQTY_weekly"] * sep_daily["dow_share"]
    sep_daily = sep_daily[["date", "yyyy-mm", "week", "destination_region", "forecast_product",
                        "shoptype", "forecasted_shop", "format_clean", "frame", "FQTY_daily"]]

    if not pcs_shares.empty:
        sep_daily["year_month"] = sep_daily["date"].dt.to_period("M").astype(str)
        sep_daily = sep_daily.merge(
            pcs_shares.rename(columns={"PRODUCT": "forecast_product", "year_month": "year_month"}),
            on=["destination_region", "forecast_product", "year_month"], how="left")
        # FIX (Issue 2): When pcs_shares has no entry for a (region, product, month),
        # the merge produces NaN. Previous logic ran to_num (which converts NaN→0)
        # then multiplied to get FQTY=0 → row dropped entirely. This wiped out
        # forecasts for products like sendmoments+Postcard Digital that didn't have
        # pcs_share coverage. Fix: mirror the base_fc_daily logic — fill NaN
        # pcs_share with 1.0 and set pcs="na" so the row passes through.
        no_pcs = sep_daily["pcs_share"].isna()
        sep_daily.loc[no_pcs, "pcs"] = "na"
        sep_daily.loc[no_pcs, "pcs_share"] = 1.0
        sep_daily["pcs_share"] = to_num(sep_daily["pcs_share"])
        sep_daily["FQTY"] = sep_daily["FQTY_daily"] * sep_daily["pcs_share"]
        sep_daily = sep_daily.loc[sep_daily["FQTY"] != 0].reset_index(drop=True)
        sep_daily.drop(columns=["year_month", "pcs_share", "FQTY_daily"], inplace=True)
    else:
        sep_daily["pcs"] = "na"
        sep_daily = sep_daily.rename(columns={"FQTY_daily": "FQTY"})

    sep_daily = sep_daily[pd.to_datetime(sep_daily["date"]) >= min_fc_date].reset_index(drop=True)

    # ── Scale current month forecast so actuals + forecast = input ────
    fc_rows = actuals_combos[actuals_combos["forecasted"] == "Yes"]
    fc_keys = set(
        fc_rows["region"] + "||" + fc_rows["product"] + "||" + fc_rows["shop"])
    del fc_rows  # only needed for fc_keys

    _sep_fc_shops = set(
        sep_fc["destination_region"] + "||" +
        sep_fc["forecast_product"] + "||" +
        sep_fc["forecasted_shop"])

    sep_act = act_full.copy()
    sep_act["destination_region"] = sep_act["region"]
    sep_act["forecast_product"] = sep_act["product"]
    sep_act["forecasted_shop"] = sep_act["forecasted_shop"].astype(str).str.strip()
    sep_act["_cov_key"] = (
        sep_act["region"] + "||" + sep_act["product"] + "||" + sep_act["forecasted_shop"])
    sep_act = sep_act[sep_act["_cov_key"].isin(fc_keys)].copy()
    sep_act.drop(columns=["_cov_key"], inplace=True)
    del fc_keys  # used only for the filter above

    sep_act["shoptype"] = sep_act["shoptype"].astype(str).str.strip()
    sep_act["_shop_fc_key"] = (
        sep_act["destination_region"] + "||" +
        sep_act["forecast_product"] + "||" +
        sep_act["forecasted_shop"])

    _no_sep_fc = ~sep_act["_shop_fc_key"].isin(_sep_fc_shops)
    sep_act["_other_shop"] = "Other " + sep_act["shoptype"]
    sep_act["_other_fc_key"] = (
        sep_act["destination_region"] + "||" +
        sep_act["forecast_product"] + "||" +
        sep_act["_other_shop"])
    _has_other = sep_act["_other_fc_key"].isin(_sep_fc_shops)
    del _sep_fc_shops  # used only for the two boolean masks above

    sep_act.loc[_no_sep_fc & _has_other, "forecasted_shop"] = (
        sep_act.loc[_no_sep_fc & _has_other, "_other_shop"])
    sep_act.loc[_no_sep_fc & ~_has_other, "forecasted_shop"] = (
        sep_act.loc[_no_sep_fc & ~_has_other, "shoptype"])
    sep_act.drop(columns=["_shop_fc_key", "_other_shop", "_other_fc_key"], inplace=True)

    sep_act["fulldate"] = pd.to_datetime(sep_act["fulldate"])

    # Current-month scaling removed: forecast.csv now shows actuals for
    # dates up to the cutoff and unscaled forecast for dates after.
    # YTD actuals are already prepended (base_actuals_ytd / sep_act_ytd),
    # so the output naturally transitions from actuals → forecast at the
    # cutoff boundary without any artificial adjustment.
    del base_actuals  # all uses complete (ytd prepend)

    # ── Append separate YTD actuals ───────────────────────────────────
    sep_act_ytd = sep_act[
        (sep_act["fulldate"] >= year_start) &
        (sep_act["fulldate"] < min_fc_date)]

    if not sep_act_ytd.empty:
        sep_act_ytd = sep_act_ytd.copy()
        sep_act_ytd["format_clean"] = sep_act_ytd["format_clean"].astype(str).str.strip()
        sep_act_ytd["frame"] = sep_act_ytd["frame"].astype(str).str.strip()
        sep_act_ytd["shoptype"] = sep_act_ytd["shoptype"].astype(str).str.strip()
        sep_act_ytd["yyyy-mm"] = sep_act_ytd["fulldate"].dt.to_period("M").astype(str)
        sep_act_ytd["week"], _ = iso_weeks_clamped(sep_act_ytd["fulldate"], t)

        grp_cols = ["fulldate", "yyyy-mm", "week", "destination_region", "forecasted_shop",
                    "shoptype", "forecast_product", "format_clean", "frame"]
        if "pcs_region" in sep_act_ytd.columns:
            grp_cols.append("pcs_region")
        sep_ytd = sep_act_ytd.groupby(grp_cols, as_index=False)["actuals"].sum()
        sep_ytd = sep_ytd.rename(columns={"fulldate": "date", "actuals": "FQTY"})
        if "pcs_region" in sep_ytd.columns:
            sep_ytd = sep_ytd.rename(columns={"pcs_region": "pcs"})
        sep_daily = pd.concat([sep_ytd, sep_daily], ignore_index=True)
        del sep_ytd  # prepended into sep_daily
    del sep_act_ytd, sep_act  # all uses complete

    # ── Combine remaining + separate → forecast.csv ───────────────────
    base_fc_daily["source"] = "remaining"
    sep_daily["source"] = "separate"

    base_fc_daily["date"] = pd.to_datetime(base_fc_daily["date"]).dt.strftime("%Y-%m-%d")
    sep_daily["date"] = pd.to_datetime(sep_daily["date"]).dt.strftime("%Y-%m-%d")

    base_fc_daily = (base_fc_daily
        .groupby(["date", "yyyy-mm", "destination_region", "pcs", "shoptype",
                "forecast_product", "format_clean", "frame", "source"], as_index=False)
        .agg({"FQTY": "sum"}))

    sep_daily = (sep_daily
        .groupby([c for c in ["date", "yyyy-mm", "destination_region", "pcs", "shoptype",
                            "forecasted_shop", "forecast_product", "format_clean", "frame", "source"]
                if c in sep_daily.columns], as_index=False)
        .agg({"FQTY": "sum"}))

    forecast_all = pd.concat([base_fc_daily, sep_daily], ignore_index=True)
    del base_fc_daily, sep_daily  # merged into forecast_all
    forecast_all = forecast_all.loc[
        (forecast_all["FQTY"] != 0) &
        (forecast_all["date"] >= f"{t-1}-01-01") &
        (forecast_all["date"] <= f"{t}-12-31")]

    _periods = pd.to_datetime(forecast_all["yyyy-mm"])
    forecast_all["year"] = _periods.dt.year
    forecast_all["monthname"] = _periods.dt.strftime("%b")

    _fc_dates = pd.to_datetime(forecast_all["date"])
    _iso = _fc_dates.dt.isocalendar()
    forecast_all["iso_week"] = [f"{y}-{int(w):02d}" for y, w in zip(_iso.year, _iso.week)]

    # ── Write forecast.csv to S3 and SharePoint ───────────────────────
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    token = get_access_token()
    csv_buf = BytesIO()
    forecast_all.to_csv(csv_buf, index=False)
    csv_buf.seek(0)
    S3.put_object(Bucket=BUCKET, Key=FORECAST_KEY, Body=csv_buf.getvalue(), ContentType="text/csv")

    try:
        result_fc = upload_to_sharepoint(token, csv_buf.getvalue(), f"forecast_{timestamp}.csv")
        print(f"[sharepoint] forecast upload: {result_fc.get('name', result_fc)}")
    except Exception as e:
        print(f"[sharepoint] forecast upload FAILED: {e}")

    try:
        # Build separate_shops: mirrors forecasted_shop classification in sep_act
        _sep_fc_shops_set = set(
            sep_fc["destination_region"] + "||" +
            sep_fc["forecast_product"] + "||" +
            sep_fc["forecasted_shop"])

        act_upload = act_full[act_full["fulldate"] >= "2026-01-01"].copy()
        act_upload["_region"]   = act_upload["region"].astype(str).str.strip()
        act_upload["_product"]  = act_upload["forecast_product"].astype(str).str.strip()
        act_upload["_shop"]     = act_upload["forecasted_shop"].astype(str).str.strip()
        act_upload["_shoptype"] = act_upload["shoptype"].astype(str).str.strip()

        _shop_key   = act_upload["_region"] + "||" + act_upload["_product"] + "||" + act_upload["_shop"]
        _other_name = "Other " + act_upload["_shoptype"]
        _other_key  = act_upload["_region"] + "||" + act_upload["_product"] + "||" + _other_name

        _has_direct = _shop_key.isin(_sep_fc_shops_set)
        _has_other  = _other_key.isin(_sep_fc_shops_set)

        act_upload["separate_shops"] = np.select(
            [_has_direct, ~_has_direct & _has_other],
            [act_upload["_shop"], _other_name],
            default=act_upload["_shoptype"])

        act_upload.drop(columns=["_region", "_product", "_shop", "_shoptype"], inplace=True)
        del _sep_fc_shops_set, _shop_key, _other_name, _other_key, _has_direct, _has_other

        act_upload_bytes = act_upload.to_csv(index=False).encode("utf-8")
        del act_upload
        result_act = upload_to_sharepoint(token, act_upload_bytes, "actuals.csv")
        print(f"[sharepoint] actuals upload: {result_act.get('name', result_act)}")
    except Exception as e:
        print(f"[sharepoint] actuals upload FAILED: {e}")
        
    # ── Write remaining_forecasts.xlsx ────────────────────────────────
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if not calc_eu.empty:
            calc_eu.to_excel(writer, sheet_name="EU + RoW", index=False)
        if not calc_us.empty:
            calc_us.to_excel(writer, sheet_name="US + CA", index=False)
        actuals_combos.to_excel(writer, sheet_name="Coverage", index=False)
        set_growth.to_excel(writer, sheet_name="Set Growth", index=False)
        if not format_shares.empty:
            format_shares.to_excel(writer, sheet_name="Format Shares", index=False)
        if not pcs_shares.empty:
            pcs_shares.to_excel(writer, sheet_name="PCS Shares", index=False)
        if not dow_shares.empty:
            dow_shares.to_excel(writer, sheet_name="DOW Shares", index=False)
        if not shoptype_shares.empty:
            shoptype_shares.to_excel(writer, sheet_name="Shoptype Shares", index=False)

        if not calc_eu.empty:
            format_summary_sheet(writer.book["EU + RoW"], "product", months_eu)
        if not calc_us.empty:
            format_summary_sheet(writer.book["US + CA"], "product", months_us)

        for sg_sheet in ["Set Growth"]:
            sg_ws = writer.book[sg_sheet]
            for c in range(1, sg_ws.max_column + 1):
                h = sg_ws.cell(row=1, column=c).value
                if h and str(h).startswith("YoY_"):
                    for r in range(2, sg_ws.max_row + 1):
                        sg_ws.cell(row=r, column=c).number_format = PCT_FMT

        for sheet_name, share_col in [
            ("Format Shares", "share"), ("PCS Shares", "pcs_share"),
            ("DOW Shares", "dow_share"), ("Shoptype Shares", "share"),
        ]:
            if sheet_name not in writer.book.sheetnames:
                continue
            ws = writer.book[sheet_name]
            for c in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=c).value
                if h == share_col:
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=c).number_format = PCT_FMT
                ws.cell(row=1, column=c).font = BOLD
                ws.cell(row=1, column=c).border = Border(bottom=THIN)
                header_len = len(str(h)) if h else 0
                ws.column_dimensions[get_column_letter(c)].width = max(10, min(40, header_len + 4))

    S3.put_object(
        Bucket=BUCKET, Key=OUTPUT_KEY, Body=buf.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    total_combos = len(actuals_combos)
    not_fc_count = int((actuals_combos["forecasted"] == "No").sum())
    slack_reply(
        event.get("response_url", ""),
        "...completed - please refresh Demand Forecast 2026.xlsx",
    )
    return {
        "ok": True,
        "s3_key": OUTPUT_KEY,
        "forecast_key": FORECAST_KEY,
        "total": total_combos,
        "forecasted": total_combos - not_fc_count,
        "not_forecasted": not_fc_count,
        "format_share_rows": len(format_shares),
        "pcs_share_rows": len(pcs_shares),
        "dow_share_rows": len(dow_shares),
        "shoptype_share_rows": len(shoptype_shares),
        "daily_forecast_rows": len(forecast_all),
        "remaining_rows": len(forecast_all[forecast_all["source"] == "remaining"]),
        "separate_rows": len(forecast_all[forecast_all["source"] == "separate"]),
    }


# ── Formatting ────────────────────────────────────────────────────────
THIN = Side(style="thin", color="000000")
BOLD = Font(bold=True)
WHITE = PatternFill(fill_type="solid", fgColor="FFFFFF")
PCT_FMT = "0.0%"


def format_summary_sheet(ws, label_col_name, month_cols):
    max_row, max_col = ws.max_row, ws.max_column
    headers = {}

    total_row = None
    for r in range(2, max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.startswith("TOTAL"):
            total_row = r
            break
    end_row = total_row or max_row

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = WHITE
            cell.border = Border()

    for c in range(1, max_col + 1):
        h = ws.cell(row=1, column=c)
        headers[h.value] = c
        h.font = BOLD
        h.alignment = Alignment(horizontal="left")
        h.border = Border(bottom=THIN)

    pct_names = {"% of Total YTD", "YTD % growth", "Last 3w % growth"} | set(month_cols)
    for name, c in headers.items():
        if name and str(name).strip() in pct_names:
            for r in range(2, end_row + 1):
                ws.cell(row=r, column=c).number_format = PCT_FMT

    for name in ("YTD % growth", "Last 3w % growth"):
        c = headers.get(name)
        if c:
            for r in range(1, end_row + 1):
                cell = ws.cell(row=r, column=c)
                b = cell.border or Border()
                cell.border = Border(left=b.left, right=THIN, top=b.top, bottom=b.bottom)

    label_c = headers.get(label_col_name)
    if label_c:
        for r in range(2, end_row + 1):
            v = ws.cell(row=r, column=label_c).value
            if isinstance(v, str) and v.startswith("TOTAL"):
                ws.cell(row=r, column=label_c).alignment = Alignment(horizontal="left")
            else:
                ws.cell(row=r, column=label_c).alignment = Alignment(horizontal="left", indent=1)

    if total_row:
        for c in range(1, max_col + 1):
            cell = ws.cell(row=total_row, column=c)
            cell.font = BOLD
            b = cell.border or Border()
            cell.border = Border(left=b.left, right=b.right, top=THIN, bottom=b.bottom)

    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        header_len = len(str(v)) if v is not None else 0
        ws.column_dimensions[get_column_letter(c)].width = max(8, min(60, header_len + 2))