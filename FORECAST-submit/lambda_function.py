import re
import csv
import json
from io import BytesIO, StringIO
from collections import defaultdict
from datetime import datetime, timezone
from dateutil.relativedelta import relativedelta

import boto3
import numpy as np
import pandas as pd
import urllib.request
from openpyxl import load_workbook

S3 = boto3.client("s3")

BUCKET = "bi-automations"
PREFIX = "Forecast/Seperate Forecasts/"
ACTUALS_KEY = "Forecast/actuals.csv"
INPUT_XLSX_KEY = "Forecast/Forecast_Input.xlsx"
OUT_KEY_COMBINED = f"{PREFIX}seperate_forecasts_combined.csv"

BLUE_RGB = "DCE6F1"
EPOCH = datetime(1970, 1, 1, tzinfo=timezone.utc)

SHOP_FILE_RE = re.compile(r"^submit_shop_(.+?)(?:_\d{8}_\d{6})?(?:\s*\(\d+\))*\.xlsx$", re.I)
SHOPTYPE_FILE_RE = re.compile(r"^submit_shoptype_(.+?)(?:_\d{8}_\d{6})?(?:\s*\(\d+\))*\.xlsx$", re.I)
REGION_FILE_RE = re.compile(r"^submit_region_(.+?)(?:_\d{8}_\d{6})?(?:\s*\(\d+\))*\.xlsx$", re.I)
ANY_SUBMIT_RE = re.compile(r"^submit_(?:shop|shoptype|region)_.+?(?:_\d{8}_\d{6})?(?:\s*\(\d+\))*\.xlsx$", re.I)
PERIOD_RE = re.compile(r"^\d{4}-\d{2}$")

CSV_FIELDS = [
    "forecasted_shop", "shoptype", "destination", "period",
    "product", "format", "frame", "forecast", "forecast_type",
    "source_file", "source_timestamp",
]


# ═══════════════════════════════════════════════════════════════════
#  SLACK HELPERS
# ═══════════════════════════════════════════════════════════════════


def get_slack_token():
    ssm = boto3.client("ssm")
    return ssm.get_parameter(
        Name="/forecast/slack-bot-token", WithDecryption=True
    )["Parameter"]["Value"]


def slack_reply(response_url, text):
    if not response_url:
        return
    req = urllib.request.Request(
        response_url,
        data=json.dumps({"response_type": "in_channel", "text": text}).encode(),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    urllib.request.urlopen(req, timeout=5).read()


def download_slack_file_to_s3(user_id, channel_id, response_url):
    token = get_slack_token()

    url = f"https://slack.com/api/conversations.history?channel={channel_id}&limit=10"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    resp = json.loads(urllib.request.urlopen(req).read())

    if not resp.get("ok"):
        slack_reply(response_url, f"Slack API error: {resp.get('error', 'unknown')}")
        return None, None

    file_info = None
    for msg in resp.get("messages", []):
        if msg.get("user") != user_id:
            continue
        for f in msg.get("files", []):
            if ANY_SUBMIT_RE.match(f.get("name", "")):
                file_info = f
                break
        if file_info:
            break

    if not file_info:
        slack_reply(response_url, "No matching file found. Upload a `submit_shop_*.xlsx`, `submit_shoptype_*.xlsx`, or `submit_region_*.xlsx` first.")
        return None, None

    file_name = file_info["name"]
    entity_name = (
        shop_from_key(file_name)
        or shoptype_from_key(file_name)
        or region_from_key(file_name)
        or file_name
    )

    req = urllib.request.Request(
        file_info["url_private"],
        headers={"Authorization": f"Bearer {token}"},
    )
    file_bytes = urllib.request.urlopen(req).read()

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    base, ext = file_name.rsplit(".", 1)
    # Strip Slack duplicate counter like " (14)" before adding timestamp
    base = re.sub(r"\s*\(\d+\)$", "", base)
    timestamped_name = f"{base}_{ts}.{ext}"
    s3_key = f"{PREFIX}{timestamped_name}"
    S3.put_object(Bucket=BUCKET, Key=s3_key, Body=file_bytes)
    slack_reply(response_url, f"Submitted `{entity_name}` — adding to seperate forecasts...")
    return s3_key, entity_name


# ═══════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════


def _safe_lm(lm):
    if lm is None:
        return EPOCH
    if lm.tzinfo is None:
        return lm.replace(tzinfo=timezone.utc)
    return lm


def cell_to_period(cv):
    if isinstance(cv, str):
        s = cv.strip()
        if PERIOD_RE.match(s):
            return s
    elif hasattr(cv, "strftime"):
        return cv.strftime("%Y-%m")
    return None


def is_na(v):
    return v is None or (isinstance(v, str) and v.strip().lower() == "na")


def is_blue(cell):
    fill = getattr(cell, "fill", None)
    if not fill or not getattr(fill, "patternType", None):
        return False
    rgb = getattr(getattr(fill, "fgColor", None), "rgb", None)
    return bool(rgb) and rgb.upper().endswith(BLUE_RGB)


def shop_from_key(key):
    name = key.rsplit("/", 1)[-1]
    m = SHOP_FILE_RE.match(name)
    return m.group(1).strip() if m else ""


def shoptype_from_key(key):
    name = key.rsplit("/", 1)[-1]
    m = SHOPTYPE_FILE_RE.match(name)
    return m.group(1).strip() if m else ""


def region_from_key(key):
    name = key.rsplit("/", 1)[-1]
    m = REGION_FILE_RE.match(name)
    return m.group(1).strip() if m else ""


# ═══════════════════════════════════════════════════════════════════
#  S3 I/O
# ═══════════════════════════════════════════════════════════════════


def list_submit_keys(bucket, prefix):
    items = []
    token = None
    while True:
        kwargs = {"Bucket": bucket, "Prefix": prefix, "MaxKeys": 1000}
        if token:
            kwargs["ContinuationToken"] = token
        resp = S3.list_objects_v2(**kwargs)
        for obj in resp.get("Contents", []):
            k = obj["Key"]
            if ANY_SUBMIT_RE.match(k.rsplit("/", 1)[-1]):
                items.append((k, obj.get("LastModified")))
        if not resp.get("IsTruncated"):
            return items
        token = resp.get("NextContinuationToken")


def s3_bytes(bucket, key):
    return S3.get_object(Bucket=bucket, Key=key)["Body"].read()


def read_existing_csv(bucket, key):
    return []


def write_csv(bucket, key, rows):
    buf = StringIO()
    w = csv.DictWriter(buf, fieldnames=CSV_FIELDS, lineterminator="\n")
    w.writeheader()
    for r in rows:
        w.writerow({f: r.get(f, "") for f in CSV_FIELDS})
    S3.put_object(Bucket=bucket, Key=key, Body=buf.getvalue().encode("utf-8"), ContentType="text/csv")


# ═══════════════════════════════════════════════════════════════════
#  SHOP → SHOPTYPE LOOKUP
# ═══════════════════════════════════════════════════════════════════


def build_shop_to_shoptype_and_actuals(bucket, actuals_key):
    try:
        raw = S3.get_object(Bucket=bucket, Key=actuals_key)["Body"].read().decode("utf-8", "replace")
    except Exception:
        return {}, pd.DataFrame()

    actuals_df = pd.read_csv(StringIO(raw))

    reader = csv.DictReader(StringIO(raw))
    totals = defaultdict(float)
    for row in reader:
        shop = (row.get("forecasted_shop") or "").strip()
        st = (row.get("shoptype") or "").strip()
        try:
            val = float(row.get("actuals") or 0)
        except (ValueError, TypeError):
            val = 0.0
        if shop and st:
            totals[(shop, st)] += val

    best = {}
    for (shop, st), total in totals.items():
        if shop not in best or total > best[shop][1]:
            best[shop] = (st, total)

    mapping = {shop: st for shop, (st, _) in best.items()}
    return mapping, actuals_df


# ═══════════════════════════════════════════════════════════════════
#  SHOP FILE READING
# ═══════════════════════════════════════════════════════════════════


def find_row(ws, label, start, end):
    for r in range(start, end + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip() == label:
            return r
    return None


def iter_avail_rows(ws, avail_row):
    r = avail_row + 1
    while r <= ws.max_row:
        v1 = ws.cell(r, 1).value
        v2 = ws.cell(r, 2).value
        if (v1 is None or (isinstance(v1, str) and not v1.strip())) and \
           (v2 is None or (isinstance(v2, str) and not str(v2).strip())):
            return
        yield (r, "" if v1 is None else str(v1).strip(), "" if v2 is None else str(v2).strip())
        r += 1


def find_summary_sheet(wb):
    for name in wb.sheetnames:
        ws = wb[name]
        v = ws.cell(1, 1).value
        if isinstance(v, str) and v.strip() == "TOP Products":
            return ws
    return wb[wb.sheetnames[0]]


def read_destination(summary_ws):
    for r in range(1, summary_ws.max_row + 1):
        v = summary_ws.cell(r, 1).value
        if isinstance(v, str) and v.strip() == "Destination:":
            dv = summary_ws.cell(r, 2).value
            s = "" if dv is None else str(dv).strip()
            return s or "Unknown"
    return "Unknown"


_HEADER_SKIP_LABELS = {
    "Actuals", "Actuals LY", "YoY %", "F-YoY %", "Forecast",
    "Calculations", "Metric", "Available formats:",
}


def _find_header_row(ws):
    for r in range(1, min(80, ws.max_row) + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip() == "Product name":
            return r

    for r in range(1, min(80, ws.max_row) + 1):
        if not (cell_to_period(ws.cell(r, 3).value) or cell_to_period(ws.cell(r, 4).value)):
            continue
        v = ws.cell(r, 1).value
        label = str(v).strip() if v else ""
        if label in _HEADER_SKIP_LABELS or label.startswith("Destination:"):
            continue
        return r

    return -1


def read_submit_xlsx(file_bytes, key):
    shop = shop_from_key(key)
    if not shop:
        return []

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    summary_ws = find_summary_sheet(wb)
    destination = read_destination(summary_ws)
    out = []

    for ws in wb.worksheets:
        if ws.title == summary_ws.title:
            continue

        header = _find_header_row(ws)
        if header < 0:
            continue

        month_cols = {}
        for c in range(3, ws.max_column + 1):
            p = cell_to_period(ws.cell(header, c).value)
            if p:
                month_cols[c] = p
        if not month_cols:
            continue

        scan_start = header
        scan_end = min(ws.max_row, header + 250)
        forecast_row = find_row(ws, "Forecast", scan_start, scan_end)
        avail_row = find_row(ws, "Available formats:", scan_start, scan_end)
        product = ws.title

        def emit(period, fmt, frame, value):
            out.append({
                "forecasted_shop": shop,
                "destination": destination,
                "period": period,
                "product": product,
                "format": fmt or "Unknown",
                "frame": frame or "Unknown",
                "forecast": value,
                "forecast_type": "Manual",
            })

        for col, period in month_cols.items():
            if forecast_row:
                cell = ws.cell(forecast_row, col)
                if is_blue(cell) and not is_na(cell.value):
                    emit(period, "Unknown", "Unknown", cell.value)
            if avail_row:
                for rr, fmt, frame in iter_avail_rows(ws, avail_row):
                    cell = ws.cell(rr, col)
                    if is_blue(cell) and not is_na(cell.value):
                        emit(period, fmt, frame, cell.value)

    return out


# ═══════════════════════════════════════════════════════════════════
#  SHOPTYPE FILE READING
# ═══════════════════════════════════════════════════════════════════


def read_submit_shoptype_xlsx(file_bytes, key):
    shoptype = shoptype_from_key(key)
    if not shoptype:
        return []

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    out = []

    for ws in wb.worksheets:
        v11 = ws.cell(1, 1).value
        if isinstance(v11, str) and v11.strip() == "TOP Products":
            continue
        if ws.title.lower().startswith("shoptype "):
            continue

        product = ws.title

        dest_rows = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.strip().startswith("Destination:"):
                dest_rows.append(r)

        for idx, dest_r in enumerate(dest_rows):
            block_end = dest_rows[idx + 1] - 1 if idx + 1 < len(dest_rows) else ws.max_row

            dest_raw = str(ws.cell(dest_r, 1).value).strip()
            dest_value = dest_raw[len("Destination:"):].strip() if ":" in dest_raw else "Unknown"

            month_cols = {}
            for c in range(3, ws.max_column + 1):
                p = cell_to_period(ws.cell(dest_r, c).value)
                if p:
                    month_cols[c] = p
            if not month_cols:
                continue

            forecast_row = None
            shop_rows = []

            for r in range(dest_r + 1, block_end + 1):
                v = ws.cell(r, 1).value
                label = str(v).strip() if v is not None else ""

                if label == "Forecast":
                    forecast_row = r
                    continue

                if forecast_row is not None and label and label not in (
                    "Actuals", "Actuals LY", "YoY %", "F-YoY %",
                    "Calculations", "Last 6w % growth:", "Metric", "Product name",
                ) and not label.startswith("Destination:"):
                    if any(is_blue(ws.cell(r, c)) for c in month_cols):
                        shop_rows.append((r, label))

            if forecast_row:
                for col, period in month_cols.items():
                    cell = ws.cell(forecast_row, col)
                    if is_blue(cell) and not is_na(cell.value):
                        out.append({
                            "forecasted_shop": shoptype,
                            "shoptype": shoptype,
                            "destination": dest_value or "Unknown",
                            "period": period,
                            "product": product,
                            "format": "Unknown",
                            "frame": "Unknown",
                            "forecast": cell.value,
                            "forecast_type": "Manual",
                            "_is_shoptype_total": True,
                        })

            for shop_r, shop_name in shop_rows:
                for col, period in month_cols.items():
                    cell = ws.cell(shop_r, col)
                    if is_blue(cell) and not is_na(cell.value):
                        out.append({
                            "forecasted_shop": shop_name,
                            "shoptype": shoptype,
                            "destination": dest_value or "Unknown",
                            "period": period,
                            "product": product,
                            "format": "Unknown",
                            "frame": "Unknown",
                            "forecast": cell.value,
                            "forecast_type": "Manual",
                            "_is_shoptype_total": False,
                        })

    return out


# ═══════════════════════════════════════════════════════════════════
#  REGION FILE READING
# ═══════════════════════════════════════════════════════════════════


def read_submit_region_xlsx(file_bytes, key):
    region = region_from_key(key)
    if not region:
        return []

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    out = []

    for ws in wb.worksheets:
        v11 = ws.cell(1, 1).value
        if isinstance(v11, str) and v11.strip() == "TOP Products":
            continue

        product = ws.title

        month_header_row = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if isinstance(v, str) and v.strip() == "Month:":
                month_header_row = r
                break
        if month_header_row is None:
            continue

        label_row = month_header_row - 1
        data_start = month_header_row + 1

        month_rows = []
        r = data_start
        while r <= ws.max_row:
            v = ws.cell(r, 2).value
            if v is None:
                break
            s = str(v).strip()
            if s == "TOTAL":
                break
            p = cell_to_period(v)
            if p:
                month_rows.append((r, p))
            r += 1

        if not month_rows:
            continue

        blocks = []
        c = 3
        while c <= ws.max_column:
            v = ws.cell(label_row, c).value
            if v is not None and str(v).strip():
                blocks.append((str(v).strip(), c))
                c += 3
            else:
                c += 1

        if not blocks:
            continue

        current_shoptype = ""
        sheet_rows = []
        shoptypes_with_shops = set()

        for block_label, col_start in blocks:
            if block_label == "TOTAL":
                continue

            if block_label.startswith("TOTAL "):
                current_shoptype = block_label[len("TOTAL "):]
                for row, period in month_rows:
                    cell = ws.cell(row, col_start)
                    if is_blue(cell) and not is_na(cell.value):
                        sheet_rows.append(({
                            "forecasted_shop": current_shoptype,
                            "shoptype": current_shoptype,
                            "destination": region,
                            "period": period,
                            "product": product,
                            "format": "Unknown",
                            "frame": "Unknown",
                            "forecast": cell.value,
                            "forecast_type": "Manual",
                        }, current_shoptype, True))
                continue

            if block_label.startswith("Other "):
                shoptype = block_label[len("Other "):]
                forecasted_shop = block_label
            else:
                shoptype = current_shoptype
                forecasted_shop = block_label

            for row, period in month_rows:
                cell = ws.cell(row, col_start)
                if is_blue(cell) and not is_na(cell.value):
                    sheet_rows.append(({
                        "forecasted_shop": forecasted_shop,
                        "shoptype": shoptype,
                        "destination": region,
                        "period": period,
                        "product": product,
                        "format": "Unknown",
                        "frame": "Unknown",
                        "forecast": cell.value,
                        "forecast_type": "Manual",
                    }, shoptype, False))
                    shoptypes_with_shops.add(shoptype)

        for row_dict, shoptype, is_total in sheet_rows:
            if is_total and shoptype in shoptypes_with_shops:
                continue
            out.append(row_dict)

    return out


# ═══════════════════════════════════════════════════════════════════
#  PROCESSING PIPELINES
# ═══════════════════════════════════════════════════════════════════


def process_shop_files(submit_items, existing_csv_key):
    submit_items.sort(key=lambda x: (x[1] is None, x[1], x[0]))
    existing = read_existing_csv(BUCKET, existing_csv_key)

    def full_key(r):
        return (
            (r.get("forecasted_shop") or "").strip(),
            (r.get("destination") or "").strip(),
            (r.get("period") or "").strip(),
            (r.get("product") or "").strip(),
            (r.get("format") or "").strip(),
            (r.get("frame") or "").strip(),
        )

    def group_key(r):
        return (
            (r.get("forecasted_shop") or "").strip(),
            (r.get("destination") or "").strip(),
            (r.get("period") or "").strip(),
            (r.get("product") or "").strip(),
        )

    dedup = {}
    file_rank_by_fk = {}
    file_lm_by_fk = {}
    for r in existing:
        r["_file_lm"] = EPOCH
        r["_source_file"] = "existing_csv"
        k = full_key(r)
        dedup[k] = r
        file_rank_by_fk[k] = 0
        file_lm_by_fk[k] = EPOCH

    new_rows = []
    file_rank = 0
    for key, lm in submit_items:
        file_rank += 1
        safe = _safe_lm(lm)
        rows = read_submit_xlsx(s3_bytes(BUCKET, key), key)
        for r in rows:
            r["_file_lm"] = safe
            r["_source_file"] = key
        new_rows.extend(rows)
        for r in rows:
            k = full_key(r)
            dedup[k] = r
            file_rank_by_fk[k] = file_rank
            file_lm_by_fk[k] = safe

    combined = list(dedup.values())

    latest_rank = {}
    for r in combined:
        gk = group_key(r)
        fr = file_rank_by_fk.get(full_key(r), 0)
        if (gk not in latest_rank) or (fr > latest_rank[gk]):
            latest_rank[gk] = fr

    latest_rows = [
        r for r in combined
        if file_rank_by_fk.get(full_key(r), 0) == latest_rank.get(group_key(r), 0)
    ]

    has_specific = {}
    for r in latest_rows:
        gk = group_key(r)
        fmt = (r.get("format") or "").strip()
        frm = (r.get("frame") or "").strip()
        if (fmt and fmt != "Unknown") or (frm and frm != "Unknown"):
            has_specific[gk] = True

    final = []
    for r in latest_rows:
        gk = group_key(r)
        fmt = (r.get("format") or "").strip()
        frm = (r.get("frame") or "").strip()
        if has_specific.get(gk) and fmt == "Unknown" and frm == "Unknown":
            continue
        final.append(r)

    return final, len(new_rows), len(existing)


def process_shoptype_files(submit_items, existing_csv_key):
    submit_items.sort(key=lambda x: (x[1] is None, x[1], x[0]))
    existing = read_existing_csv(BUCKET, existing_csv_key)

    def full_key(r):
        return (
            (r.get("forecasted_shop") or "").strip(),
            (r.get("destination") or "").strip(),
            (r.get("period") or "").strip(),
            (r.get("product") or "").strip(),
            (r.get("format") or "").strip(),
            (r.get("frame") or "").strip(),
        )

    def snapshot_key(r):
        return (
            (r.get("forecasted_shop") or "").strip(),
            (r.get("destination") or "").strip(),
            (r.get("period") or "").strip(),
            (r.get("product") or "").strip(),
        )

    dedup = {}
    file_rank_by_fk = {}
    file_lm_by_fk = {}
    for r in existing:
        r["_file_lm"] = EPOCH
        r["_source_file"] = "existing_csv"
        k = full_key(r)
        dedup[k] = r
        file_rank_by_fk[k] = 0
        file_lm_by_fk[k] = EPOCH

    new_rows = []
    file_rank = 0
    for key, lm in submit_items:
        file_rank += 1
        safe = _safe_lm(lm)
        rows = read_submit_shoptype_xlsx(s3_bytes(BUCKET, key), key)
        for r in rows:
            r["_file_lm"] = safe
            r["_source_file"] = key
        new_rows.extend(rows)
        for r in rows:
            k = full_key(r)
            dedup[k] = r
            file_rank_by_fk[k] = file_rank
            file_lm_by_fk[k] = safe

    combined = list(dedup.values())

    latest_rank = {}
    for r in combined:
        sk = snapshot_key(r)
        fr = file_rank_by_fk.get(full_key(r), 0)
        if (sk not in latest_rank) or (fr > latest_rank[sk]):
            latest_rank[sk] = fr

    latest_rows = [
        r for r in combined
        if file_rank_by_fk.get(full_key(r), 0) == latest_rank.get(snapshot_key(r), 0)
    ]

    has_shop_rows = {}
    for r in latest_rows:
        if not r.get("_is_shoptype_total", False):
            st = (r.get("shoptype") or "").strip()
            k = (st, (r.get("destination") or "").strip(),
                     (r.get("period") or "").strip(),
                     (r.get("product") or "").strip())
            has_shop_rows[k] = True

    def total_suppression_key(r):
        st = (r.get("forecasted_shop") or "").strip()
        return (st, (r.get("destination") or "").strip(),
                    (r.get("period") or "").strip(),
                    (r.get("product") or "").strip())

    final = [
        r for r in latest_rows
        if not (r.get("_is_shoptype_total", False) and has_shop_rows.get(total_suppression_key(r)))
    ]

    for r in final:
        r.pop("_is_shoptype_total", None)

    return final, len(new_rows), len(existing)


def process_region_files(submit_items, existing_csv_key):
    submit_items.sort(key=lambda x: (x[1] is None, x[1], x[0]))
    existing = read_existing_csv(BUCKET, existing_csv_key)

    def full_key(r):
        return (
            (r.get("forecasted_shop") or "").strip(),
            (r.get("destination") or "").strip(),
            (r.get("period") or "").strip(),
            (r.get("product") or "").strip(),
        )

    dedup = {}
    file_rank_by_fk = {}
    for r in existing:
        r["_file_lm"] = EPOCH
        r["_source_file"] = "existing_csv"
        k = full_key(r)
        dedup[k] = r
        file_rank_by_fk[k] = 0

    new_rows = []
    file_rank = 0
    for key, lm in submit_items:
        file_rank += 1
        safe = _safe_lm(lm)
        rows = read_submit_region_xlsx(s3_bytes(BUCKET, key), key)
        for r in rows:
            r["_file_lm"] = safe
            r["_source_file"] = key
        new_rows.extend(rows)
        for r in rows:
            k = full_key(r)
            dedup[k] = r
            file_rank_by_fk[k] = file_rank

    final = list(dedup.values())
    return final, len(new_rows), len(existing)


# ═══════════════════════════════════════════════════════════════════
#  BUILD COMBINED — region > shoptype > shop
# ═══════════════════════════════════════════════════════════════════


def build_combined(shop_rows, st_rows, region_rows, current_month):
    """
    Period-aware source priority:
    - For every (shop, product, dest, period): newest file wins.
    - If the winning file has no row for a given period, fall back to any
      other source that does — preserving explicitly submitted data regardless
      of how near or far the period is.
    """

    def _dest(r):
        return (r.get("destination") or "").strip()

    def _period(r):
        return (r.get("period") or "").strip()

    # ── File-level timestamps (for winner determination) ─────────────────
    region_lm = {}
    for r in region_rows:
        shop = (r.get("forecasted_shop") or "").strip()
        product = (r.get("product") or "").strip()
        dest = _dest(r)
        lm = r.get("_file_lm", EPOCH)
        key = (shop, product, dest)
        if shop and product:
            if key not in region_lm or lm > region_lm[key]:
                region_lm[key] = lm

    shop_lm = {}
    for r in shop_rows:
        shop = (r.get("forecasted_shop") or "").strip()
        product = (r.get("product") or "").strip()
        dest = _dest(r)
        lm = r.get("_file_lm", EPOCH)
        key = (shop, product, dest)
        if key not in shop_lm or lm > shop_lm[key]:
            shop_lm[key] = lm

    st_shop_lm = {}
    for r in st_rows:
        shop = (r.get("forecasted_shop") or "").strip()
        st = (r.get("shoptype") or "").strip()
        product = (r.get("product") or "").strip()
        dest = _dest(r)
        lm = r.get("_file_lm", EPOCH)
        if shop and product and shop != st:
            key = (shop, product, dest)
            if key not in st_shop_lm or lm > st_shop_lm[key]:
                st_shop_lm[key] = lm

    region_st_coverage = {}
    for r in region_rows:
        st = (r.get("shoptype") or "").strip()
        product = (r.get("product") or "").strip()
        dest = _dest(r)
        lm = r.get("_file_lm", EPOCH)
        if st and product:
            k = (st, product, dest)
            if k not in region_st_coverage or lm > region_st_coverage[k]:
                region_st_coverage[k] = lm

    st_total_lm = {}
    for r in st_rows:
        shop = (r.get("forecasted_shop") or "").strip()
        st = (r.get("shoptype") or "").strip()
        product = (r.get("product") or "").strip()
        dest = _dest(r)
        lm = r.get("_file_lm", EPOCH)
        if shop == st and shop and product:
            k = (shop, product, dest)
            if k not in st_total_lm or lm > st_total_lm[k]:
                st_total_lm[k] = lm

    # ── Period-level coverage: which (shop, product, dest, period) each source has ──
    region_periods = {
        ((r.get("forecasted_shop") or "").strip(), (r.get("product") or "").strip(), _dest(r), _period(r))
        for r in region_rows
    }
    shop_periods = {
        ((r.get("forecasted_shop") or "").strip(), (r.get("product") or "").strip(), _dest(r), _period(r))
        for r in shop_rows
    }
    st_periods = {
        ((r.get("forecasted_shop") or "").strip(), (r.get("product") or "").strip(), _dest(r), _period(r))
        for r in st_rows
    }

    # ── File-level winners ───────────────────────────────────────────────
    # FIX (Issue 3): include st_total_lm in the comparison so that a NEWER
    # shoptype-specific submit (where shop == shoptype, e.g., sendmoments
    # submitting its own total) can override an older region-file row.
    # Previously st_total_lm was tracked separately but never consulted in
    # winner determination, which meant freshly uploaded shoptype totals
    # were silently ignored if a region file had any timestamp at all.
    all_keys = set(region_lm) | set(shop_lm) | set(st_shop_lm) | set(st_total_lm)
    region_wins = set()
    shop_wins = set()
    st_wins = set()

    for key in all_keys:
        r_lm = region_lm.get(key, EPOCH)
        s_lm = shop_lm.get(key, EPOCH)
        t_lm = st_shop_lm.get(key, EPOCH)
        tt_lm = st_total_lm.get(key, EPOCH)
        best = max(r_lm, s_lm, t_lm, tt_lm)
        if tt_lm == best and key in st_total_lm:
            # Shoptype-as-total wins (e.g., sendmoments submitting its own total).
            # Mark as st_wins so the shoptype-rows loop appends it.
            st_wins.add(key)
        elif r_lm == best and key in region_lm:
            region_wins.add(key)
        elif t_lm == best and key in st_shop_lm:
            st_wins.add(key)
        elif key in shop_lm:
            shop_wins.add(key)

    def _winner_covers(shop, product, dest, period):
        """Does the winning source actually have a row for this period?"""
        pk = (shop, product, dest, period)
        if (shop, product, dest) in region_wins:
            return pk in region_periods
        if (shop, product, dest) in st_wins:
            return pk in st_periods
        if (shop, product, dest) in shop_wins:
            return pk in shop_periods
        return False

    def _keep_as_fallback(shop, product, dest, period):
        """
        Allow a non-winning source to contribute if the winning source
        has no row for this specific period. Real submitted data should
        never be silently dropped regardless of how near or far the period is.
        """
        return not _winner_covers(shop, product, dest, period)

    shoptype_member_shops = defaultdict(set)
    for r in st_rows:
        s = (r.get("forecasted_shop") or "").strip()
        st_name = (r.get("shoptype") or "").strip()
        p = (r.get("product") or "").strip()
        d = _dest(r)
        if s != st_name and s and p:
            shoptype_member_shops[(st_name, p, d)].add(s)

    shoptype_covered_by_shop_files = {
        key
        for key, members in shoptype_member_shops.items()
        if members and all((s, key[1], key[2]) in shop_wins for s in members)
    }

    if shoptype_covered_by_shop_files:
        print(f"[build_combined] suppressing shoptype totals for {len(shoptype_covered_by_shop_files)} "
              f"(shoptype, product, dest) combos — all member shops have shop files")

    if region_wins:
        print(f"[build_combined] region file newest for {len(region_wins)} (shop,product,dest) tuples")
    if st_wins:
        print(f"[build_combined] shoptype file newest for {len(st_wins)} (shop,product,dest) tuples")
    if shop_wins:
        print(f"[build_combined] shop file newest for {len(shop_wins)} (shop,product,dest) tuples")

    combined = []
    fallback_rows = 0

    # ── Region rows ──────────────────────────────────────────────────────
    for r in region_rows:
        shop = (r.get("forecasted_shop") or "").strip()
        product = (r.get("product") or "").strip()
        dest = _dest(r)
        period = _period(r)
        key = (shop, product, dest)

        if key in region_wins:
            if shop.startswith("Other "):
                st_name = shop[len("Other "):]
                st_tot_lm_val = st_total_lm.get((st_name, product, dest), EPOCH)
                reg_lm_val = region_lm.get(key, EPOCH)
                if st_tot_lm_val > reg_lm_val:
                    continue
            st = (r.get("shoptype") or "").strip()
            if shop == st and shop:
                # Stale shoptype total from old region file.
                # Suppress if a newer shoptype file has broken this into
                # individual shops + "Other [shoptype]".
                other_key = (f"Other {st}", product, dest)
                if other_key in st_shop_lm or (st, product, dest) in shoptype_member_shops:
                    continue
            combined.append(r)
        elif _keep_as_fallback(shop, product, dest, period):
            combined.append(r)
            fallback_rows += 1

    # ── Shoptype rows ────────────────────────────────────────────────────
    for r in st_rows:
        shop = (r.get("forecasted_shop") or "").strip()
        st = (r.get("shoptype") or "").strip()
        product = (r.get("product") or "").strip()
        dest = _dest(r)
        period = _period(r)
        key = (shop, product, dest)

        if shop == st:
            region_cov_lm = region_st_coverage.get((st, product, dest))
            if region_cov_lm is not None:
                st_tot_lm_val = st_total_lm.get((st, product, dest), EPOCH)
                if region_cov_lm >= st_tot_lm_val:
                    continue
            if (st, product, dest) in shoptype_covered_by_shop_files:
                continue
            if key not in region_wins and key not in shop_wins:
                combined.append(r)
            elif _keep_as_fallback(shop, product, dest, period):
                combined.append(r)
                fallback_rows += 1
        else:
            if key in st_wins:
                combined.append(r)
            elif key not in region_wins and key not in shop_wins:
                combined.append(r)
            elif _keep_as_fallback(shop, product, dest, period):
                combined.append(r)
                fallback_rows += 1

    # ── Shop rows ────────────────────────────────────────────────────────
    for r in shop_rows:
        shop = (r.get("forecasted_shop") or "").strip()
        product = (r.get("product") or "").strip()
        dest = _dest(r)
        period = _period(r)
        key = (shop, product, dest)

        if key in shop_wins:
            combined.append(r)
        elif key not in region_wins and key not in st_wins:
            combined.append(r)
        elif _keep_as_fallback(shop, product, dest, period):
            combined.append(r)
            fallback_rows += 1

    print(f"[build_combined] period-fallback rows retained: {fallback_rows}")

    # Deduplicate: if multiple sources contributed a row for the same
    # (shop, product, dest, period, format, frame), keep only the newest.
    dedup = {}
    for r in combined:
        dk = (
            (r.get("forecasted_shop") or "").strip(),
            (r.get("shoptype") or "").strip(),
            (r.get("destination") or "").strip(),
            (r.get("period") or "").strip(),
            (r.get("product") or "").strip(),
            (r.get("format") or "").strip(),
            (r.get("frame") or "").strip(),
        )
        existing_lm = dedup[dk].get("_file_lm", EPOCH) if dk in dedup else EPOCH
        if dk not in dedup or r.get("_file_lm", EPOCH) >= existing_lm:
            dedup[dk] = r
    combined = list(dedup.values())
    print(f"[build_combined] after dedup: {len(combined)} rows")
    return combined


# ═══════════════════════════════════════════════════════════════════
#  FORMAT DISTRIBUTION
# ═══════════════════════════════════════════════════════════════════


def get_last_q4_months(current_ym):
    y = datetime.strptime(current_ym, "%Y-%m").year - 1
    return [f"{y}-10", f"{y}-11", f"{y}-12"]


def get_last_n_months(current_ym, n=6):
    dt = datetime.strptime(current_ym, "%Y-%m")
    return [(dt - relativedelta(months=i)).strftime("%Y-%m") for i in reversed(range(n))]


def is_q4_period(period_str):
    try:
        return int(period_str.split("-")[1]) >= 10
    except Exception:
        return False


def read_forecast_config(bucket, xlsx_key):
    raw = s3_bytes(bucket, xlsx_key)
    xls = pd.ExcelFile(BytesIO(raw))

    info = pd.read_excel(xls, sheet_name="info", header=None)
    current_month = str(info.iat[3, 2]).strip()

    ignore_fmt = pd.read_excel(xls, sheet_name="adhoc", header=9, usecols="E").dropna()
    ignore_fmt.columns = ["ignore_format"]

    nf_eu = pd.read_excel(xls, sheet_name="EU formats")
    nf_us = pd.read_excel(xls, sheet_name="US formats")
    new_formats = pd.concat([nf_eu, nf_us], ignore_index=True)

    return current_month, ignore_fmt, new_formats


def make_combo_key(df, cols, new_col="combo_key"):
    df = df.copy()
    df[new_col] = df[cols[0]].astype(str).str.cat(
        [df[c].astype(str) for c in cols[1:]], sep="_"
    )
    return df


def compute_shares(
    df, *, months=None, month_col=None,
    groupby_dims, total_by_dims,
    excluded=None, exclude_col=None,
):
    data = df.copy()
    if months is not None:
        data = data[data[month_col].isin(months)]
    if excluded is not None:
        data = data[~data[exclude_col].isin(excluded)]

    g = data.groupby(groupby_dims, as_index=False, dropna=False).agg({"actuals": "sum"})
    totals = g.groupby(total_by_dims, dropna=False)["actuals"].transform("sum")
    g["denominator_actuals"] = totals
    g["share"] = np.where(totals > 0, g["actuals"] / totals, 0.0)
    return g


def share_check(df, key_cols):
    ln = df[df["avg_type"] == "last N months"]
    q4 = df[df["avg_type"] == "last Q4"]
    p = (
        ln[key_cols].assign(in_ln=True)
        .merge(q4[key_cols].assign(in_q4=True), on=key_cols, how="outer")
        .fillna({"in_ln": False, "in_q4": False})
    )
    ln_s = ln[key_cols + ["share"]].rename(columns={"share": "share_last_n"})
    q4_s = q4[key_cols + ["share"]].rename(columns={"share": "share_q4"})
    p = p.merge(ln_s, on=key_cols, how="left").merge(q4_s, on=key_cols, how="left")
    return p[~(p["in_ln"] & p["in_q4"])].reset_index(drop=True)


def fill_mismatched_shares(shares_df, key_cols):
    chk = share_check(shares_df, key_cols)
    if chk.empty:
        return shares_df

    miss_q4 = chk["share_q4"].isna() & chk["share_last_n"].notna()
    miss_n = chk["share_last_n"].isna() & chk["share_q4"].notna()

    adds = pd.concat([
        chk.loc[miss_q4, key_cols].assign(
            avg_type="last Q4", share=chk.loc[miss_q4, "share_last_n"].values
        ),
        chk.loc[miss_n, key_cols].assign(
            avg_type="last N months", share=chk.loc[miss_n, "share_q4"].values
        ),
    ], ignore_index=True)

    out = pd.concat([shares_df, adds], ignore_index=True)
    return out.groupby(key_cols + ["avg_type"], as_index=False, dropna=False).agg(
        share=("share", "max")
    )


def inject_expected_shares(shares, new_formats):
    if "expected_share" not in new_formats.columns:
        return shares
    if not pd.to_numeric(new_formats["expected_share"], errors="coerce").notna().any():
        return shares

    nf = new_formats[new_formats["expected_share"] != "x"].copy()
    nf["expected_share"] = pd.to_numeric(nf["expected_share"], errors="coerce")
    nf = nf.dropna(subset=["expected_share"])
    if nf.empty:
        return shares

    nf = pd.concat([
        nf.assign(avg_type="last N months"),
        nf.assign(avg_type="last Q4"),
    ], ignore_index=True).drop_duplicates(
        ["destination", "product", "format_clean", "frame", "avg_type"]
    )

    shares = make_combo_key(
        shares,
        cols=["destination_region", "forecast_product", "format_clean", "frame", "avg_type"],
        new_col="_mk_nf",
    )
    nf = make_combo_key(
        nf,
        cols=["destination", "product", "format_clean", "frame", "avg_type"],
        new_col="_mk_nf",
    )
    shares_keep = shares[~shares["_mk_nf"].isin(nf["_mk_nf"])].copy()

    denom = (
        shares[["destination_region", "forecast_product", "shoptype", "avg_type", "denominator_actuals"]]
        .drop_duplicates()
    )
    denom = make_combo_key(
        denom, cols=["destination_region", "forecast_product", "avg_type"], new_col="_mk_denom",
    )
    nf = make_combo_key(
        nf, cols=["destination", "product", "avg_type"], new_col="_mk_denom"
    )

    nf_expanded = nf.merge(
        denom[["_mk_denom", "shoptype", "denominator_actuals"]],
        on="_mk_denom", how="inner",
    )
    nf_expanded["actuals"] = nf_expanded["expected_share"] * nf_expanded["denominator_actuals"]
    nf_expanded = nf_expanded.rename(columns={
        "destination": "destination_region", "product": "forecast_product",
    })

    shares_agg = shares_keep.groupby(
        ["destination_region", "forecast_product", "format_clean", "frame", "shoptype", "avg_type"],
        as_index=False,
    ).agg({"actuals": "sum"})

    nf_agg = nf_expanded.groupby(
        ["destination_region", "forecast_product", "format_clean", "frame", "shoptype", "avg_type"],
        as_index=False,
    ).agg({"actuals": "sum"})

    pool = pd.concat([shares_agg, nf_agg], ignore_index=True)

    recalced = []
    for avg_type in ["last N months", "last Q4"]:
        sub = pool[pool["avg_type"] == avg_type]
        if sub.empty:
            continue
        r = compute_shares(
            sub,
            groupby_dims=["destination_region", "forecast_product", "format_clean", "frame", "shoptype"],
            total_by_dims=["forecast_product", "destination_region", "shoptype"],
        ).assign(avg_type=avg_type)
        recalced.append(r)

    return pd.concat(recalced, ignore_index=True) if recalced else shares_keep


def compute_shop_level_shares(actuals_df, last_n_months, last_q4, ignore_formats):
    child = ["destination_region", "forecast_product", "format_clean", "frame", "forecasted_shop"]
    parent = ["forecast_product", "destination_region", "forecasted_shop"]

    act = actuals_df[~actuals_df["forecast_product"].isin(ignore_formats["ignore_format"])]

    shares_n = compute_shares(
        act, months=last_n_months, month_col="yyyy-mm",
        groupby_dims=child, total_by_dims=parent,
        excluded=["EXCLUDED FORMAT", "NEW FORMAT"], exclude_col="format_clean",
    ).assign(avg_type="last N months")

    shares_q4 = compute_shares(
        act, months=last_q4, month_col="yyyy-mm",
        groupby_dims=child, total_by_dims=parent,
        excluded=["EXCLUDED FORMAT", "NEW FORMAT"], exclude_col="format_clean",
    ).assign(avg_type="last Q4")

    shares = pd.concat([shares_n, shares_q4], ignore_index=True)

    ign = actuals_df[actuals_df["forecast_product"].isin(ignore_formats["ignore_format"])]
    if not ign.empty:
        ign_agg = ign.groupby(
            ["destination_region", "forecast_product", "forecasted_shop"], as_index=False
        ).agg({"actuals": "sum"})
        for at in ["last N months", "last Q4"]:
            shares = pd.concat([
                shares,
                ign_agg.assign(format_clean="na", frame="na", share=1.0, avg_type=at),
            ], ignore_index=True)

    shares = fill_mismatched_shares(shares, child)
    return shares


def compute_shoptype_level_shares(actuals_df, last_n_months, last_q4, ignore_formats, new_formats):
    child = ["destination_region", "forecast_product", "format_clean", "frame", "shoptype"]
    parent = ["forecast_product", "destination_region", "shoptype"]

    act = actuals_df[~actuals_df["forecast_product"].isin(ignore_formats["ignore_format"])]

    shares_n = compute_shares(
        act, months=last_n_months, month_col="yyyy-mm",
        groupby_dims=child, total_by_dims=parent,
        excluded=["EXCLUDED FORMAT", "NEW FORMAT"], exclude_col="format_clean",
    ).assign(avg_type="last N months")

    shares_q4 = compute_shares(
        act, months=last_q4, month_col="yyyy-mm",
        groupby_dims=child, total_by_dims=parent,
        excluded=["EXCLUDED FORMAT", "NEW FORMAT"], exclude_col="format_clean",
    ).assign(avg_type="last Q4")

    shares = pd.concat([shares_n, shares_q4], ignore_index=True)
    shares = inject_expected_shares(shares, new_formats)

    ign = actuals_df[actuals_df["forecast_product"].isin(ignore_formats["ignore_format"])]
    if not ign.empty:
        ign_agg = ign.groupby(
            ["destination_region", "forecast_product", "shoptype"], as_index=False
        ).agg({"actuals": "sum"})
        for at in ["last N months", "last Q4"]:
            shares = pd.concat([
                shares,
                ign_agg.assign(format_clean="na", frame="na", share=1.0, avg_type=at),
            ], ignore_index=True)

    shares = fill_mismatched_shares(shares, child)
    return shares


def classify_row(r):
    shop = (r.get("forecasted_shop") or "").strip()
    st = (r.get("shoptype") or "").strip()
    if not st or not shop:
        return "shop"
    if shop == st:
        return "shoptype"
    if shop == f"Other {st}":
        return "shoptype"
    return "shop"


def distribute_unknown_formats(combined_rows, actuals_df, shop_shares, st_shares):
    result = []
    fallback_count = 0
    unresolved_count = 0

    shop_idx = {}
    for _, row in shop_shares.iterrows():
        k = (
            str(row.get("forecasted_shop", "")).strip(),
            str(row.get("destination_region", "")).strip(),
            str(row.get("forecast_product", "")).strip(),
            str(row.get("avg_type", "")).strip(),
        )
        shop_idx.setdefault(k, []).append(row)

    st_idx = {}
    for _, row in st_shares.iterrows():
        k = (
            str(row.get("shoptype", "")).strip(),
            str(row.get("destination_region", "")).strip(),
            str(row.get("forecast_product", "")).strip(),
            str(row.get("avg_type", "")).strip(),
        )
        st_idx.setdefault(k, []).append(row)

    for r in combined_rows:
        fmt = (r.get("format") or "").strip()
        frm = (r.get("frame") or "").strip()

        if fmt != "Unknown" and frm != "Unknown":
            result.append(r)
            continue

        period = (r.get("period") or "").strip()
        avg_type = "last Q4" if is_q4_period(period) else "last N months"
        shop = (r.get("forecasted_shop") or "").strip()
        st = (r.get("shoptype") or "").strip()
        dest = (r.get("destination") or "").strip()
        product = (r.get("product") or "").strip()
        forecast = r.get("forecast", 0)
        try:
            forecast = float(forecast)
        except (ValueError, TypeError):
            forecast = 0.0

        row_type = classify_row(r)
        shares_rows = None

        if row_type == "shop":
            key = (shop, dest, product, avg_type)
            shares_rows = shop_idx.get(key)
            if not shares_rows and st:
                shares_rows = st_idx.get((st, dest, product, avg_type))
                if shares_rows:
                    fallback_count += 1
        else:
            key = (st, dest, product, avg_type)
            shares_rows = st_idx.get(key)

        if not shares_rows:
            r_copy = dict(r)
            r_copy["format"] = "na"
            r_copy["frame"] = "na"
            result.append(r_copy)
            unresolved_count += 1
            continue

        for sr in shares_rows:
            share = float(sr.get("share", 0))
            if share <= 0:
                continue
            result.append({
                "forecasted_shop": r.get("forecasted_shop", ""),
                "shoptype": r.get("shoptype", ""),
                "destination": r.get("destination", ""),
                "period": period,
                "product": product,
                "format": str(sr.get("format_clean", "na")).strip(),
                "frame": str(sr.get("frame", "na")).strip(),
                "forecast": share * forecast,
                "forecast_type": "Manual",
                "source_file": r.get("source_file", ""),
                "source_timestamp": r.get("source_timestamp", ""),
            })

    print(f"[format-dist] fallback to shoptype: {fallback_count}")
    print(f"[format-dist] unresolved (no history): {unresolved_count}")
    return result


# ═══════════════════════════════════════════════════════════════════
#  MONTHLY EXTRAPOLATION
# ═══════════════════════════════════════════════════════════════════


def generate_month_range(current_ym):
    dt = datetime.strptime(current_ym, "%Y-%m")
    start = datetime(dt.year, 1, 1)
    end = dt + relativedelta(months=11)
    months = []
    cursor = start
    while cursor <= end:
        months.append(cursor.strftime("%Y-%m"))
        cursor += relativedelta(months=1)
    return months


def extrapolate_missing_months(combined_rows, actuals_df, current_month, last_n_months):
    # source_file and source_timestamp are NOT part of the dedup/skeleton key —
    # including them causes ghost skeleton entries per source file, producing
    # duplicate Calculation rows for periods already covered by a newer source.
    grp_cols = ["forecasted_shop", "shoptype", "destination", "product", "format", "frame"]
    all_months = generate_month_range(current_month)

    fc = pd.DataFrame(combined_rows)
    for c in CSV_FIELDS:
        if c not in fc.columns:
            fc[c] = ""
    fc["forecast"] = pd.to_numeric(fc["forecast"], errors="coerce").fillna(0.0)
    fc["forecast_type"] = fc.get("forecast_type", "Manual")
    fc["forecast_type"] = fc["forecast_type"].fillna("Manual")

    # Deduplicate on business key before building skeleton.
    # Keep the row with the latest source_timestamp when there are conflicts.
    fc["source_timestamp_dt"] = pd.to_datetime(fc["source_timestamp"], errors="coerce")
    fc = (
        fc.sort_values("source_timestamp_dt", ascending=False)
        .drop_duplicates(subset=grp_cols + ["period"])
        .drop(columns=["source_timestamp_dt"])
    )

    combos = fc[grp_cols].drop_duplicates()
    skeleton = combos.assign(_key=1).merge(
        pd.DataFrame({"period": all_months, "_key": 1}), on="_key"
    ).drop(columns="_key")
    skeleton = skeleton[skeleton["period"] >= current_month].reset_index(drop=True)
    skeleton["forecast"] = 0.0

    merge_cols = grp_cols + ["period"]
    merged = skeleton.merge(
        fc[merge_cols + ["forecast"]],
        on=merge_cols, how="left", suffixes=("_default", ""),
    )
    merged["forecast"] = merged["forecast"].fillna(merged["forecast_default"])
    merged.drop(columns=["forecast_default"], inplace=True)

    fc_keys = set(tuple(row) for row in fc[merge_cols].values)
    merged["fc_available"] = merged[merge_cols].apply(
        lambda row: tuple(row) in fc_keys, axis=1
    )

    merged["period_ly"] = (
        pd.to_datetime(merged["period"] + "-01") - pd.DateOffset(years=1)
    ).dt.to_period("M").astype(str)

    _total_rows = len(merged)
    _fc_rows = int(merged["fc_available"].sum())
    _missing_rows = _total_rows - _fc_rows
    _unique_combos = merged[grp_cols].drop_duplicates().shape[0]
    _unique_periods = merged["period"].nunique()
    print(f"[extrapolate] skeleton: {_total_rows} rows | {_unique_combos} combos x {_unique_periods} periods | fc={_fc_rows} | missing={_missing_rows}")

    # ── LY baselines ────────────────────────────────────────────────────
    shop_ly = (
        actuals_df
        .groupby(["forecasted_shop", "shoptype", "destination_region",
                  "forecast_product", "yyyy-mm"], as_index=False)["actuals"]
        .sum()
        .rename(columns={"destination_region": "destination", "yyyy-mm": "period_ly",
                         "forecast_product": "product", "actuals": "shop_actuals_ly"})
    )

    regional_ly = (
        actuals_df
        .groupby(["shoptype", "destination_region",
                  "forecast_product", "yyyy-mm"], as_index=False)["actuals"]
        .sum()
        .rename(columns={"destination_region": "destination", "yyyy-mm": "period_ly",
                         "forecast_product": "product", "actuals": "regional_actuals_ly"})
    )

    covered_shops = (
        fc[
            ~fc["forecasted_shop"].str.startswith("Other ", na=False) &
            (fc["forecasted_shop"] != fc["shoptype"]) &
            fc["shoptype"].notna() &
            (fc["shoptype"] != "")
        ]
        [["shoptype", "destination", "product", "forecasted_shop"]]
        .drop_duplicates()
    )

    covered_ly = (
        actuals_df
        .rename(columns={"destination_region": "destination",
                         "forecast_product": "product",
                         "yyyy-mm": "period_ly"})
        .merge(covered_shops, on=["shoptype", "destination", "product", "forecasted_shop"], how="inner")
        .groupby(["shoptype", "destination", "product", "period_ly"], as_index=False)["actuals"]
        .sum()
        .rename(columns={"actuals": "covered_actuals_ly"})
    )

    other_ly = regional_ly.merge(
        covered_ly, on=["shoptype", "destination", "product", "period_ly"], how="left"
    )
    other_ly["covered_actuals_ly"] = other_ly["covered_actuals_ly"].fillna(0)
    other_ly["other_actuals_ly"] = (
        other_ly["regional_actuals_ly"] - other_ly["covered_actuals_ly"]
    ).clip(lower=0)
    other_ly = other_ly[["shoptype", "destination", "product", "period_ly", "other_actuals_ly"]]

    merged = merged.merge(
        shop_ly, on=["forecasted_shop", "shoptype", "destination", "product", "period_ly"], how="left"
    )
    merged = merged.merge(
        regional_ly, on=["shoptype", "destination", "product", "period_ly"], how="left"
    )
    merged = merged.merge(
        other_ly, on=["shoptype", "destination", "product", "period_ly"], how="left"
    )
    merged["shop_actuals_ly"] = merged["shop_actuals_ly"].fillna(0)
    merged["regional_actuals_ly"] = merged["regional_actuals_ly"].fillna(0)
    merged["other_actuals_ly"] = merged["other_actuals_ly"].fillna(0)

    is_other = (
        merged["forecasted_shop"].str.startswith("Other ", na=False) &
        (merged["forecasted_shop"] != merged["shoptype"])
    )

    use_shop_flag = (
        merged[~is_other]
        .groupby(["forecasted_shop", "shoptype", "destination", "product"], as_index=False)
        .agg(any_zero=("shop_actuals_ly", lambda x: (x <= 0).any()))
    )
    use_shop_flag["use_shop"] = ~use_shop_flag["any_zero"]
    use_shop_flag.drop(columns=["any_zero"], inplace=True)

    merged = merged.merge(
        use_shop_flag, on=["forecasted_shop", "shoptype", "destination", "product"], how="left"
    )
    merged["use_shop"] = merged["use_shop"].fillna(False)

    shop_used = use_shop_flag["use_shop"].sum()
    regional_used = (~use_shop_flag["use_shop"]).sum()
    print(f"[extrapolate] shop-level seasonality: {shop_used} combos, "
          f"regional fallback: {regional_used} combos")

    merged["chosen_ly"] = np.where(
        is_other,
        merged["other_actuals_ly"],
        np.where(merged["use_shop"], merged["shop_actuals_ly"], merged["regional_actuals_ly"])
    )

    has_fc = merged[merged["fc_available"]].copy()
    if has_fc.empty:
        print("[extrapolate] WARNING: no forecast data found, returning combined as-is")
        return combined_rows

    coarse_cols = ["forecasted_shop", "shoptype", "destination", "product"]
    regional_coarse_cols = ["shoptype", "destination", "product"]

    has_fc_agg = (
        has_fc
        .groupby(coarse_cols + ["period"], as_index=False)
        .agg(
            forecast=("forecast", "sum"),
            regional_actuals_ly=("regional_actuals_ly", "first"),
            shop_actuals_ly=("shop_actuals_ly", "first"),
            chosen_ly=("chosen_ly", "first"),
        )
    )

    regional_monthly = (
        has_fc_agg
        .groupby(regional_coarse_cols + ["period"], as_index=False)
        .agg(
            forecast_month=("forecast", "sum"),
            regional_ly_month=("regional_actuals_ly", "first"),
        )
    )
    regional_monthly["monthly_pct"] = np.where(
        regional_monthly["regional_ly_month"] > 0,
        regional_monthly["forecast_month"] / regional_monthly["regional_ly_month"] - 1,
        np.nan,
    )
    yoy_regional = (
        regional_monthly
        .groupby(regional_coarse_cols, as_index=False)["monthly_pct"]
        .mean()
        .rename(columns={"monthly_pct": "yoy_growth_regional"})
    )
    monthly_coarse = (
        has_fc_agg
        .groupby(coarse_cols + ["period"], as_index=False)
        .agg(
            forecast_month=("forecast", "sum"),
            chosen_ly_month=("chosen_ly", "first"),
        )
    )
    monthly_coarse["monthly_pct"] = np.where(
        monthly_coarse["chosen_ly_month"] > 0,
        monthly_coarse["forecast_month"] / monthly_coarse["chosen_ly_month"] - 1,
        np.nan,
    )
    yoy_coarse = (
        monthly_coarse
        .groupby(coarse_cols, as_index=False)["monthly_pct"]
        .mean()
        .rename(columns={"monthly_pct": "yoy_growth"})
    )

    last_n_set = set(last_n_months)
    shop_ln_actuals = (
        actuals_df[actuals_df["yyyy-mm"].isin(last_n_set)]
        .groupby(["forecasted_shop", "shoptype", "destination_region", "forecast_product", "yyyy-mm"],
                 as_index=False)["actuals"]
        .sum()
        .rename(columns={"destination_region": "destination", "forecast_product": "product"})
    )
    shop_ln_valid = (
        shop_ln_actuals
        .groupby(["forecasted_shop", "shoptype", "destination", "product"], as_index=False)
        .agg(months_with_actuals=("actuals", lambda x: (x > 0).sum()))
    )
    shop_ln_valid["use_shop_growth"] = shop_ln_valid["months_with_actuals"] >= len(last_n_set)

    yoy_coarse = yoy_coarse.merge(
        shop_ln_valid[["forecasted_shop", "shoptype", "destination", "product", "use_shop_growth"]],
        on=coarse_cols, how="left"
    )
    yoy_coarse["use_shop_growth"] = yoy_coarse["use_shop_growth"].fillna(False)

    other_mask = (
        yoy_coarse["forecasted_shop"].str.startswith("Other ", na=False) &
        (yoy_coarse["forecasted_shop"] != yoy_coarse["shoptype"])
    )
    yoy_coarse.loc[other_mask, "use_shop_growth"] = False

    yoy_coarse = yoy_coarse.merge(yoy_regional, on=regional_coarse_cols, how="left")
    yoy_coarse["yoy_growth"] = np.where(
        yoy_coarse["use_shop_growth"],
        yoy_coarse["yoy_growth"],
        yoy_coarse["yoy_growth_regional"],
    )

    fallback_growth = (~yoy_coarse["use_shop_growth"]).sum()
    print(f"[extrapolate] yoy_growth fallback to regional: {fallback_growth} combos")
    print(f"[extrapolate] yoy_growth computed for {len(yoy_coarse)} combos")

    yoy_coarse = yoy_coarse[coarse_cols + ["yoy_growth"]]

    manual_df = has_fc[grp_cols + ["period", "forecast"]].copy()
    manual_df["forecast_type"] = "Manual"

    missing_coarse = (
        merged[~merged["fc_available"]]
        .groupby(coarse_cols + ["period"], as_index=False)
        .agg(chosen_ly=("chosen_ly", "first"))
    )
    missing_coarse = missing_coarse.merge(yoy_coarse, on=coarse_cols, how="left")
    missing_coarse["extrap_total"] = (
        missing_coarse["chosen_ly"] * (1 + missing_coarse["yoy_growth"].fillna(0))
    )
    print(f"[extrapolate] missing_coarse: {len(missing_coarse)} rows | "
          f"no_growth_rate={int(missing_coarse['yoy_growth'].isna().sum())} | "
          f"zero_ly={int((missing_coarse['chosen_ly'] <= 0).sum())}")
    _sample = missing_coarse.groupby(["shoptype","destination","product"], as_index=False)["extrap_total"].sum()
    print(f"[extrapolate] extrap_total by shoptype/dest/product:")
    print(_sample.to_string(index=False))

    fmt_key_cols = ["format", "frame"]
    fmt_shares = (
        has_fc
        .groupby(coarse_cols + fmt_key_cols, as_index=False)["forecast"]
        .sum()
        .rename(columns={"forecast": "fc_total"})
    )
    fmt_totals = fmt_shares.groupby(coarse_cols)["fc_total"].transform("sum")
    fmt_shares["fmt_share"] = np.where(fmt_totals > 0, fmt_shares["fc_total"] / fmt_totals, 0.0)
    print(f"[extrapolate] fmt_shares: {len(fmt_shares)} format/frame combos across {fmt_shares[coarse_cols].drop_duplicates().shape[0]} coarse combos")
    _missing_fmt = missing_coarse[~missing_coarse[coarse_cols[0]].isin(fmt_shares[coarse_cols[0]])].shape[0]
    print(f"[extrapolate] coarse combos in missing with no fmt_share match: {_missing_fmt}")

    # source_lookup must come from fc (which retains source columns),
    # not has_fc (which is derived from the skeleton and lacks them).
    source_lookup = (
        fc
        .groupby(coarse_cols + ["source_file", "source_timestamp"], as_index=False)["forecast"]
        .sum()
        .sort_values("forecast", ascending=False)
        .drop_duplicates(subset=coarse_cols)
        [coarse_cols + ["source_file", "source_timestamp"]]
    )

    extrap_expanded = missing_coarse.merge(
        fmt_shares[coarse_cols + fmt_key_cols + ["fmt_share"]],
        on=coarse_cols, how="left",
    )
    extrap_expanded = extrap_expanded.merge(source_lookup, on=coarse_cols, how="left")
    extrap_expanded["forecast"] = extrap_expanded["extrap_total"] * extrap_expanded["fmt_share"].fillna(1.0)
    extrap_expanded["forecast_type"] = "Calculation"

    result = (
        pd.concat(
            [manual_df, extrap_expanded[grp_cols + ["period", "forecast", "forecast_type"]]],
            ignore_index=True,
        )
        .groupby(grp_cols + ["period", "forecast_type"], as_index=False)["forecast"]
        .sum()
    )

    # Re-attach source info to all rows (manual + extrapolated)
    result = result.merge(source_lookup, on=coarse_cols, how="left")
    result["source_file"] = result["source_file"].fillna("")
    result["source_timestamp"] = result["source_timestamp"].fillna("")

    filled = len(missing_coarse)
    print(f"[extrapolate] rows_before={len(combined_rows)}, rows_after={len(result)}, "
          f"coarse_periods_filled={filled}")

    return result.to_dict("records")


# ═══════════════════════════════════════════════════════════════════
#  LAMBDA HANDLER
# ═══════════════════════════════════════════════════════════════════


def lambda_handler(event, context):
    response_url = event.get("response_url", "")
    entity_name = event.get("entity_name") or None
    if "user_id" in event:
        s3_key, entity_name = download_slack_file_to_s3(
            event["user_id"], event["channel_id"], response_url,
        )
        if not s3_key:
            return {"ok": False, "error": "file download failed"}

    shop_to_shoptype, actuals_df = build_shop_to_shoptype_and_actuals(BUCKET, ACTUALS_KEY)

    current_month, ignore_formats, new_formats = read_forecast_config(BUCKET, INPUT_XLSX_KEY)
    last_n_months = get_last_n_months(current_month, 6)
    last_q4 = get_last_q4_months(current_month)
    print(f"[config] current_month={current_month}, last_n={last_n_months}, last_q4={last_q4}")

    shop_shares = compute_shop_level_shares(actuals_df, last_n_months, last_q4, ignore_formats)
    st_shares = compute_shoptype_level_shares(actuals_df, last_n_months, last_q4, ignore_formats, new_formats)
    print(f"[shares] shop_shares={len(shop_shares)}, st_shares={len(st_shares)}")

    all_items = list_submit_keys(BUCKET, PREFIX)
    shop_items = [(k, lm) for k, lm in all_items if SHOP_FILE_RE.match(k.rsplit("/", 1)[-1])]
    shoptype_items = [(k, lm) for k, lm in all_items if SHOPTYPE_FILE_RE.match(k.rsplit("/", 1)[-1])]
    region_items = [(k, lm) for k, lm in all_items if REGION_FILE_RE.match(k.rsplit("/", 1)[-1])]

    shop_final, shop_new, shop_existing = process_shop_files(shop_items, None)
    for r in shop_final:
        r["shoptype"] = shop_to_shoptype.get((r.get("forecasted_shop") or "").strip(), "")

    st_final, st_new, st_existing = process_shoptype_files(shoptype_items, None)
    for r in st_final:
        shop = (r.get("forecasted_shop") or "").strip()
        if not (r.get("shoptype") or "").strip():
            looked_up = shop_to_shoptype.get(shop, "")
            r["shoptype"] = looked_up if looked_up else shop

    region_final, region_new, region_existing = process_region_files(region_items, None)

    combined = build_combined(shop_final, st_final, region_final, current_month)

    for r in combined:
        r["source_file"] = r.pop("_source_file", "")
        lm = r.pop("_file_lm", None)
        r["source_timestamp"] = lm.isoformat() if lm and lm != EPOCH else ""

    combined = [r for r in combined if (r.get("period") or "") >= current_month]

    combined = distribute_unknown_formats(combined, actuals_df, shop_shares, st_shares)

    combined = extrapolate_missing_months(combined, actuals_df, current_month, last_n_months)

    unknown_remaining = sum(
        1 for r in combined
        if (r.get("format") or "").strip() in ("Unknown", "")
        or (r.get("frame") or "").strip() in ("Unknown", "")
    )
    print(f"[validate] total_rows={len(combined)}, unknown_remaining={unknown_remaining}")

    write_csv(BUCKET, OUT_KEY_COMBINED, combined)

    if response_url and entity_name:
        slack_reply(response_url, f"Added `{entity_name}` to forecast.")

    return {
        "ok": True,
        "shops": {
            "submit_files": len(shop_items),
            "new_rows": shop_new,
            "existing_rows": shop_existing,
            "total_rows": len(shop_final),
        },
        "shoptypes": {
            "submit_files": len(shoptype_items),
            "new_rows": st_new,
            "existing_rows": st_existing,
            "total_rows": len(st_final),
        },
        "regions": {
            "submit_files": len(region_items),
            "new_rows": region_new,
            "existing_rows": region_existing,
            "total_rows": len(region_final),
        },
        "combined": {
            "total_rows": len(combined),
            "unknown_remaining": unknown_remaining,
            "out_key": OUT_KEY_COMBINED,
        },
        "config": {
            "current_month": current_month,
            "last_n_months": last_n_months,
            "last_q4": last_q4,
        },
    }